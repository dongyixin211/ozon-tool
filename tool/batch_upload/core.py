from __future__ import annotations

import base64
import copy
import hashlib
import hmac
import json
import mimetypes
import re
import time
from dataclasses import dataclass
from email.utils import formatdate
from pathlib import Path
from typing import Callable
from urllib import error, parse, request

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


SUPPORTED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
OZON_BASE_URL = "https://api-seller.ozon.ru"
RESULT_FILE_NAME = "batch_upload_results.xlsx"
UPLOAD_RESULT_HEADERS = ["是否上传成功", "上传成功SKU", "Ozon task_id", "OSS 文件夹", "错误信息"]
DESCRIPTION_ATTRIBUTE_IDS = {4191}
DESCRIPTION_NAME_MARKERS = ("description", "описание", "简介", "描述")
RICH_JSON_HEADERS = ("json富文本内容", "json富内容", "富文本json", "rich_json", "富内容json")
RICH_CONTENT_MARKERS = ("rich", "富内容", "富文本", "json content", "widget")
IMAGE_URL_KEYS = {"image", "image_url", "images", "img", "picture", "pictures", "src", "url"}
IMAGE_URL_PATTERN = re.compile(r"https?://[^\s\"'<>\\]+", re.IGNORECASE)
MERGE_CARD_NAME_MARKERS = ("merge", "combine", "объедин", "合并")
MERGE_CARD_VALUE_PREFIXES = ("98aabba9-a9aa-49a9-bba8-aa8988aa989b_",)
MERGE_CARD_ATTRIBUTE_IDS = {8292, 9048}
IMPORT_ALLOWED_FIELDS = {
    "attributes",
    "category_id",
    "color_image",
    "complex_attributes",
    "currency_code",
    "depth",
    "description",
    "description_category_id",
    "dimension_unit",
    "height",
    "images",
    "name",
    "offer_id",
    "old_price",
    "pdf_list",
    "premium_price",
    "price",
    "primary_image",
    "type_id",
    "vat",
    "weight",
    "weight_unit",
    "width",
}


@dataclass
class BatchUploadConfig:
    portrait_root: Path
    excel_path: Path
    ozon_client_id: str
    ozon_api_key: str
    template_offer_id: str
    oss_access_key_id: str
    oss_access_key_secret: str
    oss_bucket: str
    oss_endpoint: str
    oss_public_domain: str
    max_items: int = 0
    template_product: dict | None = None
    upload_template_video: bool = False
    template_video_links: list[str] | None = None


def list_sku_folders(root: Path) -> list[Path]:
    return sorted([item for item in root.iterdir() if item.is_dir()])


def list_sku_images(folder: Path) -> list[Path]:
    images = sorted(
        [item for item in folder.iterdir() if item.is_file() and item.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS]
    )
    return sorted(images, key=lambda item: (0 if item.stem.endswith("_ai_portrait") else 1, item.name.lower()))


def build_oss_object_key(shop_id: str, sku: str, image_path: Path, image_index: int = 0) -> str:
    del image_index  # 保留参数以兼容旧调用；对象名使用原始文件名
    filename = _safe_oss_filename(image_path.name)
    return f"{build_oss_folder(shop_id, sku)}/{filename}"


def build_oss_folder(shop_id: str, sku: str) -> str:
    shop = str(shop_id or "").strip()
    if not shop:
        raise ValueError("店铺 ID 不能为空")
    clean_sku = re.sub(r"[\\/]+", "_", str(sku or "").strip()).strip() or "sku"
    return f"{shop}/{clean_sku}"


def _safe_oss_filename(filename: str) -> str:
    name = Path(filename).name.strip()
    if not name:
        return "image.jpg"
    return re.sub(r"[\\/]+", "_", name)


def read_excel_rows(path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {}

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        required = {"货号", "标题", "简介"}
        missing = sorted(required - set(headers))
        if missing:
            raise RuntimeError(f"Excel 缺少表头: {', '.join(missing)}")

        indexes = {name: headers.index(name) for name in required}
        rich_index = _optional_header_index(headers, RICH_JSON_HEADERS)
        result: dict[str, dict[str, str]] = {}
        for row in rows[1:]:
            sku = _cell_text(row, indexes["货号"])
            if not sku:
                continue
            entry = {
                "title": _cell_text(row, indexes["标题"]),
                "description": _cell_text(row, indexes["简介"]),
                "rich_json": _cell_text(row, rich_index) if rich_index >= 0 else "",
            }
            result[sku] = entry
        return result
    finally:
        workbook.close()


def write_results(path: Path, rows: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "批量上架结果"
    headers = ["货号", "标题", "图片数量", "状态", "上传成功SKU", "Ozon task_id", "OSS 文件夹", "错误信息"]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.get("sku", ""),
                row.get("title", ""),
                row.get("image_count", 0),
                row.get("status", ""),
                row.get("uploaded_sku", ""),
                row.get("task_id", ""),
                row.get("oss_folder", ""),
                row.get("error", ""),
            ]
        )
    widths = [24, 42, 12, 16, 24, 24, 42, 80]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_status_to_source_excel(path: Path, results: list[dict]) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        headers = [str(sheet.cell(1, column).value).strip() if sheet.cell(1, column).value else "" for column in range(1, sheet.max_column + 1)]
        for header in UPLOAD_RESULT_HEADERS:
            if header not in headers:
                sheet.cell(1, sheet.max_column + 1, header)
                headers.append(header)

        column_map = {header: headers.index(header) + 1 for header in headers if header}
        if "货号" not in column_map:
            raise RuntimeError("Excel 缺少表头: 货号")

        result_by_sku = {str(row.get("sku", "")).strip(): row for row in results}
        for row_index in range(2, sheet.max_row + 1):
            sku = str(sheet.cell(row_index, column_map["货号"]).value or "").strip()
            if not sku or sku not in result_by_sku:
                continue
            result = result_by_sku[sku]
            sheet.cell(row_index, column_map["是否上传成功"], "是" if result.get("status") == "已提交" else "否")
            sheet.cell(row_index, column_map["上传成功SKU"], result.get("uploaded_sku", ""))
            sheet.cell(row_index, column_map["Ozon task_id"], result.get("task_id", ""))
            sheet.cell(row_index, column_map["OSS 文件夹"], result.get("oss_folder", ""))
            sheet.cell(row_index, column_map["错误信息"], result.get("error", ""))
        workbook.save(path)
    finally:
        workbook.close()


def create_upload_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "上架填写"
    headers = ["货号", "标题", "简介", "json富文本内容", *UPLOAD_RESULT_HEADERS]
    sheet.append(headers)
    sheet.append(
        [
            "SKU001",
            "Женский платок квадратный с геометрическим принтом",
            "Легкий женский платок для повседневного образа. Подходит для прогулок, поездок и сочетания с разной одеждой.",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    sheet.append(
        [
            "SKU002",
            "Женский шарф мягкий однотонный для повседневного ношения",
            "Мягкий аксессуар для базового гардероба. Можно использовать как шарф, накидку или декоративный элемент образа.",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    widths = [20, 48, 82, 36, 18, 24, 24, 70]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in range(1, 201):
        sheet.row_dimensions[row].height = 38
    sheet.freeze_panes = "A2"

    guide = workbook.create_sheet("填写说明")
    guide.append(["字段", "说明"])
    guide_rows = [
        ["货号", "必须和 3:4 输出目录下的文件夹名完全一致，例如文件夹 SKU001 对应货号 SKU001。"],
        ["标题", "上传到 Ozon 的商品标题，建议使用俄文，不能为空。"],
        ["简介", "上传到 Ozon 的商品描述/简介，建议使用俄文，不能为空。"],
        [
            "json富文本内容",
            "可选。填写 Ozon 富内容 JSON（含 content 数组）；更新已上架商品时勾选「JSON富内容」。"
            "素材导出里的文案 JSON 若无 content 字段则不会写入富内容。",
        ],
        ["图片", "不用填在 Excel 里。程序会自动读取 3:4 输出目录中同名货号文件夹里的图片。"],
        ["结果列", "是否上传成功、上传成功SKU、Ozon task_id、错误信息由程序上传后自动写入。"],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 110
    guide.freeze_panes = "A2"

    for worksheet in (sheet, guide):
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _cell_text(row: tuple, index: int) -> str:
    if index < 0 or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _optional_header_index(headers: list[str], candidates: tuple[str, ...]) -> int:
    lowered = [header.lower() for header in headers]
    for candidate in candidates:
        candidate_lower = candidate.lower()
        if candidate_lower in lowered:
            return lowered.index(candidate_lower)
    return -1


def extract_listed_image_urls(product: dict) -> list[str]:
    urls: list[str] = []
    seen: set[str] = set()

    def add(url: object) -> None:
        text = str(url or "").strip()
        if text.startswith(("http://", "https://")) and text not in seen:
            seen.add(text)
            urls.append(text)

    primary = product.get("primary_image")
    if isinstance(primary, str):
        add(primary)
    elif isinstance(primary, list):
        for item in primary:
            add(item)

    images = product.get("images")
    if isinstance(images, list):
        for item in images:
            if isinstance(item, str):
                add(item)
            elif isinstance(item, dict):
                add(item.get("url") or item.get("file_name"))
    return urls


def is_ozon_rich_content_json(text: str) -> bool:
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return False
    return isinstance(data, dict) and isinstance(data.get("content"), list)


RICH_IMG_URL_KEYS = ("src", "srcMobile", "src_mobile", "url", "image_url", "imageUrl")


def _is_http_image_url(value: object) -> bool:
    text = str(value or "").strip()
    return text.startswith(("http://", "https://"))


def _assign_rich_img_dict(img: dict, url: str) -> None:
    for key in RICH_IMG_URL_KEYS:
        if key in img and _is_http_image_url(img.get(key)):
            img[key] = url


def _replace_rich_content_images(data: dict, image_urls: list[str]) -> dict:
    """按富内容展示顺序（blocks 等）依次替换为第 1、2、3… 张商品图；同一块的 src/srcMobile 用同一张。"""
    if not image_urls:
        return data
    content = data.get("content")
    if not isinstance(content, list):
        return data

    slot_index = 0

    def next_product_url() -> str:
        nonlocal slot_index
        url = image_urls[min(slot_index, len(image_urls) - 1)]
        slot_index += 1
        return url

    for widget in content:
        if not isinstance(widget, dict):
            continue
        blocks = widget.get("blocks")
        if isinstance(blocks, list):
            for block in blocks:
                if not isinstance(block, dict):
                    continue
                img = block.get("img")
                if isinstance(img, dict):
                    _assign_rich_img_dict(img, next_product_url())
            continue
        img = widget.get("img")
        if isinstance(img, dict):
            _assign_rich_img_dict(img, next_product_url())
            continue
        if _is_http_image_url(widget.get("src")):
            widget["src"] = next_product_url()
            continue
        images = widget.get("images")
        if isinstance(images, list):
            for entry in images:
                if isinstance(entry, dict):
                    if isinstance(entry.get("img"), dict):
                        _assign_rich_img_dict(entry["img"], next_product_url())
                    elif _is_http_image_url(entry.get("src")):
                        entry["src"] = next_product_url()

    return data


def _replace_rich_json_image_urls(text: str, image_urls: list[str]) -> str:
    if not image_urls or not text.strip():
        return text
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return text
    if not isinstance(data, dict) or not isinstance(data.get("content"), list):
        return text
    return json.dumps(_replace_rich_content_images(data, image_urls), ensure_ascii=False)


def _is_rich_content_attribute(attribute: dict) -> bool:
    attr_id = int(attribute.get("id") or 0)
    if attr_id in DESCRIPTION_ATTRIBUTE_IDS:
        return False
    name = " ".join(
        str(attribute.get(key) or "") for key in ("name", "attribute_name", "title")
    ).lower()
    if any(marker in name for marker in RICH_CONTENT_MARKERS):
        return True
    values = attribute.get("values")
    if not isinstance(values, list):
        return False
    for value in values:
        raw = value.get("value") if isinstance(value, dict) else value
        if isinstance(raw, str) and is_ozon_rich_content_json(raw):
            return True
    return False


def apply_rich_json_to_item(item: dict, rich_json_text: str, *, image_urls: list[str] | None = None) -> bool:
    text = str(rich_json_text or "").strip()
    if not text:
        return False

    if is_ozon_rich_content_json(text):
        rich_value = text
    else:
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            return False
        if isinstance(payload, dict) and isinstance(payload.get("content"), list):
            rich_value = json.dumps(payload, ensure_ascii=False)
        else:
            return False

    if image_urls:
        rich_value = _replace_rich_json_image_urls(rich_value, image_urls)

    replaced = False
    for field_name in ("attributes", "complex_attributes"):
        attributes = item.get(field_name)
        if not isinstance(attributes, list):
            continue
        for attribute in attributes:
            if not isinstance(attribute, dict):
                continue
            if not _is_rich_content_attribute(attribute):
                continue
            attribute["values"] = [{"value": rich_value}]
            replaced = True
    if replaced:
        return True

    item.setdefault("attributes", []).append(
        {
            "id": 0,
            "complex_id": 0,
            "values": [{"value": rich_value}],
        }
    )
    return True


class AliyunOssClient:
    def __init__(
        self,
        access_key_id: str,
        access_key_secret: str,
        bucket: str,
        endpoint: str,
        public_domain: str,
        timeout_seconds: int = 120,
    ):
        self.access_key_id = access_key_id.strip()
        self.access_key_secret = access_key_secret.strip()
        self.bucket = bucket.strip()
        self.endpoint = endpoint.strip().replace("https://", "").replace("http://", "").rstrip("/")
        self.public_domain = public_domain.strip().rstrip("/")
        self.timeout_seconds = timeout_seconds

    def public_url(self, object_key: str) -> str:
        encoded_key = "/".join(parse.quote(part) for part in object_key.split("/"))
        if self.public_domain:
            domain = self.public_domain
            if not domain.startswith(("http://", "https://")):
                domain = f"https://{domain}"
            return f"{domain}/{encoded_key}"
        return f"https://{self.bucket}.{self.endpoint}/{encoded_key}"

    def upload_file(self, source_path: Path, object_key: str) -> str:
        content_type = mimetypes.guess_type(source_path.name)[0] or "application/octet-stream"
        payload = source_path.read_bytes()
        return self.upload_bytes(payload, object_key, content_type)

    def upload_bytes(self, payload: bytes, object_key: str, content_type: str = "application/octet-stream") -> str:
        date_header = formatdate(timeval=None, localtime=False, usegmt=True)
        headers = {
            "Content-Type": content_type,
            "Date": date_header,
            "Host": f"{self.bucket}.{self.endpoint}",
            "x-oss-object-acl": "public-read",
        }
        canonical_resource = f"/{self.bucket}/{object_key}"
        canonical_oss_headers = "x-oss-object-acl:public-read\n"
        string_to_sign = f"PUT\n\n{content_type}\n{date_header}\n{canonical_oss_headers}{canonical_resource}"
        signature = base64.b64encode(
            hmac.new(
                self.access_key_secret.encode("utf-8"),
                string_to_sign.encode("utf-8"),
                hashlib.sha1,
            ).digest()
        ).decode("ascii")
        headers["Authorization"] = f"OSS {self.access_key_id}:{signature}"

        encoded_key = "/".join(parse.quote(part) for part in object_key.split("/"))
        url = f"https://{self.bucket}.{self.endpoint}/{encoded_key}"
        req = request.Request(url, data=payload, headers=headers, method="PUT")
        try:
            with request.urlopen(req, timeout=self.timeout_seconds) as response:
                response.read()
        except error.HTTPError as exc:
            details = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"OSS 上传失败 HTTP {exc.code}: {details}") from exc
        except error.URLError as exc:
            raise RuntimeError(f"OSS 上传失败: {exc.reason}") from exc
        return self.public_url(object_key)

    def test_upload(self) -> str:
        key = f"ozon-products/_connection_test_{int(time.time())}.txt"
        payload = b"ozon tool oss test"
        date_header = formatdate(timeval=None, localtime=False, usegmt=True)
        content_type = "text/plain"
        headers = {
            "Content-Type": content_type,
            "Date": date_header,
            "Host": f"{self.bucket}.{self.endpoint}",
            "x-oss-object-acl": "public-read",
        }
        canonical_resource = f"/{self.bucket}/{key}"
        string_to_sign = f"PUT\n\n{content_type}\n{date_header}\nx-oss-object-acl:public-read\n{canonical_resource}"
        signature = base64.b64encode(
            hmac.new(
                self.access_key_secret.encode("utf-8"),
                string_to_sign.encode("utf-8"),
                hashlib.sha1,
            ).digest()
        ).decode("ascii")
        headers["Authorization"] = f"OSS {self.access_key_id}:{signature}"
        url = f"https://{self.bucket}.{self.endpoint}/{key}"
        req = request.Request(url, data=payload, headers=headers, method="PUT")
        try:
            with request.urlopen(req, timeout=self.timeout_seconds) as response:
                response.read()
        except error.HTTPError as exc:
            details = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"OSS 测试失败 HTTP {exc.code}: {details}") from exc
        except error.URLError as exc:
            raise RuntimeError(f"OSS 测试失败: {exc.reason}") from exc
        return self.public_url(key)

    def delete_object(self, object_key: str) -> None:
        date_header = formatdate(timeval=None, localtime=False, usegmt=True)
        headers = {
            "Date": date_header,
            "Host": f"{self.bucket}.{self.endpoint}",
        }
        canonical_resource = f"/{self.bucket}/{object_key}"
        string_to_sign = f"DELETE\n\n\n{date_header}\n{canonical_resource}"
        signature = base64.b64encode(
            hmac.new(
                self.access_key_secret.encode("utf-8"),
                string_to_sign.encode("utf-8"),
                hashlib.sha1,
            ).digest()
        ).decode("ascii")
        headers["Authorization"] = f"OSS {self.access_key_id}:{signature}"

        encoded_key = "/".join(parse.quote(part) for part in object_key.split("/"))
        url = f"https://{self.bucket}.{self.endpoint}/{encoded_key}"
        req = request.Request(url, headers=headers, method="DELETE")
        try:
            with request.urlopen(req, timeout=self.timeout_seconds) as response:
                response.read()
        except error.HTTPError as exc:
            if exc.code == 404:
                return
            details = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"OSS 删除失败 HTTP {exc.code}: {details}") from exc
        except error.URLError as exc:
            raise RuntimeError(f"OSS 删除失败: {exc.reason}") from exc

    def delete_objects(self, object_keys: list[str]) -> None:
        for object_key in object_keys:
            if object_key:
                self.delete_object(object_key)


class OzonSellerClient:
    def __init__(self, client_id: str, api_key: str, timeout_seconds: int = 120):
        self.client_id = client_id.strip()
        self.api_key = api_key.strip()
        self.timeout_seconds = timeout_seconds

    def _request_json(self, endpoint: str, payload: dict) -> dict:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = request.Request(
            f"{OZON_BASE_URL}{endpoint}",
            data=body,
            method="POST",
            headers={
                "Client-Id": self.client_id,
                "Api-Key": self.api_key,
                "Content-Type": "application/json",
            },
        )
        try:
            with request.urlopen(req, timeout=self.timeout_seconds) as response:
                text = response.read().decode("utf-8")
                return json.loads(text) if text else {}
        except error.HTTPError as exc:
            details = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Ozon HTTP {exc.code}: {details}") from exc
        except error.URLError as exc:
            raise RuntimeError(f"Ozon 连接失败: {exc.reason}") from exc

    def test_connection(self) -> dict:
        return self._request_json("/v3/product/list", {"filter": {}, "limit": 1})

    def product_exists(self, offer_id: str) -> bool:
        info = self._request_json("/v3/product/info/list", {"offer_id": [offer_id]})
        return bool(_extract_items(info))

    def get_template_product(self, offer_id: str) -> dict:
        info = self._request_json("/v3/product/info/list", {"offer_id": [offer_id]})
        info_items = _extract_items(info)
        if not info_items:
            raise RuntimeError(f"未找到模板商品货号: {offer_id}")

        attributes = self._request_json(
            "/v4/product/info/attributes",
            {"filter": {"offer_id": [offer_id]}, "limit": 1, "sort_dir": "ASC"},
        )
        attribute_items = _extract_items(attributes)
        merged = {}
        merged.update(info_items[0])
        if attribute_items:
            merged.update(attribute_items[0])
        return merged

    def import_products(self, items: list[dict]) -> dict:
        return self._request_json("/v3/product/import", {"items": items})

    def get_import_info(self, task_id: int | str) -> dict:
        return self._request_json("/v1/product/import/info", {"task_id": int(task_id)})

    def list_warehouses(self, *, limit: int = 200, offset: int = 0) -> list[dict]:
        data = self._request_json("/v2/warehouse/list", {"limit": limit, "offset": offset})
        top_level = data.get("warehouses")
        if isinstance(top_level, list):
            return [item for item in top_level if isinstance(item, dict)]
        warehouses = _extract_items(data)
        if warehouses:
            return warehouses
        result = data.get("result")
        if isinstance(result, list):
            return [item for item in result if isinstance(item, dict)]
        if isinstance(result, dict):
            nested = result.get("warehouses")
            if isinstance(nested, list):
                return [item for item in nested if isinstance(item, dict)]
        return []

    def list_products(self, *, last_id: str = "", limit: int = 100, visibility: str = "ALL") -> dict:
        payload: dict = {"filter": {"visibility": visibility}, "last_id": last_id, "limit": limit}
        return self._request_json("/v3/product/list", payload)

    def product_info_stocks(
        self,
        *,
        cursor: str = "",
        limit: int = 100,
        visibility: str = "",
        offer_ids: list[str] | None = None,
        product_ids: list[int] | None = None,
    ) -> dict:
        filter_payload: dict = {}
        if visibility:
            filter_payload["visibility"] = visibility
        if offer_ids:
            filter_payload["offer_id"] = offer_ids
        if product_ids:
            filter_payload["product_id"] = [str(product_id) for product_id in product_ids]
        payload: dict = {"cursor": cursor, "filter": filter_payload, "limit": limit}
        return self._request_json("/v4/product/info/stocks", payload)

    def update_stocks(self, stocks: list[dict]) -> dict:
        return self._request_json("/v2/products/stocks", {"stocks": stocks})

    def generate_barcodes(self, product_ids: list[int]) -> dict:
        return self._request_json("/v1/barcode/generate", {"product_ids": [str(product_id) for product_id in product_ids]})

    def add_barcode(self, product_id: int, barcode: str) -> dict:
        return self._request_json("/v1/barcode/add", {"product_id": int(product_id), "barcode": barcode.strip()})


def _extract_items(data: dict) -> list[dict]:
    result = data.get("result")
    if isinstance(result, dict):
        items = result.get("items") or result.get("products")
        if isinstance(items, list):
            return [item for item in items if isinstance(item, dict)]
    if isinstance(result, list):
        return [item for item in result if isinstance(item, dict)]
    items = data.get("items")
    if isinstance(items, list):
        return [item for item in items if isinstance(item, dict)]
    return []


def build_import_item(
    template: dict,
    offer_id: str,
    title: str,
    description: str,
    image_urls: list[str],
    *,
    video_links: list[str] | None = None,
    use_template_video: bool = False,
) -> dict:
    item = {key: copy.deepcopy(value) for key, value in template.items() if key in IMPORT_ALLOWED_FIELDS}
    template_offer_id = str(template.get("offer_id") or "").strip()
    _replace_template_image_urls(item, image_urls)
    _replace_merge_card_attributes(item, offer_id, template_offer_id)
    item["offer_id"] = offer_id
    item["name"] = title
    item["primary_image"] = image_urls[0]
    item["images"] = image_urls
    if "description" in item:
        item["description"] = description
    _replace_description_attribute(item, description)
    if use_template_video or video_links:
        from batch_upload.video_ops import (
            apply_video_to_import_item,
            resolve_sku_video_links,
            strip_misplaced_video_from_item,
        )

        resolved_links = resolve_sku_video_links(
            offer_id,
            default_links=video_links or [],
            template_product=template,
        )
        apply_video_to_import_item(item, resolved_links, template)
        if not resolved_links:
            strip_misplaced_video_from_item(item)
    return _drop_empty_import_fields(item)


def _replace_merge_card_attributes(item: dict, offer_id: str, template_offer_id: str = "") -> None:
    for field_name in ("attributes", "complex_attributes"):
        attributes = item.get(field_name)
        if not isinstance(attributes, list):
            continue
        for attribute in attributes:
            if _is_merge_card_attribute(attribute, template_offer_id):
                _set_attribute_value(attribute, offer_id)
            elif _is_merge_card_named_attribute(attribute):
                _set_attribute_value(attribute, offer_id)
            elif _contains_merge_card_value(attribute):
                _replace_merge_card_value(attribute, offer_id)


def _is_merge_card_attribute(attribute: object, template_offer_id: str) -> bool:
    if not isinstance(attribute, dict):
        return False
    attr_id = attribute.get("id")
    if attr_id in MERGE_CARD_ATTRIBUTE_IDS:
        return True
    if template_offer_id and _attribute_values_equal(attribute, template_offer_id):
        return True
    return False


def _attribute_values_equal(attribute: dict, expected: str) -> bool:
    values = attribute.get("values")
    if not isinstance(values, list):
        return False
    for value in values:
        if isinstance(value, dict) and str(value.get("value") or "").strip() == expected:
            return True
    return False


def _is_merge_card_named_attribute(attribute: object) -> bool:
    if not isinstance(attribute, dict):
        return False
    names = " ".join(
        str(attribute.get(key) or "") for key in ("name", "attribute_name", "title")
    ).lower()
    return any(marker in names for marker in MERGE_CARD_NAME_MARKERS)


def _set_attribute_value(attribute: object, offer_id: str) -> None:
    if not isinstance(attribute, dict):
        return
    values = attribute.get("values")
    if not isinstance(values, list) or not values:
        attribute["values"] = [{"value": offer_id}]
        return
    for value in values:
        if isinstance(value, dict):
            value["value"] = offer_id
        else:
            _replace_merge_card_value(value, offer_id)


def _replace_merge_card_value(value: object, offer_id: str) -> object:
    if isinstance(value, dict):
        for key, child in list(value.items()):
            value[key] = _replace_merge_card_value(child, offer_id)
        return value
    if isinstance(value, list):
        for index, child in enumerate(value):
            value[index] = _replace_merge_card_value(child, offer_id)
        return value
    if isinstance(value, str):
        lowered = value.lower()
        if any(prefix in lowered for prefix in MERGE_CARD_VALUE_PREFIXES):
            return offer_id
    return value


def _contains_merge_card_value(value: object) -> bool:
    if isinstance(value, dict):
        return any(_contains_merge_card_value(child) for child in value.values())
    if isinstance(value, list):
        return any(_contains_merge_card_value(child) for child in value)
    if isinstance(value, str):
        lowered = value.lower()
        return any(prefix in lowered for prefix in MERGE_CARD_VALUE_PREFIXES)
    return False


def _replace_template_image_urls(value: object, image_urls: list[str]) -> object:
    next_index = 0

    def next_image_url(old_url: str) -> str:
        nonlocal next_index
        replacement = image_urls[min(next_index, len(image_urls) - 1)]
        next_index += 1
        return replacement

    def is_image_url(url: str, key: str | None = None) -> bool:
        cleaned = url.rstrip(".,;)]}")
        lowered = cleaned.lower()
        if re.search(r"\.(?:png|jpe?g|webp|bmp)(?:$|[?#])", lowered):
            return True
        return bool(key and key.lower() in IMAGE_URL_KEYS and lowered.startswith(("http://", "https://")))

    def replace_in_string(text: str, key: str | None = None) -> str:
        if not image_urls:
            return text
        stripped = text.strip()
        if stripped.startswith("{") and '"content"' in stripped:
            replaced = _replace_rich_json_image_urls(text, image_urls)
            if replaced != text:
                return replaced

        def replace_match(match: re.Match[str]) -> str:
            url = match.group(0)
            trailing = ""
            while url and url[-1] in ".,;)]}":
                trailing = url[-1] + trailing
                url = url[:-1]
            if not is_image_url(url, key):
                return match.group(0)
            return next_image_url(url) + trailing

        return IMAGE_URL_PATTERN.sub(replace_match, text)

    def walk(node: object, key: str | None = None) -> object:
        if isinstance(node, dict):
            img_node = node.get("img")
            if isinstance(img_node, dict) and any(
                _is_http_image_url(img_node.get(url_key)) for url_key in RICH_IMG_URL_KEYS
            ):
                _assign_rich_img_dict(img_node, next_image_url(""))
            for child_key, child_value in list(node.items()):
                if child_key == "img" and isinstance(child_value, dict):
                    continue
                node[child_key] = walk(child_value, str(child_key))
            return node
        if isinstance(node, list):
            for index, child_value in enumerate(node):
                node[index] = walk(child_value, key)
            return node
        if isinstance(node, str):
            return replace_in_string(node, key)
        return node

    return walk(value)


def _replace_description_attribute(item: dict, description: str) -> None:
    attributes = item.get("attributes")
    if not isinstance(attributes, list):
        return
    replaced = False
    for attribute in attributes:
        if not isinstance(attribute, dict):
            continue
        attr_id = attribute.get("id")
        name = str(attribute.get("name") or "").lower()
        if attr_id in DESCRIPTION_ATTRIBUTE_IDS or any(marker in name for marker in DESCRIPTION_NAME_MARKERS):
            attribute["values"] = [{"value": description}]
            replaced = True
    if not replaced and description:
        attributes.append({"id": 4191, "complex_id": 0, "values": [{"value": description}]})


def _drop_empty_import_fields(item: dict) -> dict:
    cleaned = {}
    for key, value in item.items():
        if value is None:
            continue
        if isinstance(value, str) and not value.strip() and key not in {"color_image"}:
            continue
        if isinstance(value, list) and not value:
            continue
        cleaned[key] = value
    return cleaned


class BatchUploadWorker:
    def __init__(self, config: BatchUploadConfig, logger: Callable[[str], None]):
        self.config = config
        self.logger = logger
        self._cancelled = False
        self.results: list[dict] = []

    def cancel(self) -> None:
        self._cancelled = True

    def run(self) -> Path:
        if not self.config.portrait_root.is_dir():
            raise RuntimeError("3:4 输出目录不存在。")
        if not self.config.excel_path.is_file():
            raise RuntimeError("Excel 文件不存在。")

        excel_rows = read_excel_rows(self.config.excel_path)
        upload_rows = list(excel_rows.items())
        if self.config.max_items > 0:
            upload_rows = upload_rows[: self.config.max_items]
        folder_names = {folder.name for folder in list_sku_folders(self.config.portrait_root)}
        matched = sum(1 for sku, _row in upload_rows if sku in folder_names)
        self.logger(f"批量上架: Excel 货号 {len(upload_rows)} 个，匹配图片文件夹 {matched} 个，未匹配 {len(upload_rows) - matched} 个")

        oss_client = AliyunOssClient(
            self.config.oss_access_key_id,
            self.config.oss_access_key_secret,
            self.config.oss_bucket,
            self.config.oss_endpoint,
            self.config.oss_public_domain,
        )
        ozon_client = OzonSellerClient(self.config.ozon_client_id, self.config.ozon_api_key)

        self.logger(f"批量上架: 正在读取模板商品 {self.config.template_offer_id}")
        template = (
            copy.deepcopy(self.config.template_product)
            if self.config.template_product
            else ozon_client.get_template_product(self.config.template_offer_id)
        )
        self.logger("批量上架: 模板商品读取完成")
        upload_video_links: list[str] = []
        per_offer_video_links: dict[str, list[str]] = {}
        if self.config.upload_template_video:
            from batch_upload.video_ops import (
                extract_video_links_from_product,
                parse_video_links_text,
                read_video_links_from_excel,
            )

            upload_video_links = parse_video_links_text("\n".join(self.config.template_video_links or []))
            if not upload_video_links:
                upload_video_links = extract_video_links_from_product(template)
            per_offer_video_links = read_video_links_from_excel(self.config.excel_path)
            if upload_video_links:
                self.logger(f"批量上架: 默认模板视频 {len(upload_video_links)} 条")
            elif per_offer_video_links:
                self.logger(f"批量上架: Excel 视频页签已配置 {len(per_offer_video_links)} 个货号")
            elif extract_video_links_from_product(template):
                self.logger("批量上架: 将使用模板商品中的视频属性")
            else:
                self.logger("批量上架: 已勾选使用模板视频，但未找到可用视频链接或模板视频属性")

        for index, (sku, row) in enumerate(upload_rows, start=1):
            if self._cancelled:
                self.logger("批量上架: 任务已取消")
                break
            sku_folder = self.config.portrait_root / sku
            self.logger(f"[{index}/{len(upload_rows)}] 准备上架货号 {sku}")
            title = row["title"]
            description = row["description"]
            if not title or not description:
                self._add_result(sku, title, 0, "跳过", "", "", "标题或简介为空")
                self.logger("  跳过: 标题或简介为空")
                continue

            if not sku_folder.is_dir():
                self._add_result(sku, title, 0, "跳过", "", "", "未找到同名图片目录")
                self.logger("  跳过: 未找到同名图片目录")
                continue

            images = list_sku_images(sku_folder)
            if not images:
                self._add_result(sku, title, 0, "跳过", "", "", "未找到图片")
                self.logger("  跳过: 未找到图片")
                continue

            try:
                self.logger("  正在检查 Ozon 是否已存在相同货号")
                if ozon_client.product_exists(sku):
                    self._add_result(sku, title, len(images), "已存在", "", "", "Ozon 已存在相同货号，跳过上传")
                    self.logger("  跳过: Ozon 已存在相同货号")
                    continue

                shop_id = self.config.ozon_client_id
                oss_folder = build_oss_folder(shop_id, sku)
                self.logger(f"  OSS 路径: {oss_folder}/")
                image_urls: list[str] = []
                for image_index, image_path in enumerate(images, start=1):
                    object_key = build_oss_object_key(shop_id, sku, image_path, image_index)
                    self.logger(f"  [{image_index}/{len(images)}] 上传 OSS: {object_key}")
                    image_url = oss_client.upload_file(image_path, object_key)
                    image_urls.append(image_url)
                    self.logger(f"    OSS URL: {image_url}")

                sku_video_links = upload_video_links
                if self.config.upload_template_video and per_offer_video_links.get(sku):
                    sku_video_links = parse_video_links_text(
                        "\n".join(per_offer_video_links.get(sku) or [])
                    )
                item = build_import_item(
                    template,
                    sku,
                    title,
                    description,
                    image_urls,
                    video_links=sku_video_links if self.config.upload_template_video else None,
                    use_template_video=self.config.upload_template_video,
                )
                if self.config.upload_template_video:
                    from batch_upload.video_ops import count_video_url_attributes

                    video_count = count_video_url_attributes(item)
                    self.logger(
                        f"  视频: {'已写入 attributes (21841)' if video_count else '未写入，请检查模板/Excel 视频配置'}"
                    )
                self.logger(f"  正在提交 Ozon，上架图片 {len(image_urls)} 张")
                response = ozon_client.import_products([item])
                task_id = _extract_task_id(response)
                self._add_result(sku, title, len(image_urls), "已提交", sku, str(task_id or ""), "", oss_folder)
                self.logger(f"  Ozon 已提交，task_id: {task_id or '[未返回]'}")
            except Exception as exc:  # noqa: BLE001
                self._add_result(sku, title, len(images), "失败", "", "", str(exc))
                self.logger(f"  失败: {exc}")

        result_path = self.config.portrait_root / RESULT_FILE_NAME
        write_results(result_path, self.results)
        write_status_to_source_excel(self.config.excel_path, self.results)
        self.logger(f"批量上架: 上传结果已写回 Excel {self.config.excel_path}")
        self.logger(f"批量上架: 结果表已保存 {result_path}")
        return result_path

    def _add_result(
        self,
        sku: str,
        title: str,
        image_count: int,
        status: str,
        uploaded_sku: str,
        task_id: str,
        message: str,
        oss_folder: str = "",
    ) -> None:
        self.results.append(
            {
                "sku": sku,
                "title": title,
                "image_count": image_count,
                "status": status,
                "uploaded_sku": uploaded_sku,
                "task_id": task_id,
                "oss_folder": oss_folder,
                "error": message,
            }
        )


def _extract_task_id(response: dict) -> int | str | None:
    result = response.get("result")
    if isinstance(result, dict):
        return result.get("task_id") or result.get("id")
    return response.get("task_id")
