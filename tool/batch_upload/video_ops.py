"""Ozon product video helpers."""
from __future__ import annotations

import copy
import re
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook

from batch_upload.core import OzonSellerClient, _drop_empty_import_fields, _extract_items

Logger = Callable[[str], None]
VIDEO_ATTR_ID_URL = 21841
VIDEO_ATTR_ID_NAME = 21837
VIDEO_COMPLEX_ID = 100001
VIDEO_ATTR_IDS = frozenset({VIDEO_ATTR_ID_URL, VIDEO_ATTR_ID_NAME})
MAX_VIDEO_LINKS = 5
VIDEO_URL_PATTERN = re.compile(r"https?://[^\s\"'<>\\]+", re.IGNORECASE)


def _is_video_attribute(attribute: object) -> bool:
    if not isinstance(attribute, dict):
        return False
    return int(attribute.get("id") or 0) in VIDEO_ATTR_IDS


def _attribute_lists(product: dict) -> Iterable[list]:
    for field_name in ("attributes", "complex_attributes"):
        values = product.get(field_name)
        if isinstance(values, list):
            yield values


def parse_video_links_text(text: str) -> list[str]:
    links: list[str] = []
    for line in str(text or "").splitlines():
        candidate = line.strip()
        if not candidate:
            continue
        match = VIDEO_URL_PATTERN.search(candidate)
        if match:
            links.append(match.group(0).rstrip(".,;)]}"))
    if not links and str(text or "").strip().startswith(("http://", "https://")):
        links.append(str(text).strip())
    unique: list[str] = []
    seen: set[str] = set()
    for link in links:
        if link not in seen:
            seen.add(link)
            unique.append(link)
    return unique[:MAX_VIDEO_LINKS]


def _links_from_attribute(attribute: dict) -> list[str]:
    links: list[str] = []
    if int(attribute.get("id") or 0) != VIDEO_ATTR_ID_URL:
        return links
    values = attribute.get("values")
    if not isinstance(values, list):
        return links
    for value in values:
        if isinstance(value, dict):
            url = str(value.get("value") or "").strip()
        else:
            url = str(value).strip()
        if url.startswith(("http://", "https://")):
            links.append(url)
    return links


def extract_video_links_from_product(product: dict) -> list[str]:
    attributes = product.get("attributes")
    if isinstance(attributes, list):
        links = []
        for attribute in attributes:
            links.extend(_links_from_attribute(attribute))
        parsed = parse_video_links_text("\n".join(links))
        if parsed:
            return parsed
    complex_attributes = product.get("complex_attributes")
    if isinstance(complex_attributes, list):
        links = []
        for attribute in complex_attributes:
            links.extend(_links_from_attribute(attribute))
        return parse_video_links_text("\n".join(links))
    return []


def extract_video_name_from_product(product: dict, fallback: str = "product_video") -> str:
    for field_name in ("attributes", "complex_attributes"):
        attributes = product.get(field_name)
        if not isinstance(attributes, list):
            continue
        for attribute in attributes:
            if not isinstance(attribute, dict):
                continue
            if int(attribute.get("id") or 0) != VIDEO_ATTR_ID_NAME:
                continue
            values = attribute.get("values")
            if isinstance(values, list) and values:
                first = values[0]
                if isinstance(first, dict):
                    name = str(first.get("value") or "").strip()
                else:
                    name = str(first).strip()
                if name:
                    return name
    offer_id = str(product.get("offer_id") or fallback).strip()
    return f"{offer_id}_video.mp4" if offer_id else "product_video.mp4"


def build_video_attributes(video_links: list[str], video_name: str = "") -> list[dict]:
    """Ozon 视频字段 21837/21841 应放在 attributes[]，每条带 complex_id=100001。"""
    links = parse_video_links_text("\n".join(video_links))
    if not links:
        return []
    name = video_name.strip() or "product_video.mp4"
    return [
        {
            "id": VIDEO_ATTR_ID_NAME,
            "complex_id": VIDEO_COMPLEX_ID,
            "values": [{"dictionary_value_id": 0, "value": name}],
        },
        {
            "id": VIDEO_ATTR_ID_URL,
            "complex_id": VIDEO_COMPLEX_ID,
            "values": [{"dictionary_value_id": 0, "value": link} for link in links],
        },
    ]


def build_video_complex_attributes(video_links: list[str], video_name: str = "") -> list[dict]:
    return build_video_attributes(video_links, video_name)


def _strip_video_attributes(attributes: list) -> list[dict]:
    kept: list[dict] = []
    for attribute in attributes:
        if not isinstance(attribute, dict):
            continue
        if _is_video_attribute(attribute):
            continue
        kept.append(attribute)
    return kept


def strip_misplaced_video_from_item(item: dict) -> None:
    """移除误写在 complex_attributes 中的视频字段（Ozon 会 skipped）。"""
    complex_attributes = item.get("complex_attributes")
    if isinstance(complex_attributes, list):
        item["complex_attributes"] = _strip_video_attributes(complex_attributes)


def _merge_video_attributes(item: dict, video_attributes: list[dict]) -> None:
    if not video_attributes:
        return
    strip_misplaced_video_from_item(item)
    existing = item.get("attributes")
    kept = _strip_video_attributes(existing) if isinstance(existing, list) else []
    item["attributes"] = kept + [copy.deepcopy(attribute) for attribute in video_attributes]


def copy_video_attributes_from_template(template_product: dict, offer_id: str) -> list[dict]:
    copied: list[dict] = []
    seen_ids: set[int] = set()

    def collect_from(container: list) -> None:
        for attribute in container:
            if not isinstance(attribute, dict):
                continue
            attr_id = int(attribute.get("id") or 0)
            if attr_id not in VIDEO_ATTR_IDS or attr_id in seen_ids:
                continue
            seen_ids.add(attr_id)
            new_attribute = copy.deepcopy(attribute)
            if attr_id == VIDEO_ATTR_ID_NAME:
                values = new_attribute.get("values")
                if isinstance(values, list) and values:
                    first = values[0]
                    if isinstance(first, dict):
                        first["value"] = extract_video_name_from_product(template_product, offer_id)
            if "complex_id" not in new_attribute or not new_attribute.get("complex_id"):
                new_attribute["complex_id"] = VIDEO_COMPLEX_ID
            copied.append(new_attribute)

    attributes = template_product.get("attributes")
    if isinstance(attributes, list):
        collect_from(attributes)
    if len(copied) < len(VIDEO_ATTR_IDS):
        complex_attributes = template_product.get("complex_attributes")
        if isinstance(complex_attributes, list):
            collect_from(complex_attributes)
    return copied


def copy_video_complex_attributes_from_template(template_product: dict, offer_id: str) -> list[dict]:
    return copy_video_attributes_from_template(template_product, offer_id)


def read_video_links_from_excel(path: Path) -> dict[str, list[str]]:
    if not path.is_file():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            lowered = sheet_name.strip().lower()
            if "video" not in lowered:
                continue
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            offer_index = _header_index(headers, ("货号", "offer_id", "offer id", "sku"))
            link_index = _header_index(headers, ("视频链接", "视频", "video", "video_url", "url", "链接"))
            if offer_index < 0:
                offer_index = 0
            if link_index < 0:
                link_index = 1 if len(headers) > 1 else 0
            mapping: dict[str, list[str]] = {}
            for row in rows[1:]:
                if not row:
                    continue
                offer_id = _cell_text(row, offer_index)
                if not offer_id:
                    continue
                raw_link = _cell_text(row, link_index)
                links = parse_video_links_text(raw_link)
                if not links and len(row) > link_index + 1:
                    extra = "\n".join(_cell_text(row, idx) for idx in range(link_index, len(row)))
                    links = parse_video_links_text(extra)
                if links:
                    mapping[offer_id] = links
            if mapping:
                return mapping
        return {}
    finally:
        workbook.close()


def _header_index(headers: list[str], candidates: tuple[str, ...]) -> int:
    lowered = [header.lower() for header in headers]
    for candidate in candidates:
        candidate_lower = candidate.lower()
        if candidate_lower in lowered:
            return lowered.index(candidate_lower)
    return -1


def _cell_text(row: tuple, index: int) -> str:
    if index < 0 or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def build_video_update_item(template_product: dict, offer_id: str, video_links: list[str]) -> dict:
    item: dict = {}
    for key in (
        "description_category_id",
        "category_id",
        "type_id",
        "currency_code",
        "vat",
        "dimension_unit",
        "weight_unit",
        "depth",
        "width",
        "height",
        "weight",
    ):
        value = template_product.get(key)
        if value is not None and value != "":
            item[key] = copy.deepcopy(value)
    item["offer_id"] = offer_id.strip()
    video_name = extract_video_name_from_product(template_product, offer_id)
    _merge_video_attributes(item, build_video_attributes(video_links, video_name))
    return _drop_empty_import_fields(item)


def resolve_template_product_for_video(
    client: OzonSellerClient,
    *,
    template_product: dict | None,
    template_offer_id: str,
) -> dict:
    if isinstance(template_product, dict) and template_product:
        return template_product
    offer_id = str(template_offer_id or "").strip()
    if not offer_id:
        raise RuntimeError("请先查询或选择带视频的商品模板。")
    return client.get_template_product(offer_id)


def update_product_videos(
    client: OzonSellerClient,
    offer_id: str,
    video_links: list[str],
    template_product: dict,
    *,
    logger: Logger | None = None,
) -> int | str | None:
    links = parse_video_links_text("\n".join(video_links))
    if not links:
        raise RuntimeError(f"货号 {offer_id} 没有可用的视频链接。")
    item = build_video_update_item(template_product, offer_id, links)
    response = client.import_products([item])
    task_id = _extract_task_id(response)
    if logger:
        logger(f"  货号 {offer_id} 视频更新已提交，task_id={task_id}")
    return task_id


def update_listed_products_videos(
    client: OzonSellerClient,
    offer_ids: list[str],
    video_links: list[str],
    template_product: dict,
    *,
    per_offer_links: dict[str, list[str]] | None = None,
    logger: Logger | None = None,
) -> tuple[int, int]:
    default_links = parse_video_links_text("\n".join(video_links))
    if not default_links and not per_offer_links:
        raise RuntimeError("请先填写模板视频链接，或在上架 Excel 的 Ozon.Video 页签中配置链接。")
    success = 0
    failed = 0
    for offer_id in offer_ids:
        offer_id = str(offer_id).strip()
        if not offer_id:
            continue
        links = (per_offer_links or {}).get(offer_id) or default_links
        try:
            update_product_videos(client, offer_id, links, template_product, logger=logger)
            success += 1
        except Exception as exc:  # noqa: BLE001
            failed += 1
            if logger:
                logger(f"  货号 {offer_id} 视频更新失败: {exc}")
    return success, failed


def _extract_task_id(response: dict) -> int | str | None:
    result = response.get("result")
    if isinstance(result, dict):
        task_id = result.get("task_id")
        if task_id is not None:
            return task_id
    task_id = response.get("task_id")
    return task_id if task_id is not None else None


def resolve_sku_video_links(
    offer_id: str,
    *,
    default_links: list[str] | None = None,
    per_offer_links: dict[str, list[str]] | None = None,
    template_product: dict | None = None,
) -> list[str]:
    sku = str(offer_id or "").strip()
    if per_offer_links and sku:
        links = parse_video_links_text("\n".join(per_offer_links.get(sku) or []))
        if links:
            return links
    links = parse_video_links_text("\n".join(default_links or []))
    if links:
        return links
    if isinstance(template_product, dict):
        return extract_video_links_from_product(template_product)
    return []


def count_video_url_attributes(item: dict) -> int:
    attributes = item.get("attributes")
    if not isinstance(attributes, list):
        return 0
    return sum(1 for attribute in attributes if int(attribute.get("id") or 0) == VIDEO_ATTR_ID_URL)


def apply_video_to_import_item(
    item: dict,
    video_links: list[str],
    template_product: dict | None = None,
) -> bool:
    offer_id = str(item.get("offer_id") or "product").strip()
    source = template_product if isinstance(template_product, dict) else item
    links = parse_video_links_text("\n".join(video_links))
    if links:
        video_name = extract_video_name_from_product(source, offer_id)
        video_attributes = build_video_attributes(links, video_name)
    elif isinstance(template_product, dict):
        video_attributes = copy_video_attributes_from_template(template_product, offer_id)
    else:
        return False
    if not video_attributes:
        return False
    _merge_video_attributes(item, video_attributes)
    return True


def merge_video_into_import_item(item: dict, video_links: list[str], template_product: dict | None = None) -> None:
    apply_video_to_import_item(item, video_links, template_product)
