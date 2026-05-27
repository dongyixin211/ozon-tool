import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PIL import Image

from batch_upload.core import (
    apply_rich_json_to_item,
    build_oss_folder,
    build_oss_object_key,
    build_import_item,
    create_upload_template,
    is_ozon_rich_content_json,
    list_sku_images,
    OzonSellerClient,
    read_excel_rows,
    write_results,
    write_status_to_source_excel,
    _replace_rich_json_image_urls,
)


class BatchUploadTests(unittest.TestCase):
    def test_build_oss_object_key_uses_shop_sku_and_filename(self) -> None:
        shop_id = "4481877"
        sku = "头巾货号001"
        key = build_oss_object_key(shop_id, sku, Path("蓝色主图_ai_portrait.png"), 1)

        self.assertEqual(key, f"{shop_id}/{sku}/蓝色主图_ai_portrait.png")
        self.assertTrue(key.startswith(build_oss_folder(shop_id, sku) + "/"))

    def test_build_oss_folder_creates_shop_and_sku_paths(self) -> None:
        self.assertEqual(build_oss_folder("4481877", "SKU001"), "4481877/SKU001")
        self.assertNotEqual(build_oss_folder("4481877", "SKU001"), build_oss_folder("4481877", "SKU002"))
        self.assertNotEqual(build_oss_folder("4481877", "SKU001"), build_oss_folder("9999999", "SKU001"))

    def test_product_exists_checks_offer_id(self) -> None:
        class FakeOzonClient(OzonSellerClient):
            def __init__(self) -> None:
                super().__init__("client", "key")
                self.payloads = []

            def _request_json(self, endpoint: str, payload: dict) -> dict:
                self.payloads.append((endpoint, payload))
                return {"result": {"items": [{"offer_id": "SKU001"}]}}

        client = FakeOzonClient()

        self.assertTrue(client.product_exists("SKU001"))
        self.assertEqual(client.payloads[0][0], "/v3/product/info/list")
        self.assertEqual(client.payloads[0][1]["offer_id"], ["SKU001"])
        self.assertNotIn("filter", client.payloads[0][1])

    def test_product_exists_false_when_not_found(self) -> None:
        class FakeOzonClient(OzonSellerClient):
            def __init__(self) -> None:
                super().__init__("client", "key")

            def _request_json(self, endpoint: str, payload: dict) -> dict:
                return {"result": {"items": []}}

        self.assertFalse(FakeOzonClient().product_exists("SKU001"))

    def test_read_excel_rows_with_chinese_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "upload.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["货号", "标题", "简介"])
            sheet.append(["SKU001", "标题 A", "简介 A"])
            workbook.save(path)

            rows = read_excel_rows(path)

            self.assertEqual(rows["SKU001"]["title"], "标题 A")
            self.assertEqual(rows["SKU001"]["description"], "简介 A")

    def test_list_sku_images_uploads_all_images_with_ai_portrait_first(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            Image.new("RGB", (20, 20), "red").save(root / "a.png")
            Image.new("RGB", (20, 20), "blue").save(root / "b_ai_portrait.png")

            images = list_sku_images(root)

            self.assertEqual([item.name for item in images], ["b_ai_portrait.png", "a.png"])

    def test_build_import_item_replaces_only_upload_fields(self) -> None:
        template = {
            "offer_id": "TEMPLATE",
            "name": "模板标题",
            "barcode": "OZN4180474029",
            "price": "99.00",
            "vat": "0.0",
            "images": ["old"],
            "primary_image": "old",
            "attributes": [{"id": 4191, "values": [{"value": "旧简介"}]}, {"id": 123, "values": [{"value": "保留"}]}],
            "read_only": "ignored",
        }

        item = build_import_item(template, "SKU001", "新标题", "新简介", ["https://example.com/a.png"])

        self.assertEqual(item["offer_id"], "SKU001")
        self.assertEqual(item["name"], "新标题")
        self.assertEqual(item["price"], "99.00")
        self.assertEqual(item["primary_image"], "https://example.com/a.png")
        self.assertEqual(item["images"], ["https://example.com/a.png"])
        self.assertNotIn("barcode", item)
        self.assertNotIn("read_only", item)
        self.assertEqual(item["attributes"][0]["values"][0]["value"], "新简介")
        self.assertEqual(item["attributes"][1]["values"][0]["value"], "保留")

    def test_build_import_item_replaces_template_rich_image_urls(self) -> None:
        template = {
            "offer_id": "TEMPLATE",
            "name": "Template",
            "attributes": [
                {
                    "id": 999,
                    "values": [
                        {
                            "value": (
                                '{"content":[{"src":"https://old.example.com/one.jpg"},'
                                '{"src":"https://old.example.com/two.png"}]}'
                            )
                        }
                    ],
                }
            ],
            "complex_attributes": [
                {
                    "values": [
                        {"image_url": "https://old.example.com/three.webp"},
                    ]
                }
            ],
        }

        item = build_import_item(
            template,
            "SKU001",
            "Title",
            "Description",
            ["https://oss.example.com/new_001.jpg", "https://oss.example.com/new_002.jpg"],
        )

        rich_value = item["attributes"][0]["values"][0]["value"]
        self.assertIn("https://oss.example.com/new_001.jpg", rich_value)
        self.assertIn("https://oss.example.com/new_002.jpg", rich_value)
        self.assertNotIn("https://old.example.com/one.jpg", rich_value)
        self.assertEqual(item["complex_attributes"][0]["values"][0]["image_url"], "https://oss.example.com/new_002.jpg")

    def test_build_import_item_replaces_repeated_rich_urls_in_image_order(self) -> None:
        template = {
            "offer_id": "TEMPLATE",
            "name": "Template",
            "attributes": [
                {
                    "id": 999,
                    "values": [
                        {
                            "value": (
                                '{"content":[{"src":"https://old.example.com/same.jpg"},'
                                '{"src":"https://old.example.com/same.jpg"},'
                                '{"src":"https://old.example.com/same.jpg"}]}'
                            )
                        }
                    ],
                }
            ],
        }

        item = build_import_item(
            template,
            "SKU001",
            "Title",
            "Description",
            [
                "https://oss.example.com/product_001.jpg",
                "https://oss.example.com/product_002.jpg",
                "https://oss.example.com/product_003.jpg",
            ],
        )

        rich_value = item["attributes"][0]["values"][0]["value"]
        self.assertLess(rich_value.index("product_001.jpg"), rich_value.index("product_002.jpg"))
        self.assertLess(rich_value.index("product_002.jpg"), rich_value.index("product_003.jpg"))

    def test_build_import_item_sets_merge_card_attribute_to_sku(self) -> None:
        template = {
            "offer_id": "TEMPLATE",
            "name": "Template",
            "attributes": [
                {"id": 111, "name": "Merge card", "values": [{"value": "98aabba9-a9aa-49a9-bba8-aa8988aa989b_TJdNAe7zdlMF"}]},
                {"id": 222, "name": "Color", "values": [{"value": "Red"}]},
            ],
            "complex_attributes": [
                {"id": 333, "values": [{"value": "98aabba9-a9aa-49a9-bba8-aa8988aa989b_TJdNAe7zdlMF"}]},
                {"id": 444, "values": [{"value": "Keep"}]},
            ],
        }

        item = build_import_item(template, "SKU001", "Title", "Description", ["https://example.com/a.png"])

        self.assertEqual([attribute["id"] for attribute in item["attributes"]], [111, 222, 4191])
        self.assertEqual(item["attributes"][0]["values"][0]["value"], "SKU001")
        self.assertEqual([attribute["id"] for attribute in item["complex_attributes"]], [333, 444])
        self.assertEqual(item["complex_attributes"][0]["values"][0]["value"], "SKU001")
        self.assertNotIn("98aabba9-a9aa-49a9-bba8-aa8988aa989b_TJdNAe7zdlMF", str(item))

    def test_build_import_item_sets_merge_card_attribute_9048_to_sku(self) -> None:
        template = {
            "offer_id": "NP000225",
            "name": "Template",
            "attributes": [
                {"id": 9048, "complex_id": 0, "values": [{"dictionary_value_id": 0, "value": "NP000225"}]},
                {"id": 85, "complex_id": 0, "values": [{"value": "Нет бренда"}]},
            ],
        }

        item = build_import_item(template, "NP000300", "Title", "Description", ["https://example.com/a.png"])

        merge_attr = next(attribute for attribute in item["attributes"] if attribute["id"] == 9048)
        self.assertEqual(merge_attr["values"][0]["value"], "NP000300")
        self.assertNotIn("NP000225", str(item))

    def test_write_results(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            write_results(
                path,
                [
                    {
                        "sku": "SKU001",
                        "title": "标题",
                        "image_count": 2,
                        "status": "已提交",
                        "uploaded_sku": "SKU001",
                        "task_id": "123",
                        "oss_folder": "4481877/SKU001",
                        "error": "",
                    }
                ],
            )

            workbook = load_workbook(path, read_only=True)
            try:
                sheet = workbook.active
                self.assertEqual(sheet["A1"].value, "货号")
                self.assertEqual(sheet["A2"].value, "SKU001")
                self.assertEqual(sheet["E2"].value, "SKU001")
                self.assertEqual(sheet["F2"].value, "123")
                self.assertEqual(sheet["G2"].value, "4481877/SKU001")
            finally:
                workbook.close()

    def test_write_status_to_source_excel(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "upload.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["货号", "标题", "简介"])
            sheet.append(["SKU001", "标题", "简介"])
            workbook.save(path)

            write_status_to_source_excel(
                path,
                [
                    {
                        "sku": "SKU001",
                        "status": "已提交",
                        "uploaded_sku": "SKU001",
                        "task_id": "456",
                        "oss_folder": "4481877/SKU001",
                        "error": "",
                    }
                ],
            )

            workbook = load_workbook(path, read_only=True)
            try:
                sheet = workbook.active
                headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
                self.assertIn("是否上传成功", headers)
                self.assertIn("上传成功SKU", headers)
                self.assertIn("OSS 文件夹", headers)
                self.assertEqual(sheet.cell(2, headers.index("是否上传成功") + 1).value, "是")
                self.assertEqual(sheet.cell(2, headers.index("上传成功SKU") + 1).value, "SKU001")
                self.assertEqual(sheet.cell(2, headers.index("OSS 文件夹") + 1).value, "4481877/SKU001")
            finally:
                workbook.close()

    def test_create_upload_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "template.xlsx"
            create_upload_template(path)

            rows = read_excel_rows(path)

            self.assertIn("SKU001", rows)
            workbook = load_workbook(path, read_only=True)
            try:
                sheet = workbook.active
                headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
                self.assertEqual(headers[:4], ["货号", "标题", "简介", "json富文本内容"])
                self.assertIn("是否上传成功", headers)
                self.assertIn("上传成功SKU", headers)
            finally:
                workbook.close()

    def test_read_excel_rows_reads_rich_json_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "rows.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["货号", "标题", "简介", "json富文本内容"])
            sheet.append(["SKU001", "Title", "Desc", '{"content":[{"widgetName":"text"}]}'])
            workbook.save(path)
            rows = read_excel_rows(path)
            self.assertEqual(rows["SKU001"]["rich_json"], '{"content":[{"widgetName":"text"}]}')

    def test_apply_rich_json_to_item(self) -> None:
        item = {
            "attributes": [
                {
                    "id": 999,
                    "name": "Rich content JSON",
                    "values": [{"value": '{"content":[{"src":"https://old.example.com/a.jpg"}]}'}],
                }
            ]
        }
        rich = '{"content":[{"src":"https://new.example.com/b.jpg"}]}'
        self.assertTrue(is_ozon_rich_content_json(rich))
        self.assertTrue(
            apply_rich_json_to_item(
                item,
                rich,
                image_urls=["https://new.example.com/b.jpg"],
            )
        )
        self.assertIn("new.example.com", item["attributes"][0]["values"][0]["value"])

    def test_rich_content_uses_product_images_in_block_order(self) -> None:
        rich = {
            "content": [
                {
                    "widgetName": "raShowcase",
                    "type": "chess",
                    "blocks": [
                        {
                            "img": {
                                "src": "https://cdn1.ozone.ru/old-1.jpg",
                                "srcMobile": "https://cdn1.ozone.ru/old-1-mobile.jpg",
                            }
                        },
                        {
                            "img": {
                                "src": "https://cdn1.ozone.ru/old-2.jpg",
                                "srcMobile": "https://cdn1.ozone.ru/old-2-mobile.jpg",
                            }
                        },
                        {
                            "img": {
                                "src": "https://cdn1.ozone.ru/old-3.jpg",
                                "srcMobile": "https://cdn1.ozone.ru/old-3-mobile.jpg",
                            }
                        },
                    ],
                }
            ],
            "version": 0.3,
        }
        product_urls = [
            "https://oss.example.com/p1.jpg",
            "https://oss.example.com/p2.jpg",
            "https://oss.example.com/p3.jpg",
        ]
        updated = json.loads(_replace_rich_json_image_urls(json.dumps(rich, ensure_ascii=False), product_urls))
        blocks = updated["content"][0]["blocks"]
        self.assertEqual(blocks[0]["img"]["src"], product_urls[0])
        self.assertEqual(blocks[0]["img"]["srcMobile"], product_urls[0])
        self.assertEqual(blocks[1]["img"]["src"], product_urls[1])
        self.assertEqual(blocks[1]["img"]["srcMobile"], product_urls[1])
        self.assertEqual(blocks[2]["img"]["src"], product_urls[2])
        self.assertEqual(blocks[2]["img"]["srcMobile"], product_urls[2])


if __name__ == "__main__":
    unittest.main()
