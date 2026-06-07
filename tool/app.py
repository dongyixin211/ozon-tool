from __future__ import annotations

import base64
import concurrent.futures
import json
import mimetypes
import os
import queue
import re
import sys
import threading
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable
from urllib import error, request

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageFilter, ImageOps

from batch_upload.core import (
    AliyunOssClient,
    BatchUploadConfig,
    BatchUploadWorker,
    OzonSellerClient,
    create_upload_template,
    read_excel_rows,
)
from batch_upload.inventory_ops import (
    ProductInventoryRow,
    WarehouseOption,
    deserialize_warehouse_options,
    fetch_no_barcode_products,
    fetch_zero_stock_products,
    generate_barcodes_for_products,
    load_warehouses,
    serialize_warehouse_options,
    update_product_stocks,
)
from batch_upload.product_update_ops import ListedProductUpdateConfig, ListedProductUpdateWorker
from batch_upload.video_ops import (
    MAX_VIDEO_LINKS,
    extract_video_links_from_product,
    parse_video_links_text,
    read_video_links_from_excel,
    resolve_template_product_for_video,
    update_listed_products_videos,
)
from local_scene_composer import (
    DEFAULT_SIZE_LABEL,
    LOCAL_SCENE_PRESETS,
    LocalSceneJobConfig,
    LocalSceneWorker,
)
from image_providers import (
    DEFAULT_IMAGE_PROVIDER,
    IMAGE_PROVIDER_PROFILES,
    create_image_client,
    list_image_provider_ids,
    migrate_provider_api_keys,
    provider_default_base_url,
    provider_default_model,
    provider_label,
    provider_models,
    provider_uses_async_tasks,
)
from text_providers import (
    DEFAULT_TEXT_PROVIDER,
    TEXT_PROVIDER_PROFILES,
    list_text_provider_ids,
    migrate_text_provider_api_keys,
    text_provider_default_base_url,
    text_provider_default_model,
    text_provider_label,
    text_provider_models,
)
from scene_generator import (
    DEFAULT_ASPECT_RATIO,
    DEFAULT_SCENE_PROMPT_TEMPLATE,
    SCENE_PRESETS,
    SceneGenerationWorker,
    SceneJobConfig,
)


SUPPORTED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "config.json"

DEFAULT_BASE_URL = "https://breakout.wenwen-ai.com/v1"
DEFAULT_IMAGE_MODEL = "gpt-image-2"
DEFAULT_TEXT_MODEL = "gpt-5-high"
DEFAULT_GENERATION_SIZE = "1024x1536"
DEFAULT_DOWNLOAD_RETRIES = 3
LOG_FLUSH_BATCH_SIZE = 250
LOG_MAX_LINES = 4000
DEFAULT_OSS_BUCKET = "dyx-ozon-images"
DEFAULT_OSS_ENDPOINT = "oss-cn-beijing.aliyuncs.com"
DEFAULT_OSS_PUBLIC_DOMAIN = "https://dyx-ozon-images.oss-cn-beijing.aliyuncs.com"

if sys.platform == "darwin":
    UI_FONT_FAMILY = "PingFang SC"
    MONO_FONT_FAMILY = "Menlo"
elif os.name == "nt":
    UI_FONT_FAMILY = "Microsoft YaHei UI"
    MONO_FONT_FAMILY = "Consolas"
else:
    UI_FONT_FAMILY = "Noto Sans CJK SC"
    MONO_FONT_FAMILY = "DejaVu Sans Mono"

DEFAULT_IMAGE_PROMPT = (
    "你是一名 Ozon / Wildberries 俄罗斯电商视觉设计师。我将提供一张商品原图作为参考。\n\n"
    "【必须遵守】\n"
    "1. 产品主体与参考图完全一致：形状、颜色、印花、结构、比例不可修改\n"
    "2. 画面中所有中文替换为俄文，符合俄罗斯本地电商表达习惯\n"
    "3. 白底或浅色高级影棚背景，画面干净、专业、适合详情附图\n"
    "4. 可优化排版与信息层级，不添加不存在的功能或参数\n"
    "5. 输出竖版构图，比例 3:4，适合后续裁切为 900×1200 或 1200×1600\n"
    "6. 一张原图只生成一张新图，直接输出图片\n\n"
    "货号：{sku}；文件夹：{folder_name}；原图：{image_names}"
)

DEFAULT_TITLE_PROMPT = (
    "你是 Ozon 电商标题优化助手。请根据商品图片和基础信息，"
    "生成一个适合 Ozon 平台的中文商品标题。"
    "要求：简洁清晰，突出商品核心属性和卖点，不要堆砌词。"
    "货号：{sku}；文件夹：{folder_name}；图片数量：{image_count}；文件名：{image_names}。"
)

DEFAULT_DESCRIPTION_PROMPT = (
    "你是 Ozon 电商卖点文案助手。请根据商品图片和基础信息，"
    "生成适合 Ozon 商品详情页的中文描述和 5 条卖点。"
    "要求：表达自然，突出功能、材质、使用场景或优势；如果图片信息不足，不要编造危险参数。"
    "货号：{sku}；文件夹：{folder_name}；图片数量：{image_count}；文件名：{image_names}。"
)
DEFAULT_TEMPLATE_NAME = "默认模板"


def now_stamp() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def read_config_payload() -> dict:
    if not CONFIG_PATH.exists():
        return {}
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    return data if isinstance(data, dict) else {}


def write_config_payload(data: dict) -> None:
    CONFIG_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def list_subfolders(root: Path) -> list[Path]:
    return sorted([item for item in root.iterdir() if item.is_dir()])


def list_images(folder: Path) -> list[Path]:
    return sorted(
        [item for item in folder.iterdir() if item.is_file() and item.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS]
    )


def make_relative_output(root: Path, base: Path, target_root: Path) -> Path:
    return target_root / root.relative_to(base)


def image_to_data_url(path: Path) -> str:
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def encode_multipart(fields: list[tuple[str, str]], files: list[tuple[str, str, bytes, str]]) -> tuple[bytes, str]:
    boundary = f"----CodexBoundary{uuid.uuid4().hex}"
    chunks: list[bytes] = []

    for name, value in fields:
        chunks.extend(
            [
                f"--{boundary}\r\n".encode("utf-8"),
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode("utf-8"),
                value.encode("utf-8"),
                b"\r\n",
            ]
        )

    for field_name, filename, payload, content_type in files:
        chunks.extend(
            [
                f"--{boundary}\r\n".encode("utf-8"),
                f'Content-Disposition: form-data; name="{field_name}"; filename="{filename}"\r\n'.encode("utf-8"),
                f"Content-Type: {content_type}\r\n\r\n".encode("utf-8"),
                payload,
                b"\r\n",
            ]
        )

    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    body = b"".join(chunks)
    return body, f"multipart/form-data; boundary={boundary}"


def humanize_api_error(raw_message: str) -> str:
    text = raw_message.strip()
    lowered = text.lower()

    code_match = re.search(r'"code"\s*:\s*"([^"]+)"', text)
    error_code = code_match.group(1) if code_match else ""

    if "billing hard limit has been reached" in lowered or error_code == "billing_hard_limit_reached":
        return "账户已触发账单硬限制，请检查网关余额或上游账户。"
    if "insufficient_quota" in lowered:
        return "账户额度不足，请检查网关余额或令牌额度。"
    if "invalid_api_key" in lowered or "incorrect api key" in lowered:
        return "API Key 无效，请检查配置。"
    if "model_not_found" in lowered:
        return "模型不可用，请确认模型名或网关权限。"
    if "rate_limit" in lowered or "too many requests" in lowered:
        return "请求过于频繁，触发了限流，请稍后再试。"
    if "authentication" in lowered or "unauthorized" in lowered:
        return "鉴权失败，请检查 API Key 和接口地址。"
    if "error code: 1010" in lowered or " 1010" in lowered:
        return "请求被 Cloudflare 拦截（1010），请更新程序后重试；若仍失败请检查图片接口地址是否为 https://api.apimart.ai/v1。"
    if "network error" in lowered:
        return "网络连接失败，当前无法访问接口服务。"
    return text


def normalize_ozon_client_id(raw_value: str) -> str:
    value = raw_value.strip()
    if not value:
        return ""
    if value.isascii() and re.fullmatch(r"\d+", value):
        return value
    digit_groups = re.findall(r"\d+", value)
    if digit_groups:
        return digit_groups[-1]
    return value


def sanitize_json_text(raw_text: str) -> str:
    text = raw_text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    return text.strip()


def parse_copy_payload(raw_text: str) -> dict:
    text = sanitize_json_text(raw_text)
    if not text:
        raise RuntimeError("文本模型返回了空内容，无法生成文案。")
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return {
            "title": "",
            "description": text,
            "bullets": [],
        }
    bullets = data.get("bullets") or data.get("selling_points") or []
    if not isinstance(bullets, list):
        bullets = [str(bullets)]
    return {
        "title": str(data.get("title", "")).strip(),
        "description": str(data.get("description", "")).strip(),
        "bullets": [str(item).strip() for item in bullets if str(item).strip()],
    }


def parse_title_payload(raw_text: str) -> str:
    text = sanitize_json_text(raw_text)
    if not text:
        raise RuntimeError("文本模型返回了空内容，无法生成标题。")
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            return str(data.get("title", "")).strip()
    except json.JSONDecodeError:
        pass
    return text.strip()

def apply_image_watermark(
    base_image: Image.Image,
    watermark_path: Path,
    *,
    scale_ratio: float = 0.18,
    margin: int = 36,
) -> Image.Image:
    with Image.open(watermark_path) as watermark_source:
        watermark = watermark_source.convert("RGBA")

    max_width = max(1, int(base_image.width * scale_ratio))
    if watermark.width > max_width:
        resize_ratio = max_width / watermark.width
        watermark = watermark.resize(
            (max_width, max(1, int(watermark.height * resize_ratio))),
            Image.Resampling.LANCZOS,
    )

    layer = Image.new("RGBA", base_image.size, (0, 0, 0, 0))
    x = max(0, base_image.width - watermark.width - margin)
    y = max(0, base_image.height - watermark.height - margin)
    layer.alpha_composite(watermark, (x, y))
    return Image.alpha_composite(base_image.convert("RGBA"), layer).convert("RGB")


def create_portrait_variant(
    source_path: Path,
    destination_path: Path,
    size: tuple[int, int] = (1200, 1600),
    watermark_path: Path | None = None,
) -> None:
    with Image.open(source_path) as original:
        base = original.convert("RGB")
        canvas = ImageOps.fit(base, size, method=Image.Resampling.LANCZOS)
        canvas = canvas.filter(ImageFilter.GaussianBlur(radius=22))
        canvas = ImageOps.colorize(ImageOps.grayscale(canvas), black="#1e1e1e", white="#f8f8f8")

        foreground = ImageOps.contain(base, size, method=Image.Resampling.LANCZOS)
        composed = canvas.copy()
        offset = ((size[0] - foreground.width) // 2, (size[1] - foreground.height) // 2)
        composed.paste(foreground, offset)
        if watermark_path and watermark_path.is_file():
            composed = apply_image_watermark(composed, watermark_path)
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        composed.save(destination_path, format="PNG")


def save_copy_files(destination_dir: Path, sku: str, payload: dict) -> tuple[Path, Path]:
    destination_dir.mkdir(parents=True, exist_ok=True)
    json_path = destination_dir / f"{sku}_content.json"
    txt_path = destination_dir / f"{sku}_content.txt"

    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        f"标题：{payload.get('title', '')}",
        "",
        "描述：",
        payload.get("description", ""),
        "",
        "卖点：",
    ]
    for index, item in enumerate(payload.get("bullets", []), start=1):
        lines.append(f"{index}. {item}")
    txt_path.write_text("\n".join(lines), encoding="utf-8")
    return json_path, txt_path


def choose_first_portrait_image(folder: Path) -> Path | None:
    if not folder.exists():
        return None
    return next(
        (
            item
            for item in sorted(folder.iterdir())
            if item.is_file() and item.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
        ),
        None,
    )


def export_copy_to_excel(destination_dir: Path, rows: list[dict]) -> Path:
    destination_dir.mkdir(parents=True, exist_ok=True)
    output_path = destination_dir / "ozon_content_export.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "文案汇总"
    headers = ["sku", "货号", "标题", "简介", "json富文本内容", "是否上传成功"]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.get("sku", ""),
                row.get("product_code", ""),
                row.get("title", ""),
                row.get("summary", ""),
                row.get("rich_json", ""),
                row.get("upload_status", ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border_side = Side(style="thin", color="D9E2F3")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [16, 18, 36, 42, 72, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    workbook.save(output_path)
    return output_path


class GatewayClient:
    def __init__(self, api_key: str, base_url: str, timeout_seconds: int = 180):
        self.api_key = api_key.strip()
        self.base_url = base_url.rstrip("/")
        self.timeout_seconds = timeout_seconds

    def _headers(self, content_type: str | None = None) -> dict[str, str]:
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json",
        }
        if content_type:
            headers["Content-Type"] = content_type
        return headers

    def _request_json(self, method: str, endpoint: str, payload: bytes | None = None, content_type: str | None = None) -> dict:
        req = request.Request(
            f"{self.base_url}{endpoint}",
            data=payload,
            headers=self._headers(content_type),
            method=method,
        )
        try:
            with request.urlopen(req, timeout=self.timeout_seconds) as response:
                return json.loads(response.read().decode("utf-8"))
        except error.HTTPError as exc:
            details = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"HTTP {exc.code}: {details}") from exc
        except error.URLError as exc:
            raise RuntimeError(f"Network error: {exc.reason}") from exc

    def list_models(self) -> dict:
        return self._request_json("GET", "/models")


class OpenAICompatibleClient(GatewayClient):
    def __init__(self, api_key: str, base_url: str, image_model: str, text_model: str, timeout_seconds: int = 180):
        super().__init__(api_key, base_url, timeout_seconds=timeout_seconds)
        self.image_model = image_model.strip()
        self.text_model = text_model.strip()

    def _generate_image_without_references(self, prompt: str, output_size: str, quality: str) -> dict:
        payload = json.dumps(
            {"model": self.image_model, "prompt": prompt, "size": output_size, "quality": quality}
        ).encode("utf-8")
        return self._request_json("POST", "/images/generations", payload, "application/json")

    def generate_image(self, prompt: str, output_size: str, quality: str, reference_images: Iterable[Path] | None = None) -> bytes:
        refs = list(reference_images or [])
        if refs:
            fields = [("model", self.image_model), ("prompt", prompt), ("size", output_size), ("quality", quality)]
            files = []
            for ref_path in refs[:1]:
                mime = mimetypes.guess_type(ref_path.name)[0] or "application/octet-stream"
                files.append(("image", ref_path.name, ref_path.read_bytes(), mime))
            payload, content_type = encode_multipart(fields, files)
            data = self._request_json("POST", "/images/edits", payload, content_type)
        else:
            data = self._generate_image_without_references(prompt, output_size, quality)
        return self._extract_image_bytes(data)

    def generate_image_with_fallback(
        self, prompt: str, output_size: str, quality: str, reference_images: Iterable[Path] | None = None
    ) -> tuple[bytes, str]:
        refs = list(reference_images or [])
        try:
            if refs:
                return self.generate_image(prompt, output_size, quality, reference_images=refs), "edit"
            return self.generate_image(prompt, output_size, quality, reference_images=[]), "generation"
        except Exception as exc:  # noqa: BLE001
            if not refs:
                raise
            fallback_prompt = (
                f"{prompt}\n补充要求：参考原图主体样式与构图，保持商品主体清晰，适合电商详情展示。"
            )
            note = f"参考图编辑失败，已回退为文生图。原因: {humanize_api_error(str(exc))}"
            return self.generate_image(fallback_prompt, output_size, quality, reference_images=[]), note

    def _download_image_from_url(self, image_url: str) -> bytes:
        attempts = [
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
                ),
                "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            },
            {
                **self._headers(),
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
                ),
                "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            },
        ]

        last_error: Exception | None = None
        for _ in range(DEFAULT_DOWNLOAD_RETRIES):
            for headers in attempts:
                req = request.Request(image_url, headers=headers)
                try:
                    with request.urlopen(req, timeout=self.timeout_seconds) as response:
                        return response.read()
                except error.HTTPError as exc:
                    details = exc.read().decode("utf-8", errors="replace")
                    last_error = RuntimeError(f"下载图片失败 HTTP {exc.code}: {details}")
                except error.URLError as exc:
                    last_error = RuntimeError(f"下载图片失败: {exc.reason}")
            time.sleep(1.2)

        assert last_error is not None
        raise last_error

    def _extract_image_bytes(self, data: dict) -> bytes:
        items = data.get("data")
        if not isinstance(items, list) or not items:
            raise RuntimeError(f"图片接口返回缺少 data 字段: {json.dumps(data, ensure_ascii=False)[:500]}")

        first = items[0]
        if not isinstance(first, dict):
            raise RuntimeError(f"图片接口返回结构异常: {json.dumps(data, ensure_ascii=False)[:500]}")

        image_base64 = first.get("b64_json")
        if image_base64:
            return base64.b64decode(image_base64)

        image_url = first.get("url")
        if image_url:
            return self._download_image_from_url(image_url)

        raise RuntimeError(f"图片接口返回里既没有 b64_json 也没有 url: {json.dumps(first, ensure_ascii=False)[:500]}")

    def _build_multimodal_messages(self, prompt: str, reference_images: Iterable[Path] | None = None) -> list[dict]:
        content_items: list[dict] = [{"type": "text", "text": prompt}]
        for ref_path in list(reference_images or [])[:1]:
            content_items.append({"type": "image_url", "image_url": {"url": image_to_data_url(ref_path)}})
        return [{"role": "user", "content": content_items}]

    def generate_title(self, prompt: str, reference_images: Iterable[Path] | None = None) -> str:
        payload = json.dumps(
            {
                "model": self.text_model,
                "messages": [
                    {"role": "system", "content": '你是资深 Ozon 电商标题编辑。请只返回 JSON：{"title":""}'},
                    *self._build_multimodal_messages(prompt, reference_images),
                ],
                "temperature": 0.7,
            }
        ).encode("utf-8")
        data = self._request_json("POST", "/chat/completions", payload, "application/json")
        raw_text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        try:
            return parse_title_payload(raw_text)
        except Exception as exc:  # noqa: BLE001
            preview = sanitize_json_text(str(raw_text))[:200]
            raise RuntimeError(f"标题解析失败: {exc}。返回内容预览: {preview or '[empty]'}") from exc

    def generate_copy(self, prompt: str, reference_images: Iterable[Path] | None = None) -> dict:
        payload = json.dumps(
            {
                "model": self.text_model,
                "messages": [
                    {
                        "role": "system",
                        "content": '你是资深 Ozon 电商文案编辑。请只返回 JSON：{"description":"","bullets":["","","","",""]}',
                    },
                    *self._build_multimodal_messages(prompt, reference_images),
                ],
                "temperature": 0.7,
            }
        ).encode("utf-8")
        data = self._request_json("POST", "/chat/completions", payload, "application/json")
        raw_text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        try:
            return parse_copy_payload(raw_text)
        except Exception as exc:  # noqa: BLE001
            preview = sanitize_json_text(str(raw_text))[:200]
            raise RuntimeError(f"文案解析失败: {exc}。返回内容预览: {preview or '[empty]'}") from exc


class TextOnlyClient(GatewayClient):
    def __init__(self, api_key: str, base_url: str, text_model: str, timeout_seconds: int = 180):
        super().__init__(api_key, base_url, timeout_seconds=timeout_seconds)
        self.text_model = text_model.strip()

    def _build_multimodal_messages(self, prompt: str, reference_images: Iterable[Path] | None = None) -> list[dict]:
        content_items: list[dict] = [{"type": "text", "text": prompt}]
        for ref_path in list(reference_images or [])[:1]:
            content_items.append({"type": "image_url", "image_url": {"url": image_to_data_url(ref_path)}})
        return [{"role": "user", "content": content_items}]

    def generate_title(self, prompt: str, reference_images: Iterable[Path] | None = None) -> str:
        payload = json.dumps(
            {
                "model": self.text_model,
                "messages": [
                    {"role": "system", "content": '你是资深 Ozon 电商标题编辑。请只返回 JSON：{"title":""}'},
                    *self._build_multimodal_messages(prompt, reference_images),
                ],
                "temperature": 0.7,
            }
        ).encode("utf-8")
        data = self._request_json("POST", "/chat/completions", payload, "application/json")
        raw_text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        try:
            return parse_title_payload(raw_text)
        except Exception as exc:  # noqa: BLE001
            preview = sanitize_json_text(str(raw_text))[:200]
            raise RuntimeError(f"标题解析失败: {exc}。返回内容预览: {preview or '[empty]'}") from exc

    def generate_copy(self, prompt: str, reference_images: Iterable[Path] | None = None) -> dict:
        payload = json.dumps(
            {
                "model": self.text_model,
                "messages": [
                    {
                        "role": "system",
                        "content": '你是资深 Ozon 电商文案编辑。请只返回 JSON：{"description":"","bullets":["","","","",""]}',
                    },
                    *self._build_multimodal_messages(prompt, reference_images),
                ],
                "temperature": 0.7,
            }
        ).encode("utf-8")
        data = self._request_json("POST", "/chat/completions", payload, "application/json")
        raw_text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        try:
            return parse_copy_payload(raw_text)
        except Exception as exc:  # noqa: BLE001
            preview = sanitize_json_text(str(raw_text))[:200]
            raise RuntimeError(f"文案解析失败: {exc}。返回内容预览: {preview or '[empty]'}") from exc


@dataclass
class JobConfig:
    source_root: Path
    generated_root: Path
    portrait_root: Path
    content_root: Path
    image_api_key: str
    text_api_key: str
    base_url: str
    image_base_url: str
    image_provider: str
    image_model: str
    text_provider: str
    text_model: str
    image_prompt_template: str
    title_prompt_template: str
    description_prompt_template: str
    watermark_path: Path | None
    quality: str
    max_folders: int
    max_workers: int
    convert_originals: bool
    generate_copy: bool
    export_excel: bool


class BatchWorker:
    def __init__(self, config: JobConfig, logger: Callable[[str], None]):
        self.config = config
        self.logger = logger
        self._cancelled = False
        self.export_rows: list[dict] = []

    def cancel(self) -> None:
        self._cancelled = True

    def _generate_single_image(
        self,
        client: object,
        sku: str,
        sku_folder: Path,
        images: list[Path],
        source_image: Path,
        generated_folder: Path,
        portrait_folder: Path,
    ) -> tuple[Path, Path, str]:
        self.logger(f"    开始处理图片: {source_image.name}")
        image_prompt = self.config.image_prompt_template.format(
            sku=sku,
            folder_name=sku_folder.name,
            image_count=len(images),
            image_names=source_image.name,
        )
        self.logger(f"      正在请求生图接口: {source_image.name}")
        generated_bytes, mode_note = client.generate_image_with_fallback(
            prompt=image_prompt,
            output_size=DEFAULT_GENERATION_SIZE,
            quality=self.config.quality,
            reference_images=[source_image],
        )
        generated_path = generated_folder / f"{source_image.stem}_ai_vertical.png"
        generated_path.write_bytes(generated_bytes)
        self.logger(f"      已保存竖版图: {generated_path.name}")
        portrait_path = portrait_folder / f"{source_image.stem}_ai_portrait.png"
        self.logger(f"      正在生成 3:4 图: {source_image.name}")
        create_portrait_variant(generated_path, portrait_path, watermark_path=self.config.watermark_path)
        self.logger(f"      已保存 3:4 图: {portrait_path.name}")
        return generated_path, portrait_path, mode_note

    def run(self) -> None:
        folders = list_subfolders(self.config.source_root)
        if self.config.max_folders > 0:
            folders = folders[: self.config.max_folders]
        if not folders:
            raise RuntimeError("源目录下没有找到任何子文件夹。")

        client = create_image_client(
            self.config.image_provider,
            self.config.image_api_key,
            self.config.image_base_url,
            self.config.image_model,
        )
        text_client = TextOnlyClient(
            api_key=self.config.text_api_key,
            base_url=self.config.base_url,
            text_model=self.config.text_model,
        )

        total = len(folders)
        for index, sku_folder in enumerate(folders, start=1):
            if self._cancelled:
                self.logger("任务已取消。")
                return

            sku = sku_folder.name
            images = list_images(sku_folder)
            image_names = ", ".join(item.name for item in images) if images else "无"

            self.logger(f"[{index}/{total}] 处理货号 {sku}")
            generated_folder = make_relative_output(sku_folder, self.config.source_root, self.config.generated_root)
            generated_folder.mkdir(parents=True, exist_ok=True)
            portrait_folder = make_relative_output(sku_folder, self.config.source_root, self.config.portrait_root)

            if images:
                worker_count = min(max(1, self.config.max_workers), len(images))
                self.logger(f"  检测到原图 {len(images)} 张，当前并发数 {worker_count}")
                with concurrent.futures.ThreadPoolExecutor(max_workers=worker_count) as executor:
                    future_map = {}
                    for queue_index, source_image in enumerate(images, start=1):
                        self.logger(f"    [{queue_index}/{len(images)}] 已加入队列: {source_image.name}")
                        future = executor.submit(
                            self._generate_single_image,
                            client,
                            sku,
                            sku_folder,
                            images,
                            source_image,
                            generated_folder,
                            portrait_folder,
                        )
                        future_map[future] = source_image

                    completed = 0
                    for future in concurrent.futures.as_completed(future_map):
                        source_image = future_map[future]
                        completed += 1
                        if self._cancelled:
                            executor.shutdown(wait=False, cancel_futures=True)
                            self.logger("任务已取消。")
                            return
                        try:
                            generated_path, portrait_path, mode_note = future.result()
                            self.logger(
                                f"    [{completed}/{len(images)}] 已完成 {source_image.name} -> {generated_path.name}, {portrait_path.name}"
                            )
                            if mode_note != "edit":
                                self.logger(f"      {mode_note}")
                        except Exception as exc:  # noqa: BLE001
                            self.logger(f"    [{completed}/{len(images)}] 图片 {source_image.name} 失败: {humanize_api_error(str(exc))}")
            else:
                self.logger("  未发现参考图，将直接文生图生成 1 张")
                image_prompt = self.config.image_prompt_template.format(
                    sku=sku,
                    folder_name=sku_folder.name,
                    image_count=0,
                    image_names=image_names,
                )
                generated_bytes = client.generate_image(
                    prompt=image_prompt,
                    output_size=DEFAULT_GENERATION_SIZE,
                    quality=self.config.quality,
                    reference_images=[],
                )
                generated_path = generated_folder / f"{sku}_ai_vertical.png"
                generated_path.write_bytes(generated_bytes)
                portrait_path = portrait_folder / f"{sku}_ai_portrait.png"
                create_portrait_variant(generated_path, portrait_path, watermark_path=self.config.watermark_path)
                self.logger(f"  已保存竖版图: {generated_path}")
                self.logger(f"  已保存 3:4 图: {portrait_path}")

            if self.config.generate_copy:
                content_folder = make_relative_output(sku_folder, self.config.source_root, self.config.content_root)
                reference_portrait = choose_first_portrait_image(portrait_folder)
                copy_refs = [reference_portrait] if reference_portrait else []
                self.logger("  开始生成文案")
                if reference_portrait:
                    self.logger(f"    文案参考图: {reference_portrait.name}")
                else:
                    self.logger("    文案参考图: 未找到，将仅基于提示词生成")

                title_prompt = self.config.title_prompt_template.format(
                    sku=sku,
                    folder_name=sku_folder.name,
                    image_count=len(images),
                    image_names=image_names,
                )
                description_prompt = self.config.description_prompt_template.format(
                    sku=sku,
                    folder_name=sku_folder.name,
                    image_count=len(images),
                    image_names=image_names,
                )
                try:
                    self.logger("    正在生成标题")
                    title_text = text_client.generate_title(title_prompt, reference_images=copy_refs)
                    self.logger("    标题生成完成，正在生成描述和卖点")
                    copy_payload = text_client.generate_copy(description_prompt, reference_images=copy_refs)
                    copy_payload["title"] = title_text
                    json_path, txt_path = save_copy_files(content_folder, sku, copy_payload)
                    self.logger(f"  已保存文案 JSON: {json_path}")
                    self.logger(f"  已保存文案文本: {txt_path}")
                    self.export_rows.append(
                        {
                            "sku": sku,
                            "product_code": sku_folder.name,
                            "title": copy_payload.get("title", ""),
                            "summary": copy_payload.get("description", ""),
                            "rich_json": json.dumps(copy_payload, ensure_ascii=False),
                            "upload_status": "",
                        }
                    )
                except Exception as exc:  # noqa: BLE001
                    self.logger(f"  文案生成失败: {humanize_api_error(str(exc))}")

        if self.config.generate_copy and self.config.export_excel and self.export_rows:
            excel_path = export_copy_to_excel(self.config.content_root, self.export_rows)
            self.logger(f"已导出 Excel 汇总表: {excel_path}")

        self.logger("全部处理完成。")


class App(tk.Tk):
    PAGE_DEFINITIONS = (
        ("generate", "素材生成", "素材生成", "配置目录、AI 接口并批量生成商品图与文案", "_build_generate_tab"),
        ("scene", "场景图", "场景图生成", "AI 生图或本地合成（本地不消耗 API Key）", "_build_scene_tab"),
        ("prompt", "提示词", "提示词管理", "管理图片、标题与详情页的提示词模板", "_build_prompt_tab"),
        ("upload", "批量上架", "批量上架", "多店铺配置、商品模板与 Ozon 批量提交", "_build_upload_tab"),
        ("inventory", "商品运维", "商品运维", "零库存补货与无条形码商品处理", "_build_inventory_tab"),
        ("log", "运行日志", "运行日志", "查看任务执行过程与错误信息", "_build_log_tab"),
    )

    def __init__(self) -> None:
        super().__init__()
        self.title("Ozon Tool · 商品素材与上架工作台")
        self.geometry("1320x880")
        self.minsize(1100, 720)

        self.log_queue: queue.Queue[str] = queue.Queue()
        self.log_line_count = 0
        self.worker: BatchWorker | None = None
        self.worker_thread: threading.Thread | None = None
        self.upload_worker: BatchUploadWorker | None = None
        self.upload_thread: threading.Thread | None = None

        self.source_root_var = tk.StringVar()
        self.generated_root_var = tk.StringVar()
        self.portrait_root_var = tk.StringVar()
        self.content_root_var = tk.StringVar()
        self.watermark_path_var = tk.StringVar()
        self.image_api_key_var = tk.StringVar(value=os.environ.get("BREAKOUT_IMAGE_API_KEY", os.environ.get("BREAKOUT_API_KEY", "")))
        self.text_api_key_var = tk.StringVar(value=os.environ.get("BREAKOUT_TEXT_API_KEY", os.environ.get("BREAKOUT_API_KEY", "")))
        self.base_url_var = tk.StringVar(value=DEFAULT_BASE_URL)
        self.image_base_url_var = tk.StringVar(value=DEFAULT_BASE_URL)
        self.image_provider_var = tk.StringVar(value=DEFAULT_IMAGE_PROVIDER)
        self.image_model_var = tk.StringVar(value=DEFAULT_IMAGE_MODEL)
        self.provider_api_keys: dict[str, str] = {}
        self._provider_switching = False
        self.text_provider_var = tk.StringVar(value=DEFAULT_TEXT_PROVIDER)
        self.text_provider_api_keys: dict[str, str] = {}
        self._text_provider_switching = False
        self.text_model_var = tk.StringVar(value=DEFAULT_TEXT_MODEL)
        self.quality_var = tk.StringVar(value="medium")
        self.max_folders_var = tk.StringVar(value="0")
        self.max_workers_var = tk.StringVar(value="3")
        self.convert_originals_var = tk.BooleanVar(value=False)
        self.generate_copy_var = tk.BooleanVar(value=True)
        self.export_excel_var = tk.BooleanVar(value=True)
        self.ozon_client_id_var = tk.StringVar()
        self.ozon_api_key_var = tk.StringVar()
        self.ozon_template_offer_id_var = tk.StringVar()
        self.upload_excel_path_var = tk.StringVar()
        self.oss_access_key_id_var = tk.StringVar(value="")
        self.oss_access_key_secret_var = tk.StringVar(value="")
        self.oss_bucket_var = tk.StringVar(value=DEFAULT_OSS_BUCKET)
        self.oss_endpoint_var = tk.StringVar(value=DEFAULT_OSS_ENDPOINT)
        self.oss_public_domain_var = tk.StringVar(value=DEFAULT_OSS_PUBLIC_DOMAIN)
        self.upload_max_items_var = tk.StringVar(value="0")
        self.template_name_var = tk.StringVar(value=DEFAULT_TEMPLATE_NAME)
        self.shop_name_var = tk.StringVar()
        self.selected_shop_var = tk.StringVar()
        self.product_template_name_var = tk.StringVar()
        self.selected_product_template_var = tk.StringVar()
        self.shop_profiles: dict[str, dict[str, str]] = {}
        self.product_templates: dict[str, dict] = {}
        self.shop_card_vars: dict[str, dict[str, tk.Variable]] = {}
        self.upload_workers: dict[str, BatchUploadWorker] = {}
        self.upload_threads: dict[str, threading.Thread] = {}
        self._video_update_running = False
        self._product_update_running = False
        self.update_listed_title_var = tk.BooleanVar(value=True)
        self.update_listed_description_var = tk.BooleanVar(value=True)
        self.update_listed_images_var = tk.BooleanVar(value=True)
        self.update_listed_video_var = tk.BooleanVar(value=False)
        self.update_listed_rich_json_var = tk.BooleanVar(value=True)
        self.template_query_source_shop_var = tk.StringVar()
        self.template_query_offer_id_var = tk.StringVar()
        self.template_query_save_name_var = tk.StringVar()
        self.template_query_video_links_var = tk.StringVar()
        self.template_query_video_status_var = tk.StringVar(value="模板视频 0 条")
        self.template_query_summary_var = tk.StringVar(value="尚未查询模板。")
        self.template_query_product: dict | None = None
        self.template_query_shop_id = ""
        self.template_query_offer_id = ""
        self.prompt_templates: dict[str, dict[str, str]] = {
            DEFAULT_TEMPLATE_NAME: {
                "image_prompt_template": DEFAULT_IMAGE_PROMPT,
                "title_prompt_template": DEFAULT_TITLE_PROMPT,
                "description_prompt_template": DEFAULT_DESCRIPTION_PROMPT,
            }
        }
        self.inventory_shop_var = tk.StringVar()
        self.inventory_warehouse_var = tk.StringVar()
        self.inventory_stock_qty_var = tk.StringVar(value="10")
        self.inventory_warehouses: list[WarehouseOption] = []
        self.inventory_warehouses_by_shop: dict[str, dict] = {}
        self.inventory_zero_stock_rows: dict[str, ProductInventoryRow] = {}
        self.inventory_no_barcode_rows: dict[str, ProductInventoryRow] = {}
        self.inventory_zero_checked: dict[str, bool] = {}
        self.inventory_no_barcode_checked: dict[str, bool] = {}
        self.scene_source_root_var = tk.StringVar()
        self.scene_single_image_var = tk.StringVar()
        self.scene_output_root_var = tk.StringVar()
        self.scene_aspect_ratio_var = tk.StringVar(value=DEFAULT_ASPECT_RATIO)
        self.scene_count_var = tk.StringVar(value="8")
        self.scene_max_workers_var = tk.StringVar(value="2")
        self.scene_max_folders_var = tk.StringVar(value="0")
        self.scene_worker: SceneGenerationWorker | None = None
        self.scene_thread: threading.Thread | None = None
        self.scene_mockup_root_var = tk.StringVar()
        self.scene_size_label_var = tk.StringVar(value=DEFAULT_SIZE_LABEL)
        self.local_scene_worker: LocalSceneWorker | None = None
        self.local_scene_thread: threading.Thread | None = None
        self._scene_prompt_template = DEFAULT_SCENE_PROMPT_TEMPLATE

        self._apply_theme()
        self._build_ui()
        self._load_config()
        self.after(250, self._flush_logs)

    def _apply_theme(self) -> None:
        self.configure(bg="#f0f2f5")
        style = ttk.Style(self)
        try:
            if sys.platform == "darwin":
                style.theme_use("aqua")
            else:
                style.theme_use("clam")
        except tk.TclError:
            pass

        font_ui = (UI_FONT_FAMILY, 10)
        page_bg = "#f0f2f5"
        card_bg = "#ffffff"
        border = "#f0f0f0"
        primary = "#1677ff"
        primary_hover = "#4096ff"
        text_primary = "#262626"
        text_secondary = "#8c8c8c"

        style.configure(".", font=font_ui, background=page_bg, foreground=text_primary)
        style.configure("App.TFrame", background=page_bg)
        style.configure("Sidebar.TFrame", background="#ffffff")
        style.configure("SidebarLogo.TLabel", background="#ffffff", foreground=text_primary, font=(UI_FONT_FAMILY, 15, "bold"))
        style.configure("SidebarHint.TLabel", background="#ffffff", foreground=text_secondary, font=(UI_FONT_FAMILY, 9))
        style.configure("Nav.TButton", padding=(14, 10), background="#ffffff", foreground=text_primary, borderwidth=0, anchor="w")
        style.map("Nav.TButton", background=[("active", "#f5f5f5")])
        style.configure(
            "NavActive.TButton",
            padding=(14, 10),
            background="#e6f4ff",
            foreground=primary,
            borderwidth=0,
            anchor="w",
            font=(UI_FONT_FAMILY, 10, "bold"),
        )
        style.map("NavActive.TButton", background=[("active", "#bae0ff")])
        style.configure("Topbar.TFrame", background="#ffffff")
        style.configure("PageTitle.TLabel", background="#ffffff", foreground=text_primary, font=(UI_FONT_FAMILY, 18, "bold"))
        style.configure("PageHint.TLabel", background="#ffffff", foreground=text_secondary, font=(UI_FONT_FAMILY, 10))
        style.configure("Card.TFrame", background=card_bg, relief="solid", borderwidth=1, bordercolor=border)
        style.configure("CardInner.TFrame", background=card_bg)
        style.configure("ShopCard.TFrame", background=card_bg, relief="solid", borderwidth=1, bordercolor=border)
        style.configure("ShopCardHeader.TFrame", background=card_bg)
        style.configure("ShopCardBody.TFrame", background=card_bg)
        style.configure("Section.TLabel", background=card_bg, foreground=text_primary, font=(UI_FONT_FAMILY, 13, "bold"))
        style.configure("Hint.TLabel", background=card_bg, foreground=text_secondary)
        style.configure("TLabel", background=page_bg, foreground=text_primary)
        style.configure("Card.TLabel", background=card_bg, foreground=text_primary)
        style.configure("ShopName.TLabel", background=card_bg, foreground=text_primary, font=(UI_FONT_FAMILY, 12, "bold"))
        style.configure("ShopMeta.TLabel", background=card_bg, foreground=text_secondary, font=(UI_FONT_FAMILY, 9))
        style.configure("Badge.TLabel", background="#722ed1", foreground="#ffffff", font=(UI_FONT_FAMILY, 9, "bold"), padding=(10, 4))
        style.configure("Tag.TLabel", background="#fafafa", foreground=text_secondary, padding=(10, 3))
        style.configure("TagSuccess.TLabel", background="#f6ffed", foreground="#52c41a", padding=(10, 3))
        style.configure("TagDanger.TLabel", background="#fff2f0", foreground="#ff4d4f", padding=(10, 3))
        style.configure("TEntry", fieldbackground="#ffffff", bordercolor="#d9d9d9", lightcolor="#d9d9d9", darkcolor="#d9d9d9", padding=8)
        style.configure("TCombobox", fieldbackground="#ffffff", bordercolor="#d9d9d9", padding=6)
        style.configure("TButton", padding=(14, 8), background="#ffffff", foreground=text_primary, bordercolor="#d9d9d9")
        style.map("TButton", background=[("active", "#fafafa")])
        style.configure("Primary.TButton", padding=(16, 9), background=primary, foreground="#ffffff", bordercolor=primary)
        style.map("Primary.TButton", background=[("active", primary_hover)], foreground=[("active", "#ffffff")])
        style.configure("Danger.TButton", padding=(14, 8), background="#fff2f0", foreground="#ff4d4f", bordercolor="#ffccc7")
        style.map("Danger.TButton", background=[("active", "#ffe7e6")])
        style.configure("TScrollbar", background="#f0f0f0", troughcolor="#fafafa", bordercolor=border, arrowsize=13)
        style.configure("TCheckbutton", background=card_bg, foreground=text_primary)
        style.map("TCheckbutton", background=[("active", card_bg)])

    def _show_page(self, page_key: str) -> None:
        page_meta = self.page_meta.get(page_key)
        if not page_meta:
            return
        title, hint = page_meta
        self.page_title_var.set(title)
        self.page_hint_var.set(hint)
        self._ensure_page_built(page_key)
        for key, frame in self.page_frames.items():
            if key == page_key:
                frame.grid(row=0, column=0, sticky="nsew")
            else:
                frame.grid_remove()
        for key, button in self.nav_buttons.items():
            button.configure(style="NavActive.TButton" if key == page_key else "Nav.TButton")
        self._active_page = page_key

    def _ensure_page_built(self, page_key: str) -> None:
        if page_key in self._built_pages:
            return
        frame = self.page_frames[page_key]
        builder = self.page_builders[page_key]
        previous_cursor = self.cget("cursor")
        self.configure(cursor="watch")
        self.update_idletasks()
        try:
            builder(frame)
            self._built_pages.add(page_key)
        finally:
            self.configure(cursor=previous_cursor)
        self.update_idletasks()

    def _build_ui(self) -> None:
        root = ttk.Frame(self, style="App.TFrame")
        root.pack(fill="both", expand=True)
        root.columnconfigure(2, weight=1)
        root.rowconfigure(0, weight=1)

        sidebar = ttk.Frame(root, style="Sidebar.TFrame", width=220)
        sidebar.grid(row=0, column=0, sticky="ns")
        tk.Frame(root, width=1, bg="#f0f0f0").grid(row=0, column=1, sticky="ns")
        sidebar.grid_propagate(False)
        sidebar.columnconfigure(0, weight=1)

        brand = ttk.Frame(sidebar, style="Sidebar.TFrame", padding=(20, 24, 20, 18))
        brand.pack(fill="x")
        ttk.Label(brand, text="Ozon Tool", style="SidebarLogo.TLabel").pack(anchor="w")
        ttk.Label(brand, text="商品素材与上架工作台", style="SidebarHint.TLabel").pack(anchor="w", pady=(6, 0))

        nav_wrap = ttk.Frame(sidebar, style="Sidebar.TFrame", padding=(12, 8))
        nav_wrap.pack(fill="both", expand=True)
        self.nav_buttons: dict[str, ttk.Button] = {}
        self._active_page = "generate"
        self.page_meta = {key: (title, hint) for key, _nav_label, title, hint, _builder_name in self.PAGE_DEFINITIONS}
        for page_key, label, _title, _hint, _builder_name in self.PAGE_DEFINITIONS:
            button = ttk.Button(
                nav_wrap,
                text=f"  {label}",
                style="Nav.TButton",
                command=lambda key=page_key: self._show_page(key),
            )
            button.pack(fill="x", pady=3)
            self.nav_buttons[page_key] = button

        main = ttk.Frame(root, style="App.TFrame")
        main.grid(row=0, column=2, sticky="nsew")
        main.columnconfigure(0, weight=1)
        main.rowconfigure(1, weight=1)

        topbar_wrap = ttk.Frame(main, style="Topbar.TFrame")
        topbar_wrap.grid(row=0, column=0, sticky="ew")
        topbar_wrap.columnconfigure(0, weight=1)
        topbar = ttk.Frame(topbar_wrap, style="Topbar.TFrame", padding=(28, 20, 28, 16))
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.columnconfigure(0, weight=1)
        self.page_title_var = tk.StringVar(value="素材生成")
        self.page_hint_var = tk.StringVar(value="")
        ttk.Label(topbar, textvariable=self.page_title_var, style="PageTitle.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(topbar, textvariable=self.page_hint_var, style="PageHint.TLabel").grid(row=1, column=0, sticky="w", pady=(6, 0))
        tk.Frame(topbar_wrap, height=1, bg="#f0f0f0").grid(row=1, column=0, sticky="ew")

        pages_host = ttk.Frame(main, style="App.TFrame", padding=(20, 0, 24, 20))
        pages_host.grid(row=1, column=0, sticky="nsew")
        pages_host.columnconfigure(0, weight=1)
        pages_host.rowconfigure(0, weight=1)

        self.page_frames: dict[str, ttk.Frame] = {}
        self.page_builders: dict[str, Callable[[ttk.Frame], None]] = {}
        self._built_pages: set[str] = set()
        for page_key, _nav_label, _title, _hint, builder_name in self.PAGE_DEFINITIONS:
            frame = ttk.Frame(pages_host, style="App.TFrame")
            self.page_frames[page_key] = frame
            self.page_builders[page_key] = getattr(self, builder_name)

        self._show_page("generate")

    def _create_scrollable_page(self, parent: ttk.Frame) -> tuple[ttk.Frame, tk.Canvas]:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        canvas = tk.Canvas(parent, bg="#f0f2f5", highlightthickness=0, borderwidth=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scrollbar.set)
        content = ttk.Frame(canvas, style="App.TFrame")
        content_window = canvas.create_window((0, 0), window=content, anchor="nw")
        content.columnconfigure(0, weight=1)

        def update_scroll_region(_event: tk.Event) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        def update_content_width(event: tk.Event) -> None:
            canvas.itemconfigure(content_window, width=event.width)

        content.bind("<Configure>", update_scroll_region)
        canvas.bind("<Configure>", update_content_width)
        self._bind_page_mousewheel(canvas, canvas)
        self._bind_page_mousewheel(scrollbar, canvas)
        return content, canvas

    def _bind_page_mousewheel(self, widget: tk.Misc, canvas: tk.Canvas) -> None:
        if isinstance(widget, (tk.Text, ttk.Treeview)):
            return

        def on_mousewheel(event: tk.Event) -> str:
            if event.delta:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        def on_mousewheel_linux_up(_event: tk.Event) -> str:
            canvas.yview_scroll(-1, "units")
            return "break"

        def on_mousewheel_linux_down(_event: tk.Event) -> str:
            canvas.yview_scroll(1, "units")
            return "break"

        widget.bind("<MouseWheel>", on_mousewheel)
        widget.bind("<Button-4>", on_mousewheel_linux_up)
        widget.bind("<Button-5>", on_mousewheel_linux_down)
        for child in widget.winfo_children():
            self._bind_page_mousewheel(child, canvas)

    def _build_generate_tab(self, parent: ttk.Frame) -> None:
        content, canvas = self._create_scrollable_page(parent)
        paths = self._section(content, "目录与输出", "选择原图、3:4 图片、文案和水印的保存位置", 0)
        paths.columnconfigure(1, weight=1)
        self._path_row(paths, 0, "源目录", self.source_root_var)
        self._path_row(paths, 1, "AI 竖版图输出目录", self.generated_root_var)
        self._path_row(paths, 2, "3:4 输出目录", self.portrait_root_var)
        self._path_row(paths, 3, "文案输出目录", self.content_root_var)
        self._path_row(paths, 4, "水印图片", self.watermark_path_var, select_file=True)
        ttk.Button(paths, text="单独转换 3:4 图片", command=self._start_portrait_only, style="Primary.TButton").grid(
            row=5, column=0, columnspan=3, sticky="w", pady=(10, 0)
        )

        api = self._section(content, "接口与模型", "图片与文案可使用不同接口；图片接口切换后模型与 Key 自动联动", 1)
        for col in (1, 3):
            api.columnconfigure(col, weight=1)
        provider_labels = [provider_label(pid) for pid in list_image_provider_ids()]
        self._image_provider_labels = provider_labels
        self._image_provider_ids = list_image_provider_ids()
        ttk.Label(api, text="图片接口", style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=8)
        self.image_provider_combobox = ttk.Combobox(
            api,
            values=provider_labels,
            state="readonly",
            width=28,
        )
        self.image_provider_combobox.grid(row=0, column=1, columnspan=3, sticky="ew", pady=8)
        self.image_provider_combobox.bind("<<ComboboxSelected>>", self._on_image_provider_selected)
        self._field(api, 1, 0, "图片接口地址", self.image_base_url_var, colspan=3)
        text_provider_labels = [text_provider_label(pid) for pid in list_text_provider_ids()]
        ttk.Label(api, text="文案接口", style="Card.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=8)
        self.text_provider_combobox = ttk.Combobox(
            api,
            values=text_provider_labels,
            state="readonly",
            width=28,
        )
        self.text_provider_combobox.grid(row=2, column=1, columnspan=3, sticky="ew", pady=8)
        self.text_provider_combobox.bind("<<ComboboxSelected>>", self._on_text_provider_selected)
        self._field(api, 3, 0, "文案接口地址", self.base_url_var, colspan=3)
        self._field(api, 4, 0, "图片 API Key", self.image_api_key_var, show="*", colspan=3)
        self._field(api, 5, 0, "文本 API Key", self.text_api_key_var, show="*", colspan=3)
        ttk.Label(api, text="图片模型", style="Card.TLabel").grid(row=6, column=0, sticky="w", padx=(0, 10), pady=8)
        self.image_model_combobox = ttk.Combobox(api, textvariable=self.image_model_var, width=34)
        self.image_model_combobox.grid(row=6, column=1, sticky="ew", pady=8)
        ttk.Label(api, text="文案模型", style="Card.TLabel").grid(row=6, column=2, sticky="w", padx=(0, 10), pady=8)
        self.text_model_combobox = ttk.Combobox(api, textvariable=self.text_model_var, width=28)
        self.text_model_combobox.grid(row=6, column=3, sticky="ew", pady=8)
        ttk.Label(api, text="质量", style="Card.TLabel").grid(row=7, column=0, sticky="w", padx=(0, 10), pady=8)
        ttk.Combobox(api, textvariable=self.quality_var, values=("low", "medium", "high"), state="readonly", width=16).grid(
            row=7, column=1, sticky="w", pady=8
        )
        self._field(api, 7, 2, "图片并发数", self.max_workers_var, width=10)
        self._field(api, 8, 0, "最多处理文件夹", self.max_folders_var, width=10)
        ttk.Checkbutton(api, text="同时生成标题、描述、卖点文案", variable=self.generate_copy_var).grid(
            row=8, column=2, columnspan=2, sticky="w", pady=8
        )
        ttk.Checkbutton(api, text="导出 Excel 汇总表", variable=self.export_excel_var).grid(
            row=9, column=2, columnspan=2, sticky="w", pady=(0, 8)
        )
        self._apply_image_provider_ui(self.image_provider_var.get(), initial=True)
        self._apply_text_provider_ui(self.text_provider_var.get(), initial=True)

        actions = self._section(content, "操作", "保存配置后执行素材生成任务", 2)
        ttk.Button(actions, text="保存配置", command=self._save_config).pack(side="left")
        ttk.Button(actions, text="测试 API", command=self._test_api_connection).pack(side="left", padx=10)
        ttk.Button(actions, text="开始处理", command=self._start_job, style="Primary.TButton").pack(side="left", padx=10)
        ttk.Button(actions, text="取消任务", command=self._cancel_job, style="Danger.TButton").pack(side="left")
        self._bind_page_mousewheel(content, canvas)

    def _build_scene_tab(self, parent: ttk.Frame) -> None:
        content, canvas = self._create_scrollable_page(parent)
        paths = self._section(content, "输入与输出", "平面印花图 + 内置 Ozon 拼图模板（左佩戴换印花 / 右平铺，不耗 API）", 0)
        paths.columnconfigure(1, weight=1)
        self._path_row(paths, 0, "平面原图（单张）", self.scene_single_image_var, select_file=True)
        self._path_row(paths, 1, "批量源目录", self.scene_source_root_var)
        self._path_row(paths, 2, "场景图输出目录", self.scene_output_root_var)
        self._path_row(paths, 3, "模特/场景底图目录", self.scene_mockup_root_var)
        ttk.Label(
            paths,
            text="推荐直接点「本地合成」：已内置你提供的参考构图（头巾侧戴/背面/蝴蝶结/尺寸）。自定义模板见 mockups/README.txt。",
            style="Hint.TLabel",
            wraplength=720,
        ).grid(row=4, column=0, columnspan=3, sticky="w", pady=(4, 0))

        opts = self._section(content, "生成参数", "场景数量与图片比例可在下方调整", 1)
        for col in (1, 3):
            opts.columnconfigure(col, weight=1)
        ttk.Label(opts, text="图片比例", style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=8)
        ttk.Combobox(
            opts,
            textvariable=self.scene_aspect_ratio_var,
            values=("1:1", "3:4", "4:3", "16:9", "1024x1024", "1024x1536", "1536x1024"),
            width=18,
        ).grid(row=0, column=1, sticky="w", pady=8)
        self._field(opts, 0, 2, "场景数量 (1-10)", self.scene_count_var, width=8)
        self._field(opts, 1, 0, "场景并发数", self.scene_max_workers_var, width=8)
        self._field(opts, 1, 2, "最多处理文件夹", self.scene_max_folders_var, width=8)
        self._field(opts, 2, 0, "尺寸标注文字", self.scene_size_label_var, colspan=3)
        ttk.Label(opts, text="AI 质量", style="Card.TLabel").grid(row=3, column=0, sticky="w", padx=(0, 10), pady=8)
        ttk.Combobox(opts, textvariable=self.quality_var, values=("low", "medium", "high"), state="readonly", width=16).grid(
            row=3, column=1, sticky="w", pady=8
        )

        local_lines = "、".join(f"{s['id']}" for s in LOCAL_SCENE_PRESETS[:6])
        ttk.Label(
            opts,
            text=f"本地模式场景示例：{local_lines}…（共 {len(LOCAL_SCENE_PRESETS)} 种，纯 Pillow 合成）",
            style="Hint.TLabel",
            wraplength=760,
        ).grid(row=4, column=0, columnspan=4, sticky="w", pady=(6, 0))

        prompt_box = self._section(content, "AI 场景图提示词（仅 AI 模式使用）", "可用 {aspect_ratio} {output_size} {scene_description} {sku} {image_name}", 2)
        prompt_box.columnconfigure(0, weight=1)
        prompt_box.rowconfigure(0, weight=1)
        self.scene_prompt_text = self._text_box(prompt_box, height=14)
        self.scene_prompt_text.insert("1.0", self._current_scene_prompt_template())
        self.scene_prompt_text.grid(row=0, column=0, sticky="nsew")

        actions = self._section(content, "操作", "推荐：本地合成不耗 Key；AI 模式需配置 API Key", 3)
        ttk.Button(actions, text="保存配置", command=self._save_config).pack(side="left")
        ttk.Button(actions, text="本地合成（不耗 API）", command=self._start_local_scene_job, style="Primary.TButton").pack(
            side="left", padx=10
        )
        ttk.Button(actions, text="AI 生成场景图", command=self._start_scene_job).pack(side="left", padx=6)
        ttk.Button(actions, text="取消", command=self._cancel_scene_jobs, style="Danger.TButton").pack(side="left")
        self._bind_page_mousewheel(content, canvas)

    def _build_prompt_tab(self, parent: ttk.Frame) -> None:
        content, canvas = self._create_scrollable_page(parent)
        bar = self._section(content, "提示词模板", "管理图片、标题和详情文案的提示词", 0)
        bar.columnconfigure(1, weight=1)
        ttk.Label(bar, text="模板名称", style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=8)
        self.template_combobox = ttk.Combobox(bar, textvariable=self.template_name_var, state="normal", width=30)
        self.template_combobox.grid(row=0, column=1, sticky="ew", pady=8)
        ttk.Button(bar, text="加载模板", command=self._load_selected_template).grid(row=0, column=2, padx=(10, 0), pady=8)
        ttk.Button(bar, text="保存为模板", command=self._save_prompt_template).grid(row=0, column=3, padx=(8, 0), pady=8)
        ttk.Button(bar, text="删除模板", command=self._delete_prompt_template, style="Danger.TButton").grid(
            row=0, column=4, padx=(8, 0), pady=8
        )
        self.template_combobox.bind("<<ComboboxSelected>>", lambda _event: self._load_selected_template())
        self._refresh_template_options()

        editor = self._section(content, "提示词内容", "支持变量：{sku}、{folder_name}、{image_count}、{image_names}", 1)
        editor.columnconfigure(0, weight=1)
        editor.columnconfigure(1, weight=1)

        ttk.Label(editor, text="图片提示词模板", style="Card.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 6))
        self.image_prompt_text = self._text_box(editor, height=9)
        self.image_prompt_text.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 12))
        self.image_prompt_text.insert("1.0", DEFAULT_IMAGE_PROMPT)

        ttk.Label(editor, text="标题提示词模板", style="Card.TLabel").grid(row=2, column=0, sticky="w", pady=(0, 6))
        ttk.Label(editor, text="卖点文案提示词模板", style="Card.TLabel").grid(row=2, column=1, sticky="w", padx=(10, 0), pady=(0, 6))
        self.title_prompt_text = self._text_box(editor, height=10)
        self.title_prompt_text.grid(row=3, column=0, sticky="nsew", pady=(0, 0), padx=(0, 10))
        self.title_prompt_text.insert("1.0", DEFAULT_TITLE_PROMPT)
        self.description_prompt_text = self._text_box(editor, height=10)
        self.description_prompt_text.grid(row=3, column=1, sticky="nsew", pady=(0, 0), padx=(10, 0))
        self.description_prompt_text.insert("1.0", DEFAULT_DESCRIPTION_PROMPT)
        selected_name = self.template_name_var.get().strip() or DEFAULT_TEMPLATE_NAME
        if selected_name in self.prompt_templates:
            self._apply_prompt_bundle(self.prompt_templates[selected_name])
        self._bind_page_mousewheel(content, canvas)

    def _build_log_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        card = ttk.Frame(parent, style="Card.TFrame", padding=16)
        card.grid(row=0, column=0, sticky="nsew")
        card.columnconfigure(0, weight=1)
        card.rowconfigure(1, weight=1)
        ttk.Label(card, text="运行日志", style="Section.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 12))
        self.log_text = tk.Text(
            card,
            height=24,
            wrap="word",
            state="disabled",
            bg="#fafafa",
            fg="#262626",
            insertbackground="#1677ff",
            relief="flat",
            highlightthickness=1,
            highlightbackground="#f0f0f0",
            highlightcolor="#1677ff",
            padx=14,
            pady=12,
            font=(MONO_FONT_FAMILY, 10),
        )
        self.log_text.grid(row=1, column=0, sticky="nsew")

    def _build_inventory_tab(self, parent: ttk.Frame) -> None:
        content, canvas = self._create_scrollable_page(parent)

        shop_bar = self._section(content, "店铺连接", "选择已保存的店铺，使用其 Client-Id 与 Api-Key 调用 Ozon API", 0)
        shop_bar.columnconfigure(1, weight=1)
        ttk.Label(shop_bar, text="选择店铺", style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=8)
        self.inventory_shop_combobox = ttk.Combobox(shop_bar, textvariable=self.inventory_shop_var, state="readonly", width=36)
        self.inventory_shop_combobox.grid(row=0, column=1, sticky="ew", pady=8)
        self.inventory_shop_combobox.bind(
            "<<ComboboxSelected>>", lambda _event: self._on_inventory_shop_changed()
        )
        ttk.Button(shop_bar, text="刷新店铺列表", command=self._refresh_inventory_shop_options).grid(row=0, column=2, padx=(10, 0), pady=8)
        ttk.Button(shop_bar, text="加载仓库列表", command=self._load_inventory_warehouses).grid(row=0, column=3, padx=(10, 0), pady=8)

        notebook = ttk.Notebook(content)
        notebook.grid(row=1, column=0, sticky="ew", pady=(0, 0))

        zero_tab = ttk.Frame(notebook, style="App.TFrame", padding=12)
        barcode_tab = ttk.Frame(notebook, style="App.TFrame", padding=12)
        notebook.add(zero_tab, text="零库存补货")
        notebook.add(barcode_tab, text="补充条形码")

        zero_tab.columnconfigure(0, weight=1)
        zero_actions = ttk.Frame(zero_tab, style="App.TFrame")
        zero_actions.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        ttk.Button(zero_actions, text="拉取零库存商品", command=self._fetch_zero_stock_products, style="Primary.TButton").pack(
            side="left"
        )
        ttk.Label(zero_actions, text="目标仓库", style="Card.TLabel").pack(side="left", padx=(16, 8))
        self.inventory_warehouse_combobox = ttk.Combobox(
            zero_actions, textvariable=self.inventory_warehouse_var, state="readonly", width=34
        )
        self.inventory_warehouse_combobox.pack(side="left")
        self.inventory_warehouse_combobox.bind(
            "<<ComboboxSelected>>", lambda _event: self._persist_inventory_warehouse_selection()
        )
        ttk.Label(zero_actions, text="库存数量", style="Card.TLabel").pack(side="left", padx=(16, 8))
        ttk.Entry(zero_actions, textvariable=self.inventory_stock_qty_var, width=10).pack(side="left")
        ttk.Button(zero_actions, text="更新选中商品库存", command=self._apply_zero_stock_updates, style="Primary.TButton").pack(
            side="left", padx=(16, 0)
        )

        zero_select_bar = ttk.Frame(zero_tab, style="App.TFrame")
        zero_select_bar.grid(row=1, column=0, sticky="w", pady=(0, 8))
        ttk.Button(zero_select_bar, text="全选", command=lambda: self._inventory_toggle_all(self.zero_stock_tree, True)).pack(
            side="left"
        )
        ttk.Button(zero_select_bar, text="取消全选", command=lambda: self._inventory_toggle_all(self.zero_stock_tree, False)).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(zero_select_bar, text="反选", command=lambda: self._inventory_invert_checks(self.zero_stock_tree)).pack(
            side="left", padx=(8, 0)
        )
        ttk.Label(zero_select_bar, text="点击列表首列勾选框可单选", style="Hint.TLabel").pack(side="left", padx=(16, 0))

        zero_table_wrap = ttk.Frame(zero_tab, style="Card.TFrame", padding=8)
        zero_table_wrap.grid(row=2, column=0, sticky="ew")
        zero_table_wrap.columnconfigure(0, weight=1)
        self.zero_stock_tree = self._create_inventory_checkbox_tree(
            zero_table_wrap,
            (
                ("offer_id", "货号", 140),
                ("product_id", "Product ID", 120),
                ("name", "商品名称", 300),
                ("stock_summary", "各仓库存", 240),
            ),
            height=14,
        )
        zero_scroll = ttk.Scrollbar(zero_table_wrap, orient="vertical", command=self.zero_stock_tree.yview)
        self.zero_stock_tree.configure(yscrollcommand=zero_scroll.set)
        self.zero_stock_tree.grid(row=0, column=0, sticky="ew")
        zero_scroll.grid(row=0, column=1, sticky="ns")
        ttk.Label(
            zero_tab,
            text="说明：拉取 visibility=EMPTY_STOCK 的商品；更新时请选择仓库与数量（每批最多 100 个，同一仓库约 30 秒限频）。",
            style="Hint.TLabel",
            wraplength=900,
        ).grid(row=3, column=0, sticky="w", pady=(10, 0))

        barcode_tab.columnconfigure(0, weight=1)
        barcode_actions = ttk.Frame(barcode_tab, style="App.TFrame")
        barcode_actions.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        ttk.Button(
            barcode_actions, text="拉取无条形码商品", command=self._fetch_no_barcode_products, style="Primary.TButton"
        ).pack(side="left")
        ttk.Button(
            barcode_actions, text="为选中商品生成条形码", command=self._generate_barcodes_for_selected, style="Primary.TButton"
        ).pack(side="left", padx=(12, 0))

        barcode_select_bar = ttk.Frame(barcode_tab, style="App.TFrame")
        barcode_select_bar.grid(row=1, column=0, sticky="w", pady=(0, 8))
        ttk.Button(
            barcode_select_bar, text="全选", command=lambda: self._inventory_toggle_all(self.no_barcode_tree, True)
        ).pack(side="left")
        ttk.Button(
            barcode_select_bar, text="取消全选", command=lambda: self._inventory_toggle_all(self.no_barcode_tree, False)
        ).pack(side="left", padx=(8, 0))
        ttk.Button(
            barcode_select_bar, text="反选", command=lambda: self._inventory_invert_checks(self.no_barcode_tree)
        ).pack(side="left", padx=(8, 0))
        ttk.Label(barcode_select_bar, text="点击列表首列勾选框可单选", style="Hint.TLabel").pack(side="left", padx=(16, 0))

        barcode_table_wrap = ttk.Frame(barcode_tab, style="Card.TFrame", padding=8)
        barcode_table_wrap.grid(row=2, column=0, sticky="ew")
        barcode_table_wrap.columnconfigure(0, weight=1)
        self.no_barcode_tree = self._create_inventory_checkbox_tree(
            barcode_table_wrap,
            (
                ("offer_id", "货号", 140),
                ("product_id", "Product ID", 120),
                ("name", "商品名称", 400),
            ),
            height=14,
        )
        barcode_scroll = ttk.Scrollbar(barcode_table_wrap, orient="vertical", command=self.no_barcode_tree.yview)
        self.no_barcode_tree.configure(yscrollcommand=barcode_scroll.set)
        self.no_barcode_tree.grid(row=0, column=0, sticky="ew")
        barcode_scroll.grid(row=0, column=1, sticky="ns")
        ttk.Label(
            barcode_tab,
            text="说明：拉取 visibility=EMPTY_BARCODE 的商品，通过 /v1/barcode/generate 生成条形码（每批最多 100 个）。",
            style="Hint.TLabel",
            wraplength=900,
        ).grid(row=3, column=0, sticky="w", pady=(10, 0))

        self._refresh_inventory_shop_options()
        self._apply_cached_inventory_warehouses_for_current_shop()
        self._bind_page_mousewheel(content, canvas)

    def _inventory_check_maps(self, tree: ttk.Treeview) -> tuple[dict[str, bool], str, str]:
        if tree is self.zero_stock_tree:
            return self.inventory_zero_checked, "☑", "☐"
        return self.inventory_no_barcode_checked, "☑", "☐"

    def _create_inventory_checkbox_tree(
        self,
        parent: ttk.Frame,
        columns: tuple[tuple[str, str, int], ...],
        *,
        height: int,
    ) -> ttk.Treeview:
        column_ids = ("checked", *(name for name, _title, _width in columns))
        tree = ttk.Treeview(parent, columns=column_ids, show="headings", selectmode="browse", height=height)
        tree.heading("checked", text="选择")
        tree.column("checked", width=52, anchor="center", stretch=False)
        for column_id, title, width in columns:
            tree.heading(column_id, text=title)
            tree.column(column_id, width=width, anchor="w")
        tree.bind("<Button-1>", lambda event, target=tree: self._on_inventory_tree_click(target, event))
        return tree

    def _on_inventory_tree_click(self, tree: ttk.Treeview, event: tk.Event) -> None:
        if tree.identify_region(event.x, event.y) != "cell":
            return
        row_id = tree.identify_row(event.y)
        if not row_id:
            return
        if tree.identify_column(event.x) != "#1":
            return
        self._toggle_inventory_row_check(tree, row_id)

    def _toggle_inventory_row_check(self, tree: ttk.Treeview, row_id: str) -> None:
        checked_map, on_symbol, off_symbol = self._inventory_check_maps(tree)
        checked = not checked_map.get(row_id, False)
        checked_map[row_id] = checked
        values = list(tree.item(row_id, "values"))
        if not values:
            return
        values[0] = on_symbol if checked else off_symbol
        tree.item(row_id, values=values)

    def _inventory_toggle_all(self, tree: ttk.Treeview, checked: bool) -> None:
        checked_map, on_symbol, off_symbol = self._inventory_check_maps(tree)
        symbol = on_symbol if checked else off_symbol
        for row_id in tree.get_children():
            checked_map[row_id] = checked
            values = list(tree.item(row_id, "values"))
            if not values:
                continue
            values[0] = symbol
            tree.item(row_id, values=values)

    def _inventory_invert_checks(self, tree: ttk.Treeview) -> None:
        checked_map, on_symbol, off_symbol = self._inventory_check_maps(tree)
        for row_id in tree.get_children():
            checked = not checked_map.get(row_id, False)
            checked_map[row_id] = checked
            values = list(tree.item(row_id, "values"))
            if not values:
                continue
            values[0] = on_symbol if checked else off_symbol
            tree.item(row_id, values=values)

    def _refresh_inventory_shop_options(self) -> None:
        labels = [
            self._shop_profile_label(shop_id, profile) for shop_id, profile in sorted(self.shop_profiles.items())
        ]
        if hasattr(self, "inventory_shop_combobox"):
            self.inventory_shop_combobox["values"] = labels
            if labels and self.inventory_shop_var.get() not in labels:
                self.inventory_shop_var.set(labels[0])

    def _current_inventory_shop_id(self) -> str:
        selected = self.inventory_shop_var.get().strip()
        shop_id = getattr(self, "_shop_display_to_id", {}).get(selected, normalize_ozon_client_id(selected))
        return normalize_ozon_client_id(shop_id)

    def _on_inventory_shop_changed(self) -> None:
        self._refresh_inventory_shop_options()
        self._apply_cached_inventory_warehouses_for_current_shop(log=True)

    def _apply_inventory_warehouses_ui(
        self, warehouses: list[WarehouseOption], selected_warehouse_id: int | None = None
    ) -> None:
        if not hasattr(self, "inventory_warehouse_combobox"):
            return
        self.inventory_warehouses = warehouses
        labels = [warehouse.label for warehouse in warehouses]
        self.inventory_warehouse_combobox["values"] = labels
        if not labels:
            self.inventory_warehouse_var.set("")
            return
        if selected_warehouse_id is not None:
            for warehouse in warehouses:
                if warehouse.warehouse_id == selected_warehouse_id:
                    self.inventory_warehouse_var.set(warehouse.label)
                    return
        current = self.inventory_warehouse_var.get().strip()
        if current in labels:
            return
        self.inventory_warehouse_var.set(labels[0])

    def _save_inventory_warehouses_cache(
        self, shop_id: str, warehouses: list[WarehouseOption], *, selected_warehouse_id: int | None = None
    ) -> None:
        clean_shop_id = normalize_ozon_client_id(shop_id)
        if not clean_shop_id:
            return
        if selected_warehouse_id is None:
            try:
                selected_warehouse_id = self._selected_inventory_warehouse_id()
            except RuntimeError:
                selected_warehouse_id = warehouses[0].warehouse_id if warehouses else None
        self.inventory_warehouses_by_shop[clean_shop_id] = {
            "warehouses": serialize_warehouse_options(warehouses),
            "selected_warehouse_id": selected_warehouse_id,
        }
        self._save_config(silent=True)

    def _apply_cached_inventory_warehouses_for_current_shop(self, *, log: bool = False) -> None:
        shop_id = self._current_inventory_shop_id()
        if not shop_id:
            self._apply_inventory_warehouses_ui([], None)
            return
        entry = self.inventory_warehouses_by_shop.get(shop_id)
        if not isinstance(entry, dict):
            self._apply_inventory_warehouses_ui([], None)
            return
        warehouses = deserialize_warehouse_options(entry.get("warehouses") or [])
        if not warehouses:
            self._apply_inventory_warehouses_ui([], None)
            return
        selected = entry.get("selected_warehouse_id")
        selected_id = int(selected) if selected is not None and str(selected).strip() else None
        self._apply_inventory_warehouses_ui(warehouses, selected_id)
        if log:
            self._append_log(f"已使用本地缓存的仓库列表（{len(warehouses)} 个）。")

    def _persist_inventory_warehouse_selection(self) -> None:
        shop_id = self._current_inventory_shop_id()
        if not shop_id or not self.inventory_warehouses:
            return
        entry = self.inventory_warehouses_by_shop.get(shop_id)
        if not isinstance(entry, dict):
            return
        try:
            selected_id = self._selected_inventory_warehouse_id()
        except RuntimeError:
            return
        if entry.get("selected_warehouse_id") == selected_id:
            return
        entry["selected_warehouse_id"] = selected_id
        self._save_config(silent=True)

    def _inventory_client_from_shop(self) -> OzonSellerClient:
        selected = self.inventory_shop_var.get().strip()
        shop_id = getattr(self, "_shop_display_to_id", {}).get(selected, normalize_ozon_client_id(selected))
        profile = self.shop_profiles.get(shop_id)
        if not profile:
            raise RuntimeError("请先选择有效店铺。")
        client_id = normalize_ozon_client_id(str(profile.get("client_id") or shop_id))
        api_key = str(profile.get("api_key") or "").strip()
        if not client_id:
            raise RuntimeError("店铺 Client-Id 未配置。")
        if not api_key:
            raise RuntimeError("店铺 Api-Key 未配置。")
        return OzonSellerClient(client_id, api_key, timeout_seconds=120)

    def _load_inventory_warehouses(self) -> None:
        try:
            client = self._inventory_client_from_shop()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return

        def worker() -> None:
            try:
                warehouses = load_warehouses(client)
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self._append_log(f"加载仓库失败: {message}")
                self.after(0, lambda: messagebox.showerror("加载失败", message))
                return

            def apply() -> None:
                shop_id = self._current_inventory_shop_id()
                self._apply_inventory_warehouses_ui(warehouses)
                if shop_id:
                    self._save_inventory_warehouses_cache(shop_id, warehouses)
                self._append_log(f"已加载 {len(warehouses)} 个仓库，并已保存到本地。")
                messagebox.showinfo("加载成功", f"已加载 {len(warehouses)} 个仓库，并已保存到本地。")

            self.after(0, apply)

        self._append_log("正在加载 Ozon 仓库列表...")
        threading.Thread(target=worker, daemon=True).start()

    def _fill_inventory_tree(self, tree: ttk.Treeview, row_map: dict[str, ProductInventoryRow], rows: list[ProductInventoryRow]) -> None:
        tree.delete(*tree.get_children())
        row_map.clear()
        checked_map, _on_symbol, off_symbol = self._inventory_check_maps(tree)
        checked_map.clear()
        for row in rows:
            iid = str(row.product_id)
            row_map[iid] = row
            checked_map[iid] = False
            if tree is self.zero_stock_tree:
                tree.insert(
                    "",
                    "end",
                    iid=iid,
                    values=(off_symbol, row.offer_id, row.product_id, row.name, row.stock_summary),
                )
            else:
                tree.insert("", "end", iid=iid, values=(off_symbol, row.offer_id, row.product_id, row.name))

    def _selected_inventory_rows(self, tree: ttk.Treeview, row_map: dict[str, ProductInventoryRow]) -> list[ProductInventoryRow]:
        checked_map, _, _ = self._inventory_check_maps(tree)
        selected: list[ProductInventoryRow] = []
        for iid in tree.get_children():
            if not checked_map.get(iid):
                continue
            row = row_map.get(str(iid))
            if row:
                selected.append(row)
        return selected

    def _selected_inventory_warehouse_id(self) -> int:
        selected_label = self.inventory_warehouse_var.get().strip()
        for warehouse in self.inventory_warehouses:
            if warehouse.label == selected_label:
                return warehouse.warehouse_id
        raise RuntimeError("请先加载并选择目标仓库。")

    def _fetch_zero_stock_products(self) -> None:
        try:
            client = self._inventory_client_from_shop()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return

        def worker() -> None:
            try:
                rows = fetch_zero_stock_products(client, logger=self._append_log)
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self._append_log(f"拉取零库存商品失败: {message}")
                self.after(0, lambda: messagebox.showerror("拉取失败", message))
                return

            def apply() -> None:
                self._fill_inventory_tree(self.zero_stock_tree, self.inventory_zero_stock_rows, rows)
                messagebox.showinfo("拉取完成", f"共找到 {len(rows)} 个零库存商品。")

            self.after(0, apply)

        self._append_log("开始拉取零库存商品...")
        threading.Thread(target=worker, daemon=True).start()

    def _apply_zero_stock_updates(self) -> None:
        rows = self._selected_inventory_rows(self.zero_stock_tree, self.inventory_zero_stock_rows)
        if not rows:
            messagebox.showinfo("未选择商品", "请先勾选要更新库存的商品（可点击「全选」）。")
            return
        try:
            warehouse_id = self._selected_inventory_warehouse_id()
            stock = int(str(self.inventory_stock_qty_var.get()).strip() or "0")
            client = self._inventory_client_from_shop()
        except ValueError:
            messagebox.showerror("参数错误", "库存数量必须是整数。")
            return
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return

        if not messagebox.askyesno("确认更新", f"将为 {len(rows)} 个商品在仓库 {warehouse_id} 设置库存 {stock}，是否继续？"):
            return

        def worker() -> None:
            try:
                success, failed = update_product_stocks(
                    client,
                    rows,
                    warehouse_id=warehouse_id,
                    stock=stock,
                    logger=self._append_log,
                )
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self._append_log(f"更新库存失败: {message}")
                self.after(0, lambda: messagebox.showerror("更新失败", message))
                return

            def apply() -> None:
                messagebox.showinfo("更新完成", f"成功 {success} 个，失败 {failed} 个。")

            self.after(0, apply)

        self._append_log(f"开始更新 {len(rows)} 个商品库存...")
        threading.Thread(target=worker, daemon=True).start()

    def _fetch_no_barcode_products(self) -> None:
        try:
            client = self._inventory_client_from_shop()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return

        def worker() -> None:
            try:
                rows = fetch_no_barcode_products(client, logger=self._append_log)
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self._append_log(f"拉取无条形码商品失败: {message}")
                self.after(0, lambda: messagebox.showerror("拉取失败", message))
                return

            def apply() -> None:
                self._fill_inventory_tree(self.no_barcode_tree, self.inventory_no_barcode_rows, rows)
                messagebox.showinfo("拉取完成", f"共找到 {len(rows)} 个无条形码商品。")

            self.after(0, apply)

        self._append_log("开始拉取无条形码商品...")
        threading.Thread(target=worker, daemon=True).start()

    def _generate_barcodes_for_selected(self) -> None:
        rows = self._selected_inventory_rows(self.no_barcode_tree, self.inventory_no_barcode_rows)
        if not rows:
            messagebox.showinfo("未选择商品", "请先勾选要生成条形码的商品（可点击「全选」）。")
            return
        try:
            client = self._inventory_client_from_shop()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        if not messagebox.askyesno("确认生成", f"将为 {len(rows)} 个商品调用 Ozon 生成条形码，是否继续？"):
            return

        def worker() -> None:
            try:
                success, failed = generate_barcodes_for_products(client, rows, logger=self._append_log)
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self._append_log(f"生成条形码失败: {message}")
                self.after(0, lambda: messagebox.showerror("生成失败", message))
                return

            def apply() -> None:
                messagebox.showinfo("生成完成", f"成功 {success} 个，失败 {failed} 个。")

            self.after(0, apply)

        self._append_log(f"开始为 {len(rows)} 个商品生成条形码...")
        threading.Thread(target=worker, daemon=True).start()

    def _section(self, parent: ttk.Frame, title: str, hint: str, row: int) -> ttk.Frame:
        card = ttk.Frame(parent, style="Card.TFrame", padding=(20, 18))
        card.grid(row=row, column=0, sticky="ew", pady=(0, 16))
        card.columnconfigure(0, weight=1)
        ttk.Label(card, text=title, style="Section.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(card, text=hint, style="Hint.TLabel").grid(row=1, column=0, sticky="w", pady=(4, 14))
        body = ttk.Frame(card, style="CardInner.TFrame")
        body.grid(row=2, column=0, sticky="ew")
        return body

    def _field(
        self,
        parent: ttk.Frame,
        row: int,
        column: int,
        label: str,
        variable: tk.StringVar,
        *,
        show: str | None = None,
        width: int | None = None,
        colspan: int = 1,
    ) -> None:
        ttk.Label(parent, text=label, style="Card.TLabel").grid(row=row, column=column, sticky="w", padx=(0, 10), pady=8)
        ttk.Entry(parent, textvariable=variable, show=show, width=width).grid(
            row=row, column=column + 1, columnspan=colspan, sticky="ew", pady=8, padx=(0, 18 if column + colspan < 3 else 0)
        )

    def _current_image_provider_id(self) -> str:
        label = self.image_provider_var.get().strip()
        for provider_id in list_image_provider_ids():
            if provider_label(provider_id) == label or provider_id == label:
                return provider_id
        return DEFAULT_IMAGE_PROVIDER

    def _sync_provider_api_key_from_field(self) -> None:
        if self._provider_switching:
            return
        provider_id = self._current_image_provider_id()
        key = self.image_api_key_var.get().strip()
        if key:
            self.provider_api_keys[provider_id] = key

    def _apply_image_provider_ui(self, provider_id: str, *, initial: bool = False) -> None:
        if provider_id not in IMAGE_PROVIDER_PROFILES:
            provider_id = DEFAULT_IMAGE_PROVIDER
        self._provider_switching = True
        try:
            self.image_provider_var.set(provider_label(provider_id))
            if hasattr(self, "image_provider_combobox"):
                self.image_provider_combobox.set(provider_label(provider_id))
            models = provider_models(provider_id)
            if hasattr(self, "image_model_combobox"):
                self.image_model_combobox.configure(values=models)
            current_model = self.image_model_var.get().strip()
            if initial or current_model not in models:
                self.image_model_var.set(provider_default_model(provider_id))
            if initial and not self.image_base_url_var.get().strip():
                self.image_base_url_var.set(provider_default_base_url(provider_id))
            elif not initial:
                self.image_base_url_var.set(provider_default_base_url(provider_id))
            stored_key = self.provider_api_keys.get(provider_id, "")
            if not stored_key and provider_id.startswith("apimart_"):
                for other_id in list_image_provider_ids():
                    if other_id.startswith("apimart_") and other_id != provider_id:
                        stored_key = self.provider_api_keys.get(other_id, "")
                        if stored_key:
                            break
            if not stored_key and provider_id == "xiaoqian":
                stored_key = self.text_provider_api_keys.get("xiaoqian", "")
            if stored_key:
                self.image_api_key_var.set(stored_key)
        finally:
            self._provider_switching = False

    def _on_image_provider_selected(self, _event: object | None = None) -> None:
        if self._provider_switching:
            return
        self._sync_provider_api_key_from_field()
        selected_label = ""
        if hasattr(self, "image_provider_combobox"):
            selected_label = self.image_provider_combobox.get().strip()
        provider_id = DEFAULT_IMAGE_PROVIDER
        for pid in list_image_provider_ids():
            if provider_label(pid) == selected_label:
                provider_id = pid
                break
        self._apply_image_provider_ui(provider_id)

    def _current_text_provider_id(self) -> str:
        label = self.text_provider_var.get().strip()
        for provider_id in list_text_provider_ids():
            if text_provider_label(provider_id) == label or provider_id == label:
                return provider_id
        return DEFAULT_TEXT_PROVIDER

    def _sync_text_provider_api_key_from_field(self) -> None:
        if self._text_provider_switching:
            return
        provider_id = self._current_text_provider_id()
        key = self.text_api_key_var.get().strip()
        if key:
            self.text_provider_api_keys[provider_id] = key

    def _apply_text_provider_ui(self, provider_id: str, *, initial: bool = False) -> None:
        if provider_id not in TEXT_PROVIDER_PROFILES:
            provider_id = DEFAULT_TEXT_PROVIDER
        self._text_provider_switching = True
        try:
            self.text_provider_var.set(text_provider_label(provider_id))
            if hasattr(self, "text_provider_combobox"):
                self.text_provider_combobox.set(text_provider_label(provider_id))
            models = text_provider_models(provider_id)
            if hasattr(self, "text_model_combobox"):
                self.text_model_combobox.configure(values=models)
            current_model = self.text_model_var.get().strip()
            if initial or current_model not in models:
                self.text_model_var.set(text_provider_default_model(provider_id))
            if initial and not self.base_url_var.get().strip():
                self.base_url_var.set(text_provider_default_base_url(provider_id))
            elif not initial:
                self.base_url_var.set(text_provider_default_base_url(provider_id))
            stored_key = self.text_provider_api_keys.get(provider_id, "")
            if not stored_key and provider_id == "xiaoqian":
                stored_key = self.provider_api_keys.get("xiaoqian", "")
            if stored_key:
                self.text_api_key_var.set(stored_key)
        finally:
            self._text_provider_switching = False

    def _on_text_provider_selected(self, _event: object | None = None) -> None:
        if self._text_provider_switching:
            return
        self._sync_text_provider_api_key_from_field()
        selected_label = ""
        if hasattr(self, "text_provider_combobox"):
            selected_label = self.text_provider_combobox.get().strip()
        provider_id = DEFAULT_TEXT_PROVIDER
        for pid in list_text_provider_ids():
            if text_provider_label(pid) == selected_label:
                provider_id = pid
                break
        self._apply_text_provider_ui(provider_id)

    def _text_box(self, parent: ttk.Frame, *, height: int) -> tk.Text:
        return tk.Text(
            parent,
            height=height,
            wrap="word",
            bg="#ffffff",
            fg="#262626",
            insertbackground="#1677ff",
            relief="flat",
            highlightthickness=1,
            highlightbackground="#d9d9d9",
            highlightcolor="#1677ff",
            padx=10,
            pady=10,
            font=(UI_FONT_FAMILY, 10),
        )

    def _path_row(
        self,
        parent: ttk.Frame,
        row: int,
        label: str,
        variable: tk.StringVar,
        select_file: bool = False,
    ) -> None:
        ttk.Label(parent, text=label, style="Card.TLabel").grid(row=row, column=0, sticky="w", pady=7, padx=(0, 10))
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", pady=7)
        chooser = (lambda var=variable: self._choose_file(var)) if select_file else (lambda var=variable: self._choose_folder(var))
        ttk.Button(parent, text="选择", command=chooser).grid(
            row=row, column=2, sticky="e", padx=(10, 0), pady=7
        )

    def _choose_folder(self, variable: tk.StringVar) -> None:
        selected = filedialog.askdirectory()
        if selected:
            variable.set(selected)

    def _choose_file(self, variable: tk.StringVar) -> None:
        selected = filedialog.askopenfilename(
            filetypes=[
                ("Image Files", "*.png;*.jpg;*.jpeg;*.webp;*.bmp"),
                ("All Files", "*.*"),
            ]
        )
        if selected:
            variable.set(selected)

    def _choose_excel(self, variable: tk.StringVar) -> None:
        selected = filedialog.askopenfilename(
            filetypes=[
                ("Excel Files", "*.xlsx;*.xlsm"),
                ("All Files", "*.*"),
            ]
        )
        if selected:
            variable.set(selected)

    def _download_upload_template(self) -> None:
        selected = filedialog.asksaveasfilename(
            title="保存上架 Excel 模板",
            defaultextension=".xlsx",
            initialfile="ozon_batch_upload_template.xlsx",
            filetypes=[
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*"),
            ],
        )
        if not selected:
            return
        try:
            target = Path(selected)
            create_upload_template(target)
            self.upload_excel_path_var.set(str(target))
            self._append_log(f"上架 Excel 模板已保存: {target}")
            messagebox.showinfo("模板已保存", f"上架 Excel 模板已保存:\n{target}")
        except Exception as exc:  # noqa: BLE001
            message = str(exc)
            self._append_log(f"上架 Excel 模板保存失败: {message}")
            messagebox.showerror("模板保存失败", message)

    def _save_ozon_config(self) -> None:
        ozon_client_id = normalize_ozon_client_id(self.ozon_client_id_var.get())
        if ozon_client_id and ozon_client_id != self.ozon_client_id_var.get().strip():
            self.ozon_client_id_var.set(ozon_client_id)
        if not ozon_client_id:
            messagebox.showerror("配置错误", "请填写 Ozon Client-Id。")
            return
        if not ozon_client_id.isascii():
            messagebox.showerror("配置错误", "Ozon Client-Id 只能填写数字 ID，请不要包含店铺名称或中文。")
            return
        if not self.ozon_api_key_var.get().strip():
            messagebox.showerror("配置错误", "请填写 Ozon Api-Key。")
            return
        if not self.ozon_api_key_var.get().strip().isascii():
            messagebox.showerror("配置错误", "Ozon Api-Key 不能包含中文或特殊非英文字符，请检查是否粘贴了多余内容。")
            return
        self._save_config(silent=True)
        self._append_log("Ozon 配置已保存。")
        messagebox.showinfo("保存成功", "Ozon 配置已保存。")

    def _shop_label(self, shop_id: str, profile: dict[str, str]) -> str:
        name = str(profile.get("name") or "").strip()
        return f"{name} ({shop_id})" if name else shop_id

    def _refresh_shop_options(self) -> None:
        if not hasattr(self, "shop_combobox"):
            return
        labels = [self._shop_label(shop_id, profile) for shop_id, profile in sorted(self.shop_profiles.items())]
        self.shop_combobox["values"] = labels

    def _selected_shop_id(self) -> str:
        selected = self.selected_shop_var.get().strip()
        match = re.search(r"\((\d+)\)$", selected)
        if match:
            return match.group(1)
        return normalize_ozon_client_id(selected)

    def _load_selected_shop(self) -> None:
        shop_id = self._selected_shop_id()
        profile = self.shop_profiles.get(shop_id)
        if not profile:
            return
        self.shop_name_var.set(str(profile.get("name", "")))
        self.ozon_client_id_var.set(shop_id)
        self.ozon_api_key_var.set(str(profile.get("api_key", "")))
        template_offer_id = str(profile.get("template_offer_id", "")).strip()
        if template_offer_id:
            self.ozon_template_offer_id_var.set(template_offer_id)
        self._append_log(f"已切换店铺：{self._shop_label(shop_id, profile)}")

    def _save_shop_profile(self) -> None:
        shop_id = normalize_ozon_client_id(self.ozon_client_id_var.get())
        if shop_id and shop_id != self.ozon_client_id_var.get().strip():
            self.ozon_client_id_var.set(shop_id)
        if not shop_id:
            messagebox.showerror("店铺错误", "请先填写 Ozon Client-Id。")
            return
        if not self.ozon_api_key_var.get().strip():
            messagebox.showerror("店铺错误", "请先填写 Ozon Api-Key。")
            return
        profile = {
            "name": self.shop_name_var.get().strip() or shop_id,
            "api_key": self.ozon_api_key_var.get().strip(),
            "template_offer_id": self.ozon_template_offer_id_var.get().strip(),
        }
        self.shop_profiles[shop_id] = profile
        self.selected_shop_var.set(self._shop_label(shop_id, profile))
        self._refresh_shop_options()
        self._save_config(silent=True)
        self._append_log(f"已保存店铺：{self._shop_label(shop_id, profile)}")

    def _delete_shop_profile(self) -> None:
        shop_id = self._selected_shop_id() or normalize_ozon_client_id(self.ozon_client_id_var.get())
        if not shop_id or shop_id not in self.shop_profiles:
            messagebox.showerror("店铺错误", "请先选择已保存店铺。")
            return
        del self.shop_profiles[shop_id]
        self.selected_shop_var.set("")
        self._refresh_shop_options()
        self._save_config(silent=True)
        self._append_log(f"已删除店铺：{shop_id}")

    def _template_key(self, shop_id: str, template_name: str) -> str:
        return f"{shop_id}/{template_name}"

    def _template_oss_key(self, shop_id: str, template_name: str) -> str:
        clean_name = re.sub(r"[\\/]+", "_", template_name.strip()).strip() or "template"
        return f"{shop_id}/{clean_name}/{clean_name}.json"

    def _refresh_product_template_options(self) -> None:
        if not hasattr(self, "product_template_combobox"):
            return
        self.product_template_combobox["values"] = sorted(self.product_templates.keys())

    def _load_selected_product_template(self) -> None:
        key = self.selected_product_template_var.get().strip()
        template = self.product_templates.get(key)
        if not template:
            return
        self.product_template_name_var.set(str(template.get("template_name", "")))
        self.ozon_template_offer_id_var.set(str(template.get("template_offer_id", "")))
        self._append_log(f"已选择商品模板：{key}")

    def _clear_product_template_selection(self) -> None:
        self.selected_product_template_var.set("")
        self.product_template_name_var.set("")
        self._append_log("已清空商品模板选择，将按模板商品货号实时读取。")

    def _save_product_template(self) -> None:
        template_name = self.product_template_name_var.get().strip()
        if not template_name:
            messagebox.showerror("模板错误", "请填写商品模板名称。")
            return
        try:
            config = self._collect_upload_config(require_batch_files=False, require_ozon=True, require_oss=True)
        except RuntimeError as exc:
            messagebox.showerror("模板错误", str(exc))
            return
        shop_id = config.ozon_client_id
        key = self._template_key(shop_id, template_name)
        if key in self.product_templates:
            messagebox.showinfo("模板已存在", "商品模板保存后不允许修改。请换一个模板名称。")
            return

        def worker() -> None:
            try:
                self._append_log(f"正在读取并保存商品模板：{template_name}")
                ozon_client = OzonSellerClient(config.ozon_client_id, config.ozon_api_key, timeout_seconds=120)
                template_product = ozon_client.get_template_product(config.template_offer_id)
                oss_client = AliyunOssClient(
                    config.oss_access_key_id,
                    config.oss_access_key_secret,
                    config.oss_bucket,
                    config.oss_endpoint,
                    config.oss_public_domain,
                )
                oss_key = self._template_oss_key(shop_id, template_name)
                payload = {
                    "shop_id": shop_id,
                    "shop_name": self.shop_name_var.get().strip(),
                    "template_name": template_name,
                    "template_offer_id": config.template_offer_id,
                    "saved_at": now_stamp(),
                    "oss_key": oss_key,
                    "template_product": template_product,
                }
                payload_bytes = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
                oss_url = oss_client.upload_bytes(payload_bytes, oss_key, "application/json; charset=utf-8")
                payload["oss_url"] = oss_url
                self.product_templates[key] = payload
                self.selected_product_template_var.set(key)
                self.after(0, self._refresh_product_template_options)
                self._save_config(silent=True)
                self._append_log(f"商品模板已保存到 OSS: {oss_url}")
                self.after(0, lambda: messagebox.showinfo("保存成功", f"商品模板已保存到 OSS:\n{oss_url}"))
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self._append_log(f"商品模板保存失败: {message}")
                self.after(0, lambda: messagebox.showerror("保存失败", message))

        threading.Thread(target=worker, daemon=True).start()

    def _append_log(self, message: str) -> None:
        self.log_queue.put(f"{now_stamp()}  {message}")

    def _current_prompt_bundle(self) -> dict[str, str]:
        if not all(
            hasattr(self, name)
            for name in ("image_prompt_text", "title_prompt_text", "description_prompt_text")
        ):
            selected_name = self.template_name_var.get().strip() or DEFAULT_TEMPLATE_NAME
            bundle = self.prompt_templates.get(selected_name) or self.prompt_templates.get(DEFAULT_TEMPLATE_NAME) or {}
            return {
                "image_prompt_template": str(bundle.get("image_prompt_template", DEFAULT_IMAGE_PROMPT)),
                "title_prompt_template": str(bundle.get("title_prompt_template", DEFAULT_TITLE_PROMPT)),
                "description_prompt_template": str(bundle.get("description_prompt_template", DEFAULT_DESCRIPTION_PROMPT)),
            }
        return {
            "image_prompt_template": self.image_prompt_text.get("1.0", "end").strip(),
            "title_prompt_template": self.title_prompt_text.get("1.0", "end").strip(),
            "description_prompt_template": self.description_prompt_text.get("1.0", "end").strip(),
        }

    def _apply_prompt_bundle(self, bundle: dict[str, str]) -> None:
        if not all(
            hasattr(self, name)
            for name in ("image_prompt_text", "title_prompt_text", "description_prompt_text")
        ):
            return
        self.image_prompt_text.delete("1.0", "end")
        self.image_prompt_text.insert("1.0", bundle.get("image_prompt_template", DEFAULT_IMAGE_PROMPT))
        self.title_prompt_text.delete("1.0", "end")
        self.title_prompt_text.insert("1.0", bundle.get("title_prompt_template", DEFAULT_TITLE_PROMPT))
        self.description_prompt_text.delete("1.0", "end")
        self.description_prompt_text.insert(
            "1.0",
            bundle.get("description_prompt_template", DEFAULT_DESCRIPTION_PROMPT),
        )

    def _refresh_template_options(self) -> None:
        names = sorted(self.prompt_templates.keys())
        if hasattr(self, "template_combobox"):
            self.template_combobox["values"] = names
        if not self.template_name_var.get().strip():
            self.template_name_var.set(DEFAULT_TEMPLATE_NAME)

    def _current_scene_prompt_template(self) -> str:
        if hasattr(self, "scene_prompt_text"):
            self._scene_prompt_template = self.scene_prompt_text.get("1.0", "end").strip()
        return self._scene_prompt_template or DEFAULT_SCENE_PROMPT_TEMPLATE

    def _load_selected_template(self) -> None:
        name = self.template_name_var.get().strip()
        if not name:
            messagebox.showerror("模板错误", "请先输入或选择模板名。")
            return
        bundle = self.prompt_templates.get(name)
        if not bundle:
            messagebox.showerror("模板不存在", f"未找到模板：{name}")
            return
        self._apply_prompt_bundle(bundle)
        self._append_log(f"已加载提示词模板：{name}")

    def _save_prompt_template(self) -> None:
        name = self.template_name_var.get().strip()
        if not name:
            messagebox.showerror("模板错误", "请先输入模板名。")
            return
        self.prompt_templates[name] = self._current_prompt_bundle()
        self._refresh_template_options()
        self.template_name_var.set(name)
        self._save_config(silent=True)
        self._append_log(f"已保存提示词模板：{name}")

    def _delete_prompt_template(self) -> None:
        name = self.template_name_var.get().strip()
        if not name:
            messagebox.showerror("模板错误", "请先选择模板名。")
            return
        if name == DEFAULT_TEMPLATE_NAME:
            messagebox.showinfo("无法删除", "默认模板不能删除。")
            return
        if name not in self.prompt_templates:
            messagebox.showerror("模板不存在", f"未找到模板：{name}")
            return
        del self.prompt_templates[name]
        self.template_name_var.set(DEFAULT_TEMPLATE_NAME)
        self._refresh_template_options()
        self._apply_prompt_bundle(self.prompt_templates[DEFAULT_TEMPLATE_NAME])
        self._save_config(silent=True)
        self._append_log(f"已删除提示词模板：{name}")

    def _flush_logs(self) -> None:
        if not hasattr(self, "log_text"):
            self.after(250, self._flush_logs)
            return
        lines = []
        for _ in range(LOG_FLUSH_BATCH_SIZE):
            try:
                lines.append(self.log_queue.get_nowait())
            except queue.Empty:
                break
        if lines:
            self.log_text.configure(state="normal")
            self.log_text.insert("end", "\n".join(lines) + "\n")
            self.log_line_count += len(lines)
            if self.log_line_count > LOG_MAX_LINES:
                remove_count = self.log_line_count - LOG_MAX_LINES
                self.log_text.delete("1.0", f"{remove_count + 1}.0")
                self.log_line_count = LOG_MAX_LINES
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        delay = 50 if not self.log_queue.empty() else 250
        self.after(delay, self._flush_logs)

    def _collect_config(self) -> JobConfig:
        source_root = Path(self.source_root_var.get()).expanduser()
        generated_root = Path(self.generated_root_var.get()).expanduser()
        portrait_root = Path(self.portrait_root_var.get()).expanduser()
        content_root = Path(self.content_root_var.get()).expanduser()
        watermark_text = self.watermark_path_var.get().strip()
        watermark_path = Path(watermark_text).expanduser() if watermark_text else None

        if not source_root.is_dir():
            raise RuntimeError("请先选择有效的源目录。")
        if not self.generated_root_var.get().strip():
            raise RuntimeError("请先选择 AI 竖版图输出目录。")
        if not self.portrait_root_var.get().strip():
            raise RuntimeError("请先选择 3:4 输出目录。")
        if self.generate_copy_var.get() and not self.content_root_var.get().strip():
            raise RuntimeError("请先选择文案输出目录。")
        if not self.image_api_key_var.get().strip():
            raise RuntimeError("请填写图片 API Key。")
        if self.generate_copy_var.get() and not self.text_api_key_var.get().strip():
            raise RuntimeError("请填写文本 API Key。")
        if not self.image_base_url_var.get().strip():
            raise RuntimeError("请填写图片接口地址。")
        if self.generate_copy_var.get() and not self.base_url_var.get().strip():
            raise RuntimeError("请填写文案接口地址。")
        if watermark_path and not watermark_path.is_file():
            raise RuntimeError("水印图片不存在，请重新选择。")

        try:
            max_folders = max(0, int(self.max_folders_var.get().strip() or "0"))
            max_workers = max(1, int(self.max_workers_var.get().strip() or "1"))
        except ValueError as exc:
            raise RuntimeError("最多处理文件夹和图片并发数必须是整数。") from exc

        self._sync_provider_api_key_from_field()
        self._sync_text_provider_api_key_from_field()
        provider_id = self._current_image_provider_id()
        text_provider_id = self._current_text_provider_id()
        prompt_bundle = self._current_prompt_bundle()
        return JobConfig(
            source_root=source_root,
            generated_root=generated_root,
            portrait_root=portrait_root,
            content_root=content_root,
            image_api_key=self.image_api_key_var.get().strip(),
            text_api_key=(self.text_api_key_var.get().strip() or self.image_api_key_var.get().strip()),
            base_url=self.base_url_var.get().strip().rstrip("/"),
            image_base_url=self.image_base_url_var.get().strip().rstrip("/"),
            image_provider=provider_id,
            image_model=self.image_model_var.get().strip() or provider_default_model(provider_id),
            text_provider=text_provider_id,
            text_model=self.text_model_var.get().strip() or text_provider_default_model(text_provider_id),
            image_prompt_template=prompt_bundle["image_prompt_template"],
            title_prompt_template=prompt_bundle["title_prompt_template"],
            description_prompt_template=prompt_bundle["description_prompt_template"],
            watermark_path=watermark_path,
            quality=self.quality_var.get().strip(),
            max_folders=max_folders,
            max_workers=max_workers,
            convert_originals=self.convert_originals_var.get(),
            generate_copy=self.generate_copy_var.get(),
            export_excel=self.export_excel_var.get(),
        )

    def _collect_upload_config(
        self,
        *,
        require_batch_files: bool = True,
        require_ozon: bool = True,
        require_oss: bool = True,
    ) -> BatchUploadConfig:
        portrait_root = Path(self.portrait_root_var.get()).expanduser()
        excel_path = Path(self.upload_excel_path_var.get()).expanduser()
        if require_batch_files and not portrait_root.is_dir():
            raise RuntimeError("请先选择有效的 3:4 输出目录。")
        if require_batch_files and not excel_path.is_file():
            raise RuntimeError("请先选择有效的上架 Excel 文件。")
        ozon_client_id = normalize_ozon_client_id(self.ozon_client_id_var.get())
        if ozon_client_id and ozon_client_id != self.ozon_client_id_var.get().strip():
            self.ozon_client_id_var.set(ozon_client_id)
        if require_ozon and not ozon_client_id:
            raise RuntimeError("请填写 Ozon Client-Id。")
        if require_ozon and not ozon_client_id.isascii():
            raise RuntimeError("Ozon Client-Id 只能填写数字 ID，请不要包含店铺名称或中文。")
        if require_ozon and not self.ozon_api_key_var.get().strip():
            raise RuntimeError("请填写 Ozon Api-Key。")
        if require_ozon and not self.ozon_api_key_var.get().strip().isascii():
            raise RuntimeError("Ozon Api-Key 不能包含中文或特殊非英文字符，请检查是否粘贴了多余内容。")
        if require_ozon and not self.ozon_template_offer_id_var.get().strip():
            raise RuntimeError("请填写模板商品货号。")
        if require_oss and not self.oss_access_key_id_var.get().strip():
            raise RuntimeError("请填写 OSS AccessKeyId。")
        if require_oss and not self.oss_access_key_secret_var.get().strip():
            raise RuntimeError("请填写 OSS AccessKeySecret。")
        if require_oss and not self.oss_bucket_var.get().strip():
            raise RuntimeError("请填写 OSS Bucket。")
        if require_oss and not self.oss_endpoint_var.get().strip():
            raise RuntimeError("请填写 OSS Endpoint。")
        try:
            max_items = max(0, int(self.upload_max_items_var.get().strip() or "0"))
        except ValueError as exc:
            raise RuntimeError("最多上架货号必须是整数。") from exc

        selected_template = self.product_templates.get(self.selected_product_template_var.get().strip())
        template_product = None
        template_offer_id = self.ozon_template_offer_id_var.get().strip()
        if selected_template:
            template_product = selected_template.get("template_product") if isinstance(selected_template, dict) else None
            template_offer_id = str(selected_template.get("template_offer_id") or template_offer_id)

        return BatchUploadConfig(
            portrait_root=portrait_root,
            excel_path=excel_path,
            ozon_client_id=ozon_client_id,
            ozon_api_key=self.ozon_api_key_var.get().strip(),
            template_offer_id=template_offer_id,
            oss_access_key_id=self.oss_access_key_id_var.get().strip(),
            oss_access_key_secret=self.oss_access_key_secret_var.get().strip(),
            oss_bucket=self.oss_bucket_var.get().strip(),
            oss_endpoint=self.oss_endpoint_var.get().strip(),
            oss_public_domain=self.oss_public_domain_var.get().strip(),
            max_items=max_items,
            template_product=template_product,
        )

    def _collect_portrait_only_config(self) -> tuple[Path, Path, Path | None, int]:
        source_root = Path(self.source_root_var.get()).expanduser()
        portrait_root = Path(self.portrait_root_var.get()).expanduser()
        watermark_text = self.watermark_path_var.get().strip()
        watermark_path = Path(watermark_text).expanduser() if watermark_text else None
        if not source_root.is_dir():
            raise RuntimeError("请先选择有效的源目录。")
        if not self.portrait_root_var.get().strip():
            raise RuntimeError("请先选择 3:4 输出目录。")
        if watermark_path and not watermark_path.is_file():
            raise RuntimeError("水印图片不存在，请重新选择。")
        try:
            max_folders = max(0, int(self.max_folders_var.get().strip() or "0"))
        except ValueError as exc:
            raise RuntimeError("最多处理文件夹必须是整数。") from exc
        return source_root, portrait_root, watermark_path, max_folders

    def _start_portrait_only(self) -> None:
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showinfo("任务运行中", "当前已有任务在运行。")
            return
        try:
            source_root, portrait_root, watermark_path, max_folders = self._collect_portrait_only_config()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        self._save_config(silent=True)
        self.worker = None
        self.worker_thread = threading.Thread(
            target=self._run_portrait_only,
            args=(source_root, portrait_root, watermark_path, max_folders),
            daemon=True,
        )
        self.worker_thread.start()
        self._append_log("仅转换 3:4 任务已启动。")

    def _run_portrait_only(
        self,
        source_root: Path,
        portrait_root: Path,
        watermark_path: Path | None,
        max_folders: int,
    ) -> None:
        try:
            folders = list_subfolders(source_root)
            if max_folders > 0:
                folders = folders[:max_folders]
            if not folders:
                raise RuntimeError("源目录下没有找到任何货号子文件夹。")
            total_images = 0
            total_folders = 0
            for folder_index, sku_folder in enumerate(folders, start=1):
                images = list_images(sku_folder)
                if not images:
                    self._append_log(f"[{folder_index}/{len(folders)}] 跳过 {sku_folder.name}: 未找到图片")
                    continue
                output_folder = make_relative_output(sku_folder, source_root, portrait_root)
                output_folder.mkdir(parents=True, exist_ok=True)
                self._append_log(f"[{folder_index}/{len(folders)}] 转换货号 {sku_folder.name}: {len(images)} 张")
                for image_index, source_image in enumerate(images, start=1):
                    output_path = output_folder / f"{source_image.stem}_ai_portrait.png"
                    create_portrait_variant(source_image, output_path, watermark_path=watermark_path)
                    total_images += 1
                    self._append_log(f"  [{image_index}/{len(images)}] 已保存 3:4 图: {output_path.name}")
                total_folders += 1
            self._append_log(f"仅转换 3:4 完成：处理货号 {total_folders} 个，图片 {total_images} 张。")
            self.after(0, lambda: messagebox.showinfo("转换完成", f"已处理货号 {total_folders} 个，图片 {total_images} 张。"))
        except Exception as exc:  # noqa: BLE001
            message = humanize_api_error(str(exc))
            self._append_log(f"仅转换 3:4 失败: {message}")
            self.after(0, lambda: messagebox.showerror("转换失败", message))

    def _start_job(self) -> None:
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showinfo("任务运行中", "当前已有任务在运行。")
            return

        try:
            config = self._collect_config()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return

        self._save_config(silent=True)
        self.worker = BatchWorker(config, self._append_log)
        self.worker_thread = threading.Thread(target=self._run_worker, daemon=True)
        self.worker_thread.start()
        self._append_log("任务已启动。")

    def _run_worker(self) -> None:
        try:
            assert self.worker is not None
            self.worker.run()
        except Exception as exc:  # noqa: BLE001
            self._append_log(f"任务失败: {humanize_api_error(str(exc))}")

    def _cancel_job(self) -> None:
        if self.worker:
            self.worker.cancel()
            self._append_log("已请求取消任务。")

    def _collect_scene_config(self) -> SceneJobConfig:
        single_text = self.scene_single_image_var.get().strip()
        single_image = Path(single_text).expanduser() if single_text else None
        source_root = Path(self.scene_source_root_var.get()).expanduser() if self.scene_source_root_var.get().strip() else Path(".")
        output_text = self.scene_output_root_var.get().strip()
        if not output_text:
            raise RuntimeError("请先选择场景图输出目录。")
        if single_image and not single_image.is_file():
            raise RuntimeError("平面原图文件不存在，请重新选择。")
        if not single_image and not source_root.is_dir():
            raise RuntimeError("请选择平面原图，或选择有效的批量源目录。")
        if not self.image_api_key_var.get().strip():
            raise RuntimeError("请先在「素材生成」页填写图片 API Key。")
        if not self.image_base_url_var.get().strip():
            raise RuntimeError("请先在「素材生成」页填写图片接口地址。")
        try:
            scene_count = max(1, min(10, int(self.scene_count_var.get().strip() or "8")))
            max_workers = max(1, int(self.scene_max_workers_var.get().strip() or "2"))
            max_folders = max(0, int(self.scene_max_folders_var.get().strip() or "0"))
        except ValueError as exc:
            raise RuntimeError("场景数量、并发数、最多处理文件夹必须是整数。") from exc
        return SceneJobConfig(
            source_root=source_root,
            output_root=Path(output_text).expanduser(),
            single_image=single_image,
            aspect_ratio=self.scene_aspect_ratio_var.get().strip() or DEFAULT_ASPECT_RATIO,
            scene_count=scene_count,
            scene_prompt_template=self._current_scene_prompt_template(),
            quality=self.quality_var.get().strip(),
            max_workers=max_workers,
            max_folders=max_folders,
        )

    def _start_scene_job(self) -> None:
        if self.scene_thread and self.scene_thread.is_alive():
            messagebox.showinfo("任务运行中", "场景图任务正在运行。")
            return
        if self.local_scene_thread and self.local_scene_thread.is_alive():
            messagebox.showinfo("任务运行中", "请先取消正在运行的本地场景图任务。")
            return
        try:
            config = self._collect_scene_config()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        self._save_config(silent=True)
        self._sync_provider_api_key_from_field()
        provider_id = self._current_image_provider_id()
        client_holder: dict[str, object] = {}

        def client_factory() -> object:
            if "client" not in client_holder:
                client_holder["client"] = create_image_client(
                    provider_id,
                    self.image_api_key_var.get().strip(),
                    self.image_base_url_var.get().strip().rstrip("/"),
                    self.image_model_var.get().strip() or provider_default_model(provider_id),
                )
            return client_holder["client"]

        self.scene_worker = SceneGenerationWorker(config, client_factory, self._append_log)
        self.scene_thread = threading.Thread(target=self._run_scene_worker, daemon=True)
        self.scene_thread.start()
        self._append_log("场景图任务已启动。")

    def _run_scene_worker(self) -> None:
        try:
            assert self.scene_worker is not None
            self.scene_worker.run()
        except Exception as exc:  # noqa: BLE001
            self._append_log(f"场景图任务失败: {humanize_api_error(str(exc))}")

    def _cancel_scene_job(self) -> None:
        if self.scene_worker:
            self.scene_worker.cancel()
            self._append_log("已请求取消场景图任务。")

    def _collect_local_scene_config(self) -> LocalSceneJobConfig:
        single_text = self.scene_single_image_var.get().strip()
        single_image = Path(single_text).expanduser() if single_text else None
        source_root = Path(self.scene_source_root_var.get()).expanduser() if self.scene_source_root_var.get().strip() else Path(".")
        output_text = self.scene_output_root_var.get().strip()
        if not output_text:
            raise RuntimeError("请先选择场景图输出目录。")
        if single_image and not single_image.is_file():
            raise RuntimeError("平面原图文件不存在。")
        if not single_image and not source_root.is_dir():
            raise RuntimeError("请选择平面原图或有效的批量源目录。")
        mockup_text = self.scene_mockup_root_var.get().strip()
        mockup_root = Path(mockup_text).expanduser() if mockup_text else None
        if mockup_root and not mockup_root.is_dir():
            raise RuntimeError("模特底图目录不存在。")
        try:
            scene_count = max(1, min(10, int(self.scene_count_var.get().strip() or "8")))
            max_folders = max(0, int(self.scene_max_folders_var.get().strip() or "0"))
        except ValueError as exc:
            raise RuntimeError("场景数量与最多处理文件夹必须是整数。") from exc
        return LocalSceneJobConfig(
            source_root=source_root,
            output_root=Path(output_text).expanduser(),
            single_image=single_image,
            aspect_ratio=self.scene_aspect_ratio_var.get().strip() or DEFAULT_ASPECT_RATIO,
            scene_count=scene_count,
            mockup_root=mockup_root,
            size_label=self.scene_size_label_var.get().strip() or DEFAULT_SIZE_LABEL,
            max_folders=max_folders,
        )

    def _start_local_scene_job(self) -> None:
        if self.local_scene_thread and self.local_scene_thread.is_alive():
            messagebox.showinfo("任务运行中", "本地场景图任务正在运行。")
            return
        if self.scene_thread and self.scene_thread.is_alive():
            messagebox.showinfo("任务运行中", "请先取消正在运行的 AI 场景图任务。")
            return
        try:
            config = self._collect_local_scene_config()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        self._save_config(silent=True)
        self.local_scene_worker = LocalSceneWorker(config, self._append_log)
        self.local_scene_thread = threading.Thread(target=self._run_local_scene_worker, daemon=True)
        self.local_scene_thread.start()
        self._append_log("本地场景图任务已启动（不消耗 API Key）。")

    def _run_local_scene_worker(self) -> None:
        try:
            assert self.local_scene_worker is not None
            self.local_scene_worker.run()
        except Exception as exc:  # noqa: BLE001
            self._append_log(f"本地场景图失败: {exc}")

    def _cancel_scene_jobs(self) -> None:
        if self.scene_worker:
            self.scene_worker.cancel()
        if self.local_scene_worker:
            self.local_scene_worker.cancel()
        self._append_log("已请求取消场景图任务。")

    def _test_api_connection(self) -> None:
        try:
            config = self._collect_config()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return

        self._append_log("开始测试接口连通性...")
        threading.Thread(target=self._run_api_test, args=(config,), daemon=True).start()

    def _run_api_test(self, config: JobConfig) -> None:
        try:
            if provider_uses_async_tasks(config.image_provider):
                client = create_image_client(
                    config.image_provider,
                    config.image_api_key,
                    config.image_base_url,
                    config.image_model,
                    timeout_seconds=120,
                )
                generated, mode = client.generate_image_with_fallback(
                    prompt="纯白底上放置一枚红色苹果，简洁电商静物摄影，无文字。",
                    output_size="1:1",
                    quality=config.quality,
                    reference_images=[],
                )
                provider_name = provider_label(config.image_provider)
                self._append_log(f"{provider_name} 图片接口测试成功（{mode}），返回 {len(generated)} 字节")
                self.after(
                    0,
                    lambda pn=provider_name, model=config.image_model, size=len(generated): messagebox.showinfo(
                        "测试成功",
                        f"{pn} 接口正常。\n模型：{model}\n测试图大小：{size} 字节",
                    ),
                )
                return
            gateway = GatewayClient(config.image_api_key, config.image_base_url, timeout_seconds=90)
            data = gateway.list_models()
            model_count = len(data.get("data", [])) if isinstance(data, dict) else 0
            self._append_log(f"图片接口测试成功：模型列表 {model_count} 个")
            text_note = ""
            if config.text_api_key:
                try:
                    text_client = TextOnlyClient(
                        config.text_api_key,
                        config.base_url,
                        config.text_model,
                        timeout_seconds=90,
                    )
                    title = text_client.generate_title('请只返回 JSON：{"title":"测试"}')
                    text_note = f"\n文案接口正常，模型 {config.text_model}，示例标题: {title[:40]}"
                    self._append_log(f"文案接口测试成功: {config.text_model}")
                except Exception as text_exc:  # noqa: BLE001
                    text_note = f"\n文案接口测试失败: {humanize_api_error(str(text_exc))}"
                    self._append_log(text_note)
            self.after(
                0,
                lambda note=text_note: messagebox.showinfo(
                    "测试成功",
                    f"图片接口连接正常。\n已获取模型数量：{model_count}{note}",
                ),
            )
        except Exception as exc:  # noqa: BLE001
            friendly = humanize_api_error(str(exc))
            self._append_log(f"接口测试失败: {friendly}")
            self.after(0, lambda: messagebox.showerror("测试失败", friendly))

    def _test_ozon_api(self) -> None:
        try:
            config = self._collect_upload_config(require_batch_files=False, require_oss=False)
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        self._save_config(silent=True)
        self._append_log("开始测试 Ozon API...")
        threading.Thread(target=self._run_ozon_test, args=(config,), daemon=True).start()

    def _run_ozon_test(self, config: BatchUploadConfig) -> None:
        try:
            client = OzonSellerClient(config.ozon_client_id, config.ozon_api_key, timeout_seconds=90)
            client.get_template_product(config.template_offer_id)
            self._append_log(f"Ozon API 测试成功，已读取模板商品: {config.template_offer_id}")
            self.after(0, lambda: messagebox.showinfo("测试成功", "Ozon API 连接正常，模板商品可读取。"))
        except Exception as exc:  # noqa: BLE001
            message = str(exc)
            self._append_log(f"Ozon API 测试失败: {message}")
            self.after(0, lambda: messagebox.showerror("测试失败", message))

    def _test_oss(self) -> None:
        try:
            config = self._collect_upload_config(require_batch_files=False, require_ozon=False)
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        self._save_config(silent=True)
        self._append_log("开始测试 OSS...")
        threading.Thread(target=self._run_oss_test, args=(config,), daemon=True).start()

    def _run_oss_test(self, config: BatchUploadConfig) -> None:
        try:
            client = AliyunOssClient(
                config.oss_access_key_id,
                config.oss_access_key_secret,
                config.oss_bucket,
                config.oss_endpoint,
                config.oss_public_domain,
                timeout_seconds=90,
            )
            url = client.test_upload()
            self._append_log(f"OSS 测试成功: {url}")
            self.after(0, lambda: messagebox.showinfo("测试成功", f"OSS 上传正常。\n{url}"))
        except Exception as exc:  # noqa: BLE001
            message = str(exc)
            self._append_log(f"OSS 测试失败: {message}")
            self.after(0, lambda: messagebox.showerror("测试失败", message))

    def _start_batch_upload(self) -> None:
        if self.upload_thread and self.upload_thread.is_alive():
            messagebox.showinfo("任务运行中", "当前已有批量上架任务在运行。")
            return
        try:
            config = self._collect_upload_config()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        self._save_config(silent=True)
        self.upload_worker = BatchUploadWorker(config, self._append_log)
        self.upload_thread = threading.Thread(target=self._run_batch_upload, daemon=True)
        self.upload_thread.start()
        self._append_log("批量上架任务已启动。")

    def _run_batch_upload(self) -> None:
        try:
            assert self.upload_worker is not None
            result_path = self.upload_worker.run()
            self.after(0, lambda: messagebox.showinfo("批量上架完成", f"结果表已保存:\n{result_path}"))
        except Exception as exc:  # noqa: BLE001
            message = str(exc)
            self._append_log(f"批量上架失败: {message}")
            self.after(0, lambda: messagebox.showerror("批量上架失败", message))

    def _cancel_batch_upload(self) -> None:
        if self.upload_worker:
            self.upload_worker.cancel()
            self._append_log("已请求取消批量上架任务。")

    def _build_upload_tab(self, parent: ttk.Frame) -> None:
        content, canvas = self._create_scrollable_page(parent)
        self._upload_tab_canvas = canvas

        actions = self._section(
            content,
            "快捷操作",
            "新上架用「开始多店铺上传」；已上架商品用下方「按货号更新」同步标题/图/视频/富内容（Ozon /v3/product/import）",
            0,
        )
        ttk.Button(actions, text="新增店铺", command=self._add_shop_card, style="Primary.TButton").pack(side="left")
        ttk.Button(actions, text="按货号更新已上架", command=self._start_listed_product_update, style="Primary.TButton").pack(
            side="left", padx=10
        )
        ttk.Button(actions, text="按模板更新已上架视频", command=self._start_template_video_update).pack(side="left", padx=6)
        ttk.Button(actions, text="开始多店铺上传", command=self._start_batch_upload, style="Primary.TButton").pack(
            side="left", padx=10
        )
        ttk.Button(actions, text="取消全部上传", command=self._cancel_batch_upload, style="Danger.TButton").pack(side="left", padx=10)
        ttk.Button(actions, text="下载 Excel 模板", command=self._download_upload_template).pack(side="left")
        ttk.Button(actions, text="测试 OSS", command=self._test_oss).pack(side="left", padx=10)

        update_opts = self._section(
            content,
            "按货号更新已上架商品",
            "读取各店铺上架 Excel + 3:4 图片目录；以 Ozon 当前商品为底稿提交 /v3/product/import（须传全量信息）",
            1,
        )
        ttk.Checkbutton(update_opts, text="更新标题", variable=self.update_listed_title_var).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(update_opts, text="更新简介", variable=self.update_listed_description_var).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(update_opts, text="更新图片（OSS）", variable=self.update_listed_images_var).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(update_opts, text="更新视频", variable=self.update_listed_video_var).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(update_opts, text="更新 JSON 富内容", variable=self.update_listed_rich_json_var).pack(side="left", padx=(0, 12))
        ttk.Label(
            update_opts,
            text="Excel 列：货号、标题、简介、json富文本内容（可选）；视频可在模板区或 Ozon.Video 页签按货号填写。",
            style="Hint.TLabel",
            wraplength=1040,
        ).pack(side="left", padx=(16, 0))

        templates = self._section(content, "商品模板库 / 查询模板", "通过来源店铺 Ozon API 查询模板，保存后所有店铺都可引用", 2)
        templates.columnconfigure(1, weight=1)
        templates.columnconfigure(3, weight=1)
        ttk.Label(templates, text="来源店铺", style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=8)
        self.template_query_shop_combobox = ttk.Combobox(
            templates, textvariable=self.template_query_source_shop_var, state="readonly", width=30
        )
        self.template_query_shop_combobox.grid(row=0, column=1, sticky="ew", pady=8)
        self._field(templates, 0, 2, "模板商品货号", self.template_query_offer_id_var)
        self._field(templates, 1, 0, "保存模板名称", self.template_query_save_name_var)
        ttk.Button(templates, text="查询模板信息", command=self._query_product_template).grid(
            row=1, column=2, sticky="w", padx=(10, 0), pady=8
        )
        ttk.Button(templates, text="保存为商品模板到 OSS", command=self._save_queried_product_template).grid(
            row=1, column=3, sticky="e", padx=(10, 0), pady=8
        )
        ttk.Label(templates, textvariable=self.template_query_summary_var, style="Hint.TLabel", wraplength=1040).grid(
            row=2, column=0, columnspan=4, sticky="ew", pady=(8, 0)
        )
        ttk.Label(templates, text="模板视频链接", style="Card.TLabel").grid(
            row=3, column=0, sticky="nw", padx=(0, 10), pady=(12, 0)
        )
        self.template_query_video_text = self._text_box(templates, height=5)
        self.template_query_video_text.grid(row=3, column=1, columnspan=3, sticky="ew", pady=(12, 0))
        self.template_query_video_text.bind("<KeyRelease>", lambda _event: self._refresh_template_video_status())
        ttk.Label(
            templates,
            text=f"每行一个视频链接，最多 {MAX_VIDEO_LINKS} 条。更新已上架商品视频时优先使用此处链接；"
            "也可在 Excel 的 Ozon.Video 相关页签按货号填写。查询模板后会自动填充模板中的视频。",
            style="Hint.TLabel",
            wraplength=1040,
        ).grid(row=4, column=0, columnspan=4, sticky="ew", pady=(6, 0))
        ttk.Label(templates, textvariable=self.template_query_video_status_var, style="Hint.TLabel").grid(
            row=5, column=0, columnspan=4, sticky="ew", pady=(4, 0)
        )

        shops = self._section(content, "店铺管理", "每个店铺独立配置 Excel、图片目录、模板和上传数量", 3)
        shops.columnconfigure(0, weight=1)
        self.shops_cards_frame = ttk.Frame(shops, style="CardInner.TFrame")
        self.shops_cards_frame.grid(row=0, column=0, sticky="ew")
        self._refresh_shop_options()
        self._refresh_product_template_options()
        self._bind_page_mousewheel(content, canvas)
        self._upload_tab_mousewheel_ready = True

    def _shop_profile_label(self, shop_id: str, profile: dict) -> str:
        name = str(profile.get("name") or "").strip()
        return f"{name} ({shop_id})" if name else shop_id

    def _template_display_label(self, key: str, template: dict) -> str:
        name = str(template.get("template_name") or key.split("/", 1)[-1])
        source_shop_id = str(template.get("source_shop_id") or template.get("shop_id") or key.split("/", 1)[0])
        source_shop_name = str(template.get("source_shop_name") or template.get("shop_name") or source_shop_id)
        video_count = len(self._template_video_links_for_saved_template(template))
        suffix = f" | 视频 {video_count} 条" if video_count else ""
        return f"{name} - {source_shop_name}/{source_shop_id}{suffix}"

    def _get_template_query_video_links_text(self) -> str:
        if hasattr(self, "template_query_video_text"):
            return self.template_query_video_text.get("1.0", "end").strip()
        return self.template_query_video_links_var.get().strip()

    def _set_template_query_video_links_text(self, text: str) -> None:
        if hasattr(self, "template_query_video_text"):
            self.template_query_video_text.delete("1.0", "end")
            if text:
                self.template_query_video_text.insert("1.0", text)
        self.template_query_video_links_var.set(text)
        self._refresh_template_video_status()

    def _refresh_template_video_status(self) -> None:
        links = parse_video_links_text(self._get_template_query_video_links_text())
        offer_id = self.template_query_offer_id_var.get().strip() or "-"
        self.template_query_video_status_var.set(f"模板视频 {len(links)} 条 | 来源货号: {offer_id}")

    def _template_video_links_for_saved_template(self, template: dict) -> list[str]:
        stored = template.get("template_video_links")
        if isinstance(stored, list):
            links = parse_video_links_text("\n".join(str(item) for item in stored))
            if links:
                return links
        product = template.get("template_product")
        if isinstance(product, dict):
            return extract_video_links_from_product(product)
        return []

    def _refresh_shop_options(self) -> None:
        self._shop_display_to_id = {
            self._shop_profile_label(shop_id, profile): shop_id for shop_id, profile in sorted(self.shop_profiles.items())
        }
        if hasattr(self, "template_query_shop_combobox"):
            values = list(self._shop_display_to_id.keys())
            self.template_query_shop_combobox["values"] = values
            if values and self.template_query_source_shop_var.get() not in values:
                self.template_query_source_shop_var.set(values[0])
        if hasattr(self, "shop_combobox"):
            self.shop_combobox["values"] = list(self._shop_display_to_id.keys())
        self._refresh_inventory_shop_options()
        self._render_shop_cards()

    def _refresh_product_template_options(self) -> None:
        self._product_template_display_to_key = {
            self._template_display_label(key, template): key
            for key, template in sorted(self.product_templates.items())
            if isinstance(template, dict)
        }
        values = list(self._product_template_display_to_key.keys())
        if hasattr(self, "product_template_combobox"):
            self.product_template_combobox["values"] = values
        for variables in self.shop_card_vars.values():
            combo = variables.get("template_combo")
            if isinstance(combo, ttk.Combobox):
                combo["values"] = values

    def _template_key_from_display(self, display: str) -> str:
        return getattr(self, "_product_template_display_to_key", {}).get(display, display)

    def _template_display_from_key(self, key: str) -> str:
        template = self.product_templates.get(key)
        if isinstance(template, dict):
            return self._template_display_label(key, template)
        return ""

    def _source_shop_id_from_display(self) -> str:
        selected = self.template_query_source_shop_var.get().strip()
        return getattr(self, "_shop_display_to_id", {}).get(selected, normalize_ozon_client_id(selected))

    def _render_shop_cards(self) -> None:
        if not hasattr(self, "shops_cards_frame"):
            return
        for child in self.shops_cards_frame.winfo_children():
            child.destroy()
        self.shop_card_vars = {}
        if not self.shop_profiles:
            ttk.Label(
                self.shops_cards_frame,
                text="暂无店铺。点击“新增店铺”创建第一张店铺卡片。",
                style="Hint.TLabel",
            ).grid(row=0, column=0, sticky="w", pady=8)
        else:
            for row, (shop_id, profile) in enumerate(sorted(self.shop_profiles.items())):
                self._build_shop_card(self.shops_cards_frame, row, shop_id, profile)
        self._refresh_product_template_options()
        if getattr(self, "_upload_tab_mousewheel_ready", False) and hasattr(self, "_upload_tab_canvas"):
            self._bind_page_mousewheel(self.shops_cards_frame, self._upload_tab_canvas)

    def _shop_status_tag_style(self, status: str) -> str:
        text = str(status or "").strip()
        if any(token in text for token in ("失败", "错误", "离线", "取消")):
            return "TagDanger.TLabel"
        if any(token in text for token in ("完成", "成功", "在线", "通过")):
            return "TagSuccess.TLabel"
        return "Tag.TLabel"

    def _build_shop_card(self, parent: ttk.Frame, row: int, shop_id: str, profile: dict) -> None:
        card = ttk.Frame(parent, style="ShopCard.TFrame", padding=0)
        card.grid(row=row, column=0, sticky="ew", pady=(0, 16))
        card.columnconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)

        selected_key = str(profile.get("selected_product_template_key") or "")
        variables: dict[str, tk.Variable] = {
            "enabled": tk.BooleanVar(value=bool(profile.get("enabled", False))),
            "name": tk.StringVar(value=str(profile.get("name") or "")),
            "client_id": tk.StringVar(value=shop_id if not shop_id.startswith("new_") else str(profile.get("client_id", ""))),
            "api_key": tk.StringVar(value=str(profile.get("api_key") or "")),
            "portrait_root": tk.StringVar(value=str(profile.get("portrait_root") or self.portrait_root_var.get())),
            "upload_excel_path": tk.StringVar(value=str(profile.get("upload_excel_path") or self.upload_excel_path_var.get())),
            "max_items": tk.StringVar(value=str(profile.get("max_items") or self.upload_max_items_var.get() or "0")),
            "template_offer_id": tk.StringVar(value=str(profile.get("template_offer_id") or self.ozon_template_offer_id_var.get())),
            "selected_product_template_key": tk.StringVar(value=selected_key),
            "selected_product_template_display": tk.StringVar(value=self._template_display_from_key(selected_key)),
            "upload_template_video": tk.BooleanVar(value=bool(profile.get("upload_template_video", False))),
            "status": tk.StringVar(value=str(profile.get("status") or "待上传")),
        }
        self.shop_card_vars[shop_id] = variables

        header = ttk.Frame(card, style="ShopCardHeader.TFrame", padding=(18, 16))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(1, weight=1)

        ttk.Label(header, text="Ozon", style="Badge.TLabel").grid(row=0, column=0, rowspan=2, sticky="nw", padx=(0, 14))
        ttk.Label(header, textvariable=variables["name"], style="ShopName.TLabel").grid(row=0, column=1, sticky="w")
        client_id = str(variables["client_id"].get()).strip() or shop_id
        ttk.Label(header, text=f"店铺 ID：{client_id}", style="ShopMeta.TLabel").grid(row=1, column=1, sticky="w", pady=(4, 0))

        header_actions = ttk.Frame(header, style="ShopCardHeader.TFrame")
        header_actions.grid(row=0, column=2, rowspan=2, sticky="e")
        status_style = self._shop_status_tag_style(str(variables["status"].get()))
        ttk.Label(header_actions, textvariable=variables["status"], style=status_style).pack(side="left", padx=(0, 10))
        ttk.Checkbutton(header_actions, text="参与上传", variable=variables["enabled"]).pack(side="left", padx=(0, 10))
        ttk.Button(header_actions, text="保存", command=lambda sid=shop_id: self._save_shop_card(sid), style="Primary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(header_actions, text="删除", command=lambda sid=shop_id: self._delete_shop_card(sid), style="Danger.TButton").pack(side="left")

        body = ttk.Frame(card, style="ShopCardBody.TFrame", padding=(18, 4, 18, 18))
        body.grid(row=1, column=0, sticky="ew")
        body.columnconfigure(1, weight=1)
        body.columnconfigure(3, weight=1)

        self._field(body, 0, 0, "店铺名称", variables["name"])
        self._field(body, 0, 2, "Client-Id", variables["client_id"])
        self._field(body, 1, 0, "Api-Key", variables["api_key"], show="*")
        self._field(body, 1, 2, "最多上架货号", variables["max_items"], width=12)
        self._path_row(body, 2, "3:4 图片目录", variables["portrait_root"])
        ttk.Label(body, text="上架 Excel", style="Card.TLabel").grid(row=3, column=0, sticky="w", pady=7, padx=(0, 10))
        ttk.Entry(body, textvariable=variables["upload_excel_path"]).grid(row=3, column=1, sticky="ew", pady=7)
        ttk.Button(body, text="选择", command=lambda var=variables["upload_excel_path"]: self._choose_excel(var)).grid(
            row=3, column=2, sticky="e", padx=(10, 0), pady=7
        )
        self._field(body, 4, 0, "模板商品货号", variables["template_offer_id"])
        ttk.Label(body, text="商品模板", style="Card.TLabel").grid(row=4, column=2, sticky="w", padx=(0, 10), pady=8)
        combo = ttk.Combobox(body, textvariable=variables["selected_product_template_display"], state="readonly", width=34)
        combo.grid(row=4, column=3, sticky="ew", pady=8)
        combo.bind("<<ComboboxSelected>>", lambda _event, sid=shop_id: self._on_shop_template_selected(sid))
        variables["template_combo"] = combo

        video_row = ttk.Frame(body, style="ShopCardBody.TFrame")
        video_row.grid(row=5, column=0, columnspan=4, sticky="ew", pady=(6, 0))
        ttk.Checkbutton(
            video_row,
            text="使用模板中的视频",
            variable=variables["upload_template_video"],
            command=lambda sid=shop_id: self._refresh_shop_template_video_label(sid),
        ).pack(side="left")
        variables["template_video_count_label"] = ttk.Label(video_row, text="模板视频 0 条", style="Hint.TLabel")
        variables["template_video_count_label"].pack(side="left", padx=(12, 0))
        self._refresh_shop_template_video_label(shop_id)

        tools = ttk.Frame(body, style="ShopCardBody.TFrame")
        tools.grid(row=6, column=0, columnspan=4, sticky="ew", pady=(10, 0))
        ttk.Button(tools, text="测试 Ozon API", command=lambda sid=shop_id: self._test_shop_ozon_api(sid)).pack(side="left")
        ttk.Button(tools, text="清空模板选择", command=lambda sid=shop_id: self._clear_shop_template(sid)).pack(side="left", padx=10)

    def _add_shop_card(self) -> None:
        shop_id = f"new_{uuid.uuid4().hex[:8]}"
        self.shop_profiles[shop_id] = {
            "name": "新店铺",
            "client_id": "",
            "api_key": "",
            "portrait_root": self.portrait_root_var.get(),
            "upload_excel_path": self.upload_excel_path_var.get(),
            "max_items": self.upload_max_items_var.get() or "0",
            "template_offer_id": self.ozon_template_offer_id_var.get(),
            "selected_product_template_key": "",
            "upload_template_video": False,
            "enabled": True,
        }
        self._refresh_shop_options()

    def _on_shop_template_selected(self, shop_id: str) -> None:
        variables = self.shop_card_vars.get(shop_id)
        if not variables:
            return
        display = str(variables["selected_product_template_display"].get())
        key = self._template_key_from_display(display)
        variables["selected_product_template_key"].set(key)
        template = self.product_templates.get(key)
        if isinstance(template, dict):
            variables["template_offer_id"].set(str(template.get("template_offer_id") or ""))
        self._refresh_shop_template_video_label(shop_id)

    def _refresh_shop_template_video_label(self, shop_id: str) -> None:
        variables = self.shop_card_vars.get(shop_id)
        if not variables:
            return
        label = variables.get("template_video_count_label")
        if not isinstance(label, ttk.Label):
            return
        key = str(variables["selected_product_template_key"].get()).strip()
        template = self.product_templates.get(key)
        count = len(self._template_video_links_for_saved_template(template)) if isinstance(template, dict) else 0
        label.configure(text=f"模板视频 {count} 条")

    def _clear_shop_template(self, shop_id: str) -> None:
        variables = self.shop_card_vars.get(shop_id)
        if not variables:
            return
        variables["selected_product_template_display"].set("")
        variables["selected_product_template_key"].set("")
        variables["status"].set("已清空模板选择")

    def _profile_from_shop_card(self, shop_id: str) -> tuple[str, dict]:
        variables = self.shop_card_vars.get(shop_id)
        if not variables:
            raise RuntimeError("未找到店铺卡片。")
        client_id = normalize_ozon_client_id(str(variables["client_id"].get()))
        profile = {
            "name": str(variables["name"].get()).strip() or client_id or "新店铺",
            "client_id": client_id,
            "api_key": str(variables["api_key"].get()).strip(),
            "portrait_root": str(variables["portrait_root"].get()).strip(),
            "upload_excel_path": str(variables["upload_excel_path"].get()).strip(),
            "max_items": str(variables["max_items"].get()).strip() or "0",
            "template_offer_id": str(variables["template_offer_id"].get()).strip(),
            "selected_product_template_key": str(variables["selected_product_template_key"].get()).strip(),
            "upload_template_video": bool(variables["upload_template_video"].get()),
            "enabled": bool(variables["enabled"].get()),
        }
        return client_id or shop_id, profile

    def _sync_shop_profiles_from_cards(self) -> None:
        if not self.shop_card_vars:
            return
        synced: dict[str, dict] = {}
        for shop_id in list(self.shop_card_vars.keys()):
            new_id, profile = self._profile_from_shop_card(shop_id)
            synced[new_id] = profile
        self.shop_profiles = synced

    def _save_shop_card(self, shop_id: str) -> None:
        try:
            new_id, profile = self._profile_from_shop_card(shop_id)
            if not normalize_ozon_client_id(str(profile.get("client_id", ""))):
                messagebox.showerror("店铺错误", "请填写 Ozon Client-Id。")
                return
            if not str(profile.get("api_key", "")).isascii() or not str(profile.get("api_key", "")).strip():
                messagebox.showerror("店铺错误", "请填写有效的 Ozon Api-Key。")
                return
            if shop_id in self.shop_profiles and new_id != shop_id:
                del self.shop_profiles[shop_id]
            self.shop_profiles[new_id] = profile
            self._save_config(silent=True)
            self._refresh_shop_options()
            self._append_log(f"已保存店铺：{self._shop_profile_label(new_id, profile)}")
        except RuntimeError as exc:
            messagebox.showerror("店铺错误", str(exc))

    def _delete_shop_card(self, shop_id: str) -> None:
        self.shop_profiles.pop(shop_id, None)
        self.shop_card_vars.pop(shop_id, None)
        self._save_config(silent=True)
        self._refresh_shop_options()
        self._append_log(f"已删除店铺：{shop_id}")

    def _collect_upload_config_for_profile(self, shop_id: str, profile: dict, *, require_batch_files: bool = True) -> BatchUploadConfig:
        portrait_root = Path(str(profile.get("portrait_root") or "")).expanduser()
        excel_path = Path(str(profile.get("upload_excel_path") or "")).expanduser()
        if require_batch_files and not portrait_root.is_dir():
            raise RuntimeError("请先选择有效的 3:4 图片目录。")
        if require_batch_files and not excel_path.is_file():
            raise RuntimeError("请先选择有效的上架 Excel 文件。")
        client_id = normalize_ozon_client_id(str(profile.get("client_id") or shop_id))
        api_key = str(profile.get("api_key") or "").strip()
        if not client_id:
            raise RuntimeError("请填写 Ozon Client-Id。")
        if not api_key or not api_key.isascii():
            raise RuntimeError("请填写有效的 Ozon Api-Key。")
        try:
            max_items = max(0, int(str(profile.get("max_items") or "0").strip() or "0"))
        except ValueError as exc:
            raise RuntimeError("最多上架货号必须是整数。") from exc
        template_key = str(profile.get("selected_product_template_key") or "").strip()
        selected_template = self.product_templates.get(template_key)
        template_product = selected_template.get("template_product") if isinstance(selected_template, dict) else None
        template_offer_id = str(profile.get("template_offer_id") or "").strip()
        if isinstance(selected_template, dict):
            template_offer_id = str(selected_template.get("template_offer_id") or template_offer_id)
        if not template_product and not template_offer_id:
            raise RuntimeError("请选择已保存商品模板或填写模板商品货号。")
        upload_template_video = bool(profile.get("upload_template_video", False))
        template_video_links: list[str] | None = None
        if upload_template_video and isinstance(selected_template, dict):
            template_video_links = self._template_video_links_for_saved_template(selected_template)
            if not template_video_links:
                query_links = parse_video_links_text(self._get_template_query_video_links_text())
                if query_links:
                    template_video_links = query_links
        return BatchUploadConfig(
            portrait_root=portrait_root,
            excel_path=excel_path,
            ozon_client_id=client_id,
            ozon_api_key=api_key,
            template_offer_id=template_offer_id,
            oss_access_key_id=self.oss_access_key_id_var.get().strip(),
            oss_access_key_secret=self.oss_access_key_secret_var.get().strip(),
            oss_bucket=self.oss_bucket_var.get().strip(),
            oss_endpoint=self.oss_endpoint_var.get().strip(),
            oss_public_domain=self.oss_public_domain_var.get().strip(),
            max_items=max_items,
            template_product=template_product,
            upload_template_video=upload_template_video,
            template_video_links=template_video_links,
        )

    def _test_shop_ozon_api(self, shop_id: str) -> None:
        try:
            new_id, profile = self._profile_from_shop_card(shop_id)
            config = self._collect_upload_config_for_profile(new_id, profile, require_batch_files=False)
        except RuntimeError as exc:
            messagebox.showerror("测试失败", str(exc))
            return
        variables = self.shop_card_vars.get(shop_id)
        if variables:
            variables["status"].set("正在测试 Ozon API...")

        def worker() -> None:
            try:
                client = OzonSellerClient(config.ozon_client_id, config.ozon_api_key, timeout_seconds=90)
                if config.template_product:
                    client.test_connection()
                else:
                    client.get_template_product(config.template_offer_id)
                self._append_log(f"[{profile.get('name') or config.ozon_client_id}] Ozon API 测试成功")
                if variables:
                    self.after(0, lambda: variables["status"].set("Ozon API 正常"))
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self._append_log(f"[{profile.get('name') or config.ozon_client_id}] Ozon API 测试失败: {message}")
                if variables:
                    self.after(0, lambda: variables["status"].set("Ozon API 测试失败"))
                self.after(0, lambda: messagebox.showerror("测试失败", message))

        threading.Thread(target=worker, daemon=True).start()

    def _query_product_template(self) -> None:
        shop_id = self._source_shop_id_from_display()
        profile = self.shop_profiles.get(shop_id)
        offer_id = self.template_query_offer_id_var.get().strip()
        if not profile:
            messagebox.showerror("模板查询失败", "请先选择来源店铺。")
            return
        if not offer_id:
            messagebox.showerror("模板查询失败", "请填写模板商品货号。")
            return
        self.template_query_summary_var.set("正在查询模板信息...")

        def worker() -> None:
            try:
                client = OzonSellerClient(shop_id, str(profile.get("api_key") or ""), timeout_seconds=120)
                template_product = client.get_template_product(offer_id)
                self.template_query_product = template_product
                self.template_query_shop_id = shop_id
                self.template_query_offer_id = offer_id
                video_links = extract_video_links_from_product(template_product)
                video_text = "\n".join(video_links)
                summary = self._format_template_summary(template_product, shop_id, profile, offer_id, saved=False)
                self.after(0, lambda: self.template_query_summary_var.set(summary))
                self.after(0, lambda text=video_text: self._set_template_query_video_links_text(text))
                self._append_log(f"商品模板查询成功：{offer_id}")
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self.template_query_product = None
                self._append_log(f"商品模板查询失败: {message}")
                self.after(0, lambda: self.template_query_summary_var.set(f"查询失败：{message}"))
                self.after(0, lambda: messagebox.showerror("模板查询失败", message))

        threading.Thread(target=worker, daemon=True).start()

    def _format_template_summary(self, template_product: dict, shop_id: str, profile: dict, offer_id: str, *, saved: bool) -> str:
        images = template_product.get("images")
        image_count = len(images) if isinstance(images, list) else int(bool(template_product.get("primary_image")))
        attributes = template_product.get("attributes")
        attr_count = len(attributes) if isinstance(attributes, list) else 0
        name = str(template_product.get("name") or "")
        category_id = template_product.get("category_id") or template_product.get("description_category_id") or ""
        type_id = template_product.get("type_id") or ""
        price = template_product.get("price") or ""
        video_count = len(extract_video_links_from_product(template_product))
        status = "已保存" if saved else "未保存"
        return (
            f"模板：{name or offer_id} | 来源店铺：{profile.get('name') or shop_id}/{shop_id} | "
            f"货号：{offer_id} | 类目：{category_id} | 类型：{type_id} | 价格：{price} | "
            f"图片数：{image_count} | 属性数：{attr_count} | 视频数：{video_count} | OSS 状态：{status}"
        )

    def _save_queried_product_template(self) -> None:
        if not self.template_query_product:
            messagebox.showerror("保存失败", "请先查询模板信息。")
            return
        template_name = self.template_query_save_name_var.get().strip()
        if not template_name:
            messagebox.showerror("保存失败", "请填写保存模板名称。")
            return
        shop_id = self.template_query_shop_id
        profile = self.shop_profiles.get(shop_id)
        if not profile:
            messagebox.showerror("保存失败", "来源店铺不存在。")
            return
        key = self._template_key(shop_id, template_name)
        if key in self.product_templates:
            messagebox.showinfo("模板已存在", "同一来源店铺下同名模板已存在，保存后不允许覆盖。请换一个模板名称。")
            return

        def worker() -> None:
            try:
                oss_client = AliyunOssClient(
                    self.oss_access_key_id_var.get(),
                    self.oss_access_key_secret_var.get(),
                    self.oss_bucket_var.get(),
                    self.oss_endpoint_var.get(),
                    self.oss_public_domain_var.get(),
                )
                oss_key = self._template_oss_key(shop_id, template_name)
                video_links = parse_video_links_text(self._get_template_query_video_links_text())
                if not video_links and self.template_query_product:
                    video_links = extract_video_links_from_product(self.template_query_product)
                payload = {
                    "source_shop_id": shop_id,
                    "source_shop_name": str(profile.get("name") or ""),
                    "template_name": template_name,
                    "template_offer_id": self.template_query_offer_id,
                    "saved_at": now_stamp(),
                    "oss_key": oss_key,
                    "template_product": self.template_query_product,
                    "template_video_links": video_links,
                    "template_video_count": len(video_links),
                    "template_video_source_offer_id": self.template_query_offer_id,
                }
                payload_bytes = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
                oss_url = oss_client.upload_bytes(payload_bytes, oss_key, "application/json; charset=utf-8")
                payload["oss_url"] = oss_url
                self.product_templates[key] = payload
                self._save_config(silent=True)
                self.after(0, self._refresh_product_template_options)
                summary = self._format_template_summary(self.template_query_product or {}, shop_id, profile, self.template_query_offer_id, saved=True)
                self.after(0, lambda: self.template_query_summary_var.set(f"{summary} | OSS：{oss_url}"))
                self._append_log(f"商品模板已保存到 OSS: {oss_url}")
                self.after(0, lambda: messagebox.showinfo("保存成功", f"商品模板已保存到 OSS:\n{oss_url}"))
            except Exception as exc:  # noqa: BLE001
                message = str(exc)
                self._append_log(f"商品模板保存失败: {message}")
                self.after(0, lambda: messagebox.showerror("保存失败", message))

        threading.Thread(target=worker, daemon=True).start()

    def _start_template_video_update(self) -> None:
        if self._video_update_running:
            messagebox.showinfo("任务运行中", "视频更新任务正在运行。")
            return
        try:
            self._sync_shop_profiles_from_cards()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        global_links = parse_video_links_text(self._get_template_query_video_links_text())
        if not global_links and isinstance(self.template_query_product, dict):
            global_links = extract_video_links_from_product(self.template_query_product)
        selected = [(shop_id, profile) for shop_id, profile in self.shop_profiles.items() if bool(profile.get("enabled"))]
        if not selected:
            messagebox.showerror("配置错误", "请至少勾选一个参与上传的店铺。")
            return
        if not global_links:
            has_excel_video = False
            for _shop_id, profile in selected:
                excel_path = Path(str(profile.get("upload_excel_path") or "")).expanduser()
                if excel_path.is_file() and read_video_links_from_excel(excel_path):
                    has_excel_video = True
                    break
            if not has_excel_video:
                messagebox.showerror("配置错误", "请先填写模板视频链接，或在 Excel 的 Ozon.Video 页签中配置视频。")
                return
        self._video_update_running = True
        self._save_config(silent=True)
        self._append_log(f"开始按模板更新已上架视频，店铺数 {len(selected)}")

        def worker() -> None:
            try:
                for shop_id, profile in selected:
                    self._update_shop_listed_videos(shop_id, profile, global_links)
            finally:
                self._video_update_running = False
                self._append_log("按模板更新已上架视频任务结束")

        threading.Thread(target=worker, daemon=True).start()

    def _start_listed_product_update(self) -> None:
        if self._product_update_running:
            messagebox.showinfo("任务运行中", "按货号更新任务正在运行。")
            return
        if any(thread.is_alive() for thread in self.upload_threads.values()):
            messagebox.showinfo("任务运行中", "请先等待批量上架任务结束。")
            return
        if not any(
            (
                self.update_listed_title_var.get(),
                self.update_listed_description_var.get(),
                self.update_listed_images_var.get(),
                self.update_listed_video_var.get(),
                self.update_listed_rich_json_var.get(),
            )
        ):
            messagebox.showerror("配置错误", "请至少勾选一项要更新的内容。")
            return
        try:
            self._sync_shop_profiles_from_cards()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        selected = [(shop_id, profile) for shop_id, profile in self.shop_profiles.items() if bool(profile.get("enabled"))]
        if not selected:
            messagebox.showerror("配置错误", "请至少勾选一个参与上传的店铺。")
            return
        self._save_config(silent=True)
        self._product_update_running = True
        self._append_log(f"开始按货号更新已上架商品，店铺数 {len(selected)}")

        def worker() -> None:
            try:
                for shop_id, profile in selected:
                    self._update_shop_listed_products(shop_id, profile)
            finally:
                self._product_update_running = False
                self._append_log("按货号更新已上架商品任务结束")

        threading.Thread(target=worker, daemon=True).start()

    def _update_shop_listed_products(self, shop_id: str, profile: dict) -> None:
        label = str(profile.get("name") or shop_id)
        excel_path = Path(str(profile.get("upload_excel_path") or "")).expanduser()
        portrait_root = Path(str(profile.get("portrait_root") or self.portrait_root_var.get())).expanduser()
        if not excel_path.is_file():
            self._append_log(f"[{label}] 跳过：未配置上架 Excel")
            return
        api_key = str(profile.get("api_key") or "").strip()
        client_id = normalize_ozon_client_id(str(profile.get("client_id") or shop_id))
        if not client_id or not api_key:
            self._append_log(f"[{label}] 跳过：缺少 Ozon 凭证")
            return
        try:
            upload_cfg = self._collect_upload_config(require_batch_files=False, require_ozon=True, require_oss=True)
        except RuntimeError as exc:
            self._append_log(f"[{label}] 跳过：{exc}")
            return
        template_key = str(profile.get("selected_product_template_key") or "").strip()
        saved_template = self.product_templates.get(template_key)
        template_product = saved_template.get("template_product") if isinstance(saved_template, dict) else None
        video_links = parse_video_links_text(self._get_template_query_video_links_text())
        if not video_links and isinstance(saved_template, dict):
            video_links = self._template_video_links_for_saved_template(saved_template)
        try:
            max_items = max(0, int(str(profile.get("max_items") or "0").strip() or "0"))
        except ValueError:
            max_items = 0
        config = ListedProductUpdateConfig(
            portrait_root=portrait_root,
            excel_path=excel_path,
            ozon_client_id=client_id,
            ozon_api_key=api_key,
            oss_access_key_id=upload_cfg.oss_access_key_id,
            oss_access_key_secret=upload_cfg.oss_access_key_secret,
            oss_bucket=upload_cfg.oss_bucket,
            oss_endpoint=upload_cfg.oss_endpoint,
            oss_public_domain=upload_cfg.oss_public_domain,
            max_items=max_items,
            update_title=self.update_listed_title_var.get(),
            update_description=self.update_listed_description_var.get(),
            update_images=self.update_listed_images_var.get(),
            update_video=self.update_listed_video_var.get(),
            update_rich_json=self.update_listed_rich_json_var.get(),
            template_video_links=video_links,
            template_product=template_product,
        )
        worker = ListedProductUpdateWorker(
            config,
            logger=lambda message, shop_label=label: self._append_log(f"[{shop_label}] {message}"),
        )
        try:
            result_path = worker.run()
            self._append_log(f"[{label}] 更新完成，结果表: {result_path}")
        except Exception as exc:  # noqa: BLE001
            self._append_log(f"[{label}] 更新失败: {exc}")

    def _update_shop_listed_videos(self, shop_id: str, profile: dict, global_links: list[str]) -> None:
        label = str(profile.get("name") or shop_id)
        excel_path = Path(str(profile.get("upload_excel_path") or "")).expanduser()
        if not excel_path.is_file():
            self._append_log(f"[{label}] 跳过：未配置上架 Excel")
            return
        api_key = str(profile.get("api_key") or "").strip()
        client_id = normalize_ozon_client_id(str(profile.get("client_id") or shop_id))
        if not client_id or not api_key:
            self._append_log(f"[{label}] 跳过：缺少 Ozon 凭证")
            return
        try:
            client = OzonSellerClient(client_id, api_key, timeout_seconds=120)
            template_key = str(profile.get("selected_product_template_key") or "").strip()
            saved_template = self.product_templates.get(template_key)
            template_product = saved_template.get("template_product") if isinstance(saved_template, dict) else None
            template_offer_id = str(profile.get("template_offer_id") or "").strip()
            if isinstance(saved_template, dict):
                template_offer_id = str(saved_template.get("template_offer_id") or template_offer_id)
            if not template_product and isinstance(self.template_query_product, dict):
                template_product = self.template_query_product
                template_offer_id = str(self.template_query_offer_id or template_offer_id)
            template_product = resolve_template_product_for_video(
                client,
                template_product=template_product,
                template_offer_id=template_offer_id,
            )
            shop_links = list(global_links)
            if not shop_links and isinstance(saved_template, dict):
                shop_links = self._template_video_links_for_saved_template(saved_template)
            per_offer_links = read_video_links_from_excel(excel_path)
            offer_ids = list(read_excel_rows(excel_path).keys())
            if not offer_ids:
                self._append_log(f"[{label}] 跳过：Excel 中无货号")
                return
            self._append_log(f"[{label}] 准备更新 {len(offer_ids)} 个货号的视频")
            success, failed = update_listed_products_videos(
                client,
                offer_ids,
                shop_links,
                template_product,
                per_offer_links=per_offer_links,
                logger=lambda message, shop_label=label: self._append_log(f"[{shop_label}] {message}"),
            )
            self._append_log(f"[{label}] 视频更新完成：成功 {success}，失败 {failed}")
        except Exception as exc:  # noqa: BLE001
            self._append_log(f"[{label}] 视频更新失败: {exc}")

    def _start_batch_upload(self) -> None:
        if any(thread.is_alive() for thread in self.upload_threads.values()):
            messagebox.showinfo("任务运行中", "当前已有店铺上传任务在运行。")
            return
        try:
            self._sync_shop_profiles_from_cards()
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        selected = [(shop_id, profile) for shop_id, profile in self.shop_profiles.items() if bool(profile.get("enabled"))]
        if not selected:
            messagebox.showerror("配置错误", "请至少勾选一个店铺。")
            return
        configs: list[tuple[str, dict, BatchUploadConfig]] = []
        try:
            for shop_id, profile in selected:
                configs.append((shop_id, profile, self._collect_upload_config_for_profile(shop_id, profile)))
        except RuntimeError as exc:
            messagebox.showerror("配置错误", str(exc))
            return
        self._save_config(silent=True)
        self.upload_workers = {}
        self.upload_threads = {}
        for shop_id, profile, config in configs:
            label = profile.get("name") or shop_id
            worker = BatchUploadWorker(config, lambda message, label=label: self._append_log(f"[{label}] {message}"))
            thread = threading.Thread(target=self._run_shop_batch_upload, args=(shop_id, profile, worker), daemon=True)
            self.upload_workers[shop_id] = worker
            self.upload_threads[shop_id] = thread
            variables = self.shop_card_vars.get(shop_id)
            if variables:
                variables["status"].set("上传中")
            thread.start()
        self._append_log(f"多店铺上传已启动：{len(configs)} 个店铺")

    def _run_shop_batch_upload(self, shop_id: str, profile: dict, worker: BatchUploadWorker) -> None:
        label = str(profile.get("name") or shop_id)
        try:
            result_path = worker.run()
            self._append_log(f"[{label}] 上传完成，结果表：{result_path}")
            variables = self.shop_card_vars.get(shop_id)
            if variables:
                self.after(0, lambda: variables["status"].set("上传完成"))
        except Exception as exc:  # noqa: BLE001
            message = str(exc)
            self._append_log(f"[{label}] 上传失败: {message}")
            variables = self.shop_card_vars.get(shop_id)
            if variables:
                self.after(0, lambda: variables["status"].set("上传失败"))
            self.after(0, lambda: messagebox.showerror("店铺上传失败", f"{label}\n{message}"))

    def _cancel_batch_upload(self) -> None:
        if not self.upload_workers:
            self._append_log("当前没有运行中的店铺上传任务。")
            return
        for worker in self.upload_workers.values():
            worker.cancel()
        self._append_log("已请求取消全部店铺上传任务。")

    def _save_config(self, silent: bool = False) -> None:
        self._sync_provider_api_key_from_field()
        self._sync_text_provider_api_key_from_field()
        if hasattr(self, "shop_card_vars") and self.shop_card_vars:
            try:
                self._sync_shop_profiles_from_cards()
            except RuntimeError:
                pass
        payload = {
            "source_root": self.source_root_var.get(),
            "generated_root": self.generated_root_var.get(),
            "portrait_root": self.portrait_root_var.get(),
            "content_root": self.content_root_var.get(),
            "watermark_path": self.watermark_path_var.get(),
            "image_api_key": self.image_api_key_var.get(),
            "text_api_key": self.text_api_key_var.get(),
            "base_url": self.base_url_var.get(),
            "image_base_url": self.image_base_url_var.get(),
            "image_provider": self._current_image_provider_id(),
            "provider_api_keys": dict(self.provider_api_keys),
            "text_provider": self._current_text_provider_id(),
            "text_provider_api_keys": dict(self.text_provider_api_keys),
            "image_model": self.image_model_var.get(),
            "text_model": self.text_model_var.get(),
            "quality": self.quality_var.get(),
            "max_folders": self.max_folders_var.get(),
            "max_workers": self.max_workers_var.get(),
            "convert_originals": self.convert_originals_var.get(),
            "generate_copy": self.generate_copy_var.get(),
            "export_excel": self.export_excel_var.get(),
            "ozon_client_id": self.ozon_client_id_var.get(),
            "ozon_api_key": self.ozon_api_key_var.get(),
            "ozon_template_offer_id": self.ozon_template_offer_id_var.get(),
            "upload_excel_path": self.upload_excel_path_var.get(),
            "upload_max_items": self.upload_max_items_var.get(),
            "selected_template_name": self.template_name_var.get().strip() or DEFAULT_TEMPLATE_NAME,
            "prompt_templates": self.prompt_templates,
            "shop_profiles": self.shop_profiles,
            "inventory_warehouses_by_shop": self.inventory_warehouses_by_shop,
            "product_templates": self.product_templates,
            "template_query_source_shop": self.template_query_source_shop_var.get(),
            "template_query_offer_id": self.template_query_offer_id_var.get(),
            "template_query_save_name": self.template_query_save_name_var.get(),
            "template_query_video_links": self._get_template_query_video_links_text(),
            "scene_source_root": self.scene_source_root_var.get(),
            "scene_single_image": self.scene_single_image_var.get(),
            "scene_output_root": self.scene_output_root_var.get(),
            "scene_aspect_ratio": self.scene_aspect_ratio_var.get(),
            "scene_count": self.scene_count_var.get(),
            "scene_max_workers": self.scene_max_workers_var.get(),
            "scene_max_folders": self.scene_max_folders_var.get(),
            "scene_mockup_root": self.scene_mockup_root_var.get(),
            "scene_size_label": self.scene_size_label_var.get(),
            "scene_prompt_template": self._current_scene_prompt_template(),
            **self._current_prompt_bundle(),
        }
        write_config_payload(payload)
        if not silent:
            self._append_log(f"配置已保存到 {CONFIG_PATH}")

    def _load_config(self) -> None:
        if not CONFIG_PATH.exists():
            self._refresh_shop_options()
            return
        try:
            data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            self._append_log("配置文件读取失败，已忽略旧配置。")
            return

        self.source_root_var.set(data.get("source_root", ""))
        self.generated_root_var.set(data.get("generated_root", ""))
        self.portrait_root_var.set(data.get("portrait_root", ""))
        self.content_root_var.set(data.get("content_root", ""))
        self.watermark_path_var.set(data.get("watermark_path", ""))
        self.provider_api_keys = migrate_provider_api_keys(data)
        saved_provider = str(data.get("image_provider") or DEFAULT_IMAGE_PROVIDER).strip()
        if saved_provider not in IMAGE_PROVIDER_PROFILES:
            saved_provider = DEFAULT_IMAGE_PROVIDER
        self.image_provider_var.set(provider_label(saved_provider))
        active_key = str(data.get("image_api_key") or "").strip()
        if active_key:
            self.provider_api_keys[saved_provider] = active_key
        elif saved_provider not in self.provider_api_keys:
            env_key = os.environ.get("BREAKOUT_IMAGE_API_KEY", os.environ.get("BREAKOUT_API_KEY", ""))
            if env_key:
                self.provider_api_keys[saved_provider] = env_key
        self.image_api_key_var.set(self.provider_api_keys.get(saved_provider, active_key))
        self.text_provider_api_keys = migrate_text_provider_api_keys(data)
        saved_text_provider = str(data.get("text_provider") or DEFAULT_TEXT_PROVIDER).strip()
        if saved_text_provider not in TEXT_PROVIDER_PROFILES:
            saved_text_provider = DEFAULT_TEXT_PROVIDER
        self.text_provider_var.set(text_provider_label(saved_text_provider))
        active_text_key = str(data.get("text_api_key") or "").strip()
        if active_text_key:
            self.text_provider_api_keys[saved_text_provider] = active_text_key
        elif saved_text_provider not in self.text_provider_api_keys:
            env_text_key = os.environ.get("BREAKOUT_TEXT_API_KEY", os.environ.get("BREAKOUT_API_KEY", ""))
            if env_text_key:
                self.text_provider_api_keys[saved_text_provider] = env_text_key
        self.text_api_key_var.set(self.text_provider_api_keys.get(saved_text_provider, active_text_key))
        self.base_url_var.set(data.get("base_url", text_provider_default_base_url(saved_text_provider)))
        legacy_image_url = data.get("image_base_url") or data.get("base_url")
        self.image_base_url_var.set(
            legacy_image_url if legacy_image_url else provider_default_base_url(saved_provider)
        )
        saved_image_model = (data.get("image_model") or data.get("model") or "").strip()
        if not saved_image_model:
            saved_image_model = provider_default_model(saved_provider)
        elif saved_image_model in {"gpt-image-1.5", "gpt-image-1"} and saved_provider == "wenwen":
            saved_image_model = provider_default_model(saved_provider)
        self.image_model_var.set(saved_image_model)
        if hasattr(self, "image_provider_combobox"):
            self._apply_image_provider_ui(saved_provider, initial=True)
        saved_text_model = (data.get("text_model") or DEFAULT_TEXT_MODEL).strip()
        if not saved_text_model or saved_text_model not in text_provider_models(saved_text_provider):
            saved_text_model = text_provider_default_model(saved_text_provider)
        self.text_model_var.set(saved_text_model)
        if hasattr(self, "text_provider_combobox"):
            self._apply_text_provider_ui(saved_text_provider, initial=True)
        self.quality_var.set(data.get("quality", "medium"))
        self.max_folders_var.set(str(data.get("max_folders", "0")))
        self.max_workers_var.set(str(data.get("max_workers", "3")))
        self.scene_source_root_var.set(data.get("scene_source_root", ""))
        self.scene_single_image_var.set(data.get("scene_single_image", ""))
        self.scene_output_root_var.set(data.get("scene_output_root", ""))
        self.scene_aspect_ratio_var.set(data.get("scene_aspect_ratio", DEFAULT_ASPECT_RATIO))
        self.scene_count_var.set(str(data.get("scene_count", "8")))
        self.scene_max_workers_var.set(str(data.get("scene_max_workers", data.get("max_workers", "2"))))
        self.scene_max_folders_var.set(str(data.get("scene_max_folders", "0")))
        self.scene_mockup_root_var.set(data.get("scene_mockup_root", ""))
        self.scene_size_label_var.set(data.get("scene_size_label", DEFAULT_SIZE_LABEL))
        scene_prompt = str(data.get("scene_prompt_template") or "").strip()
        if scene_prompt:
            self._scene_prompt_template = scene_prompt
            if hasattr(self, "scene_prompt_text"):
                self.scene_prompt_text.delete("1.0", "end")
                self.scene_prompt_text.insert("1.0", scene_prompt)
        self.convert_originals_var.set(bool(data.get("convert_originals", False)))
        self.generate_copy_var.set(bool(data.get("generate_copy", True)))
        self.export_excel_var.set(bool(data.get("export_excel", True)))
        self.ozon_client_id_var.set(data.get("ozon_client_id", ""))
        self.ozon_api_key_var.set(data.get("ozon_api_key", ""))
        self.ozon_template_offer_id_var.set(data.get("ozon_template_offer_id", ""))
        self.upload_excel_path_var.set(data.get("upload_excel_path", ""))
        self.oss_access_key_id_var.set(data.get("oss_access_key_id", ""))
        self.oss_access_key_secret_var.set(data.get("oss_access_key_secret", ""))
        self.oss_bucket_var.set(data.get("oss_bucket", DEFAULT_OSS_BUCKET))
        self.oss_endpoint_var.set(data.get("oss_endpoint", DEFAULT_OSS_ENDPOINT))
        self.oss_public_domain_var.set(data.get("oss_public_domain", DEFAULT_OSS_PUBLIC_DOMAIN))
        self.upload_max_items_var.set(str(data.get("upload_max_items", "0")))

        saved_shops = data.get("shop_profiles")
        cleaned_shops: dict[str, dict] = {}
        if isinstance(saved_shops, dict):
            for shop_id, profile in saved_shops.items():
                if not isinstance(profile, dict):
                    continue
                clean_id = normalize_ozon_client_id(str(profile.get("client_id") or shop_id))
                if not clean_id and str(shop_id).startswith("new_"):
                    clean_id = str(shop_id)
                if clean_id:
                    cleaned_shops[clean_id] = {
                        "name": str(profile.get("name", "")),
                        "client_id": clean_id if not clean_id.startswith("new_") else str(profile.get("client_id", "")),
                        "api_key": str(profile.get("api_key", "")),
                        "portrait_root": str(profile.get("portrait_root", data.get("portrait_root", ""))),
                        "upload_excel_path": str(profile.get("upload_excel_path", data.get("upload_excel_path", ""))),
                        "max_items": str(profile.get("max_items", data.get("upload_max_items", "0"))),
                        "template_offer_id": str(profile.get("template_offer_id", data.get("ozon_template_offer_id", ""))),
                        "selected_product_template_key": str(
                            profile.get("selected_product_template_key", profile.get("selected_product_template", ""))
                        ),
                        "upload_template_video": bool(profile.get("upload_template_video", False)),
                        "enabled": bool(profile.get("enabled", False)),
                    }
        legacy_client_id = normalize_ozon_client_id(str(data.get("ozon_client_id", "")))
        if legacy_client_id and legacy_client_id not in cleaned_shops:
            cleaned_shops[legacy_client_id] = {
                "name": str(data.get("shop_name", "")) or legacy_client_id,
                "client_id": legacy_client_id,
                "api_key": str(data.get("ozon_api_key", "")),
                "portrait_root": str(data.get("portrait_root", "")),
                "upload_excel_path": str(data.get("upload_excel_path", "")),
                "max_items": str(data.get("upload_max_items", "0")),
                "template_offer_id": str(data.get("ozon_template_offer_id", "")),
                "selected_product_template_key": str(data.get("selected_product_template", "")),
                "enabled": False,
            }
        self.shop_profiles = cleaned_shops

        saved_inventory_warehouses = data.get("inventory_warehouses_by_shop")
        cleaned_inventory_warehouses: dict[str, dict] = {}
        if isinstance(saved_inventory_warehouses, dict):
            for shop_id, entry in saved_inventory_warehouses.items():
                if not isinstance(entry, dict):
                    continue
                clean_id = normalize_ozon_client_id(str(shop_id))
                if not clean_id:
                    continue
                warehouses = deserialize_warehouse_options(entry.get("warehouses") or [])
                if not warehouses:
                    continue
                selected = entry.get("selected_warehouse_id")
                selected_id = int(selected) if selected is not None and str(selected).strip() else None
                cleaned_inventory_warehouses[clean_id] = {
                    "warehouses": serialize_warehouse_options(warehouses),
                    "selected_warehouse_id": selected_id,
                }
        self.inventory_warehouses_by_shop = cleaned_inventory_warehouses

        saved_product_templates = data.get("product_templates")
        if isinstance(saved_product_templates, dict):
            cleaned_templates: dict[str, dict] = {}
            for key, template in saved_product_templates.items():
                if not isinstance(template, dict):
                    continue
                source_shop_id = str(template.get("source_shop_id") or template.get("shop_id") or str(key).split("/", 1)[0])
                template_name = str(template.get("template_name") or str(key).split("/", 1)[-1])
                normalized_key = self._template_key(source_shop_id, template_name)
                template["source_shop_id"] = source_shop_id
                template["source_shop_name"] = str(template.get("source_shop_name") or template.get("shop_name") or "")
                template["template_name"] = template_name
                cleaned_templates[normalized_key] = template
            self.product_templates = cleaned_templates

        self.template_query_source_shop_var.set(str(data.get("template_query_source_shop", "")))
        self.template_query_offer_id_var.set(str(data.get("template_query_offer_id", "")))
        self.template_query_save_name_var.set(str(data.get("template_query_save_name", "")))
        if hasattr(self, "template_query_video_text"):
            self._set_template_query_video_links_text(str(data.get("template_query_video_links", "")))
        else:
            self.template_query_video_links_var.set(str(data.get("template_query_video_links", "")))

        saved_templates = data.get("prompt_templates")
        if isinstance(saved_templates, dict):
            cleaned_prompt_templates: dict[str, dict[str, str]] = {}
            for name, bundle in saved_templates.items():
                if isinstance(name, str) and isinstance(bundle, dict):
                    cleaned_prompt_templates[name] = {
                        "image_prompt_template": str(bundle.get("image_prompt_template", DEFAULT_IMAGE_PROMPT)),
                        "title_prompt_template": str(bundle.get("title_prompt_template", DEFAULT_TITLE_PROMPT)),
                        "description_prompt_template": str(bundle.get("description_prompt_template", DEFAULT_DESCRIPTION_PROMPT)),
                    }
            if cleaned_prompt_templates:
                self.prompt_templates = cleaned_prompt_templates
        if DEFAULT_TEMPLATE_NAME not in self.prompt_templates:
            self.prompt_templates[DEFAULT_TEMPLATE_NAME] = {
                "image_prompt_template": DEFAULT_IMAGE_PROMPT,
                "title_prompt_template": DEFAULT_TITLE_PROMPT,
                "description_prompt_template": DEFAULT_DESCRIPTION_PROMPT,
            }
        legacy_bundle = {
            "image_prompt_template": data.get("image_prompt_template", DEFAULT_IMAGE_PROMPT),
            "title_prompt_template": data.get("title_prompt_template", DEFAULT_TITLE_PROMPT),
            "description_prompt_template": data.get("description_prompt_template", DEFAULT_DESCRIPTION_PROMPT),
        }
        selected_name = str(data.get("selected_template_name", DEFAULT_TEMPLATE_NAME)).strip() or DEFAULT_TEMPLATE_NAME
        if selected_name not in self.prompt_templates:
            self.prompt_templates[selected_name] = {
                "image_prompt_template": str(legacy_bundle["image_prompt_template"]),
                "title_prompt_template": str(legacy_bundle["title_prompt_template"]),
                "description_prompt_template": str(legacy_bundle["description_prompt_template"]),
            }
        self.template_name_var.set(selected_name)
        self._refresh_template_options()
        self._apply_prompt_bundle(self.prompt_templates[selected_name])
        self._refresh_product_template_options()
        self._refresh_shop_options()
        if hasattr(self, "inventory_warehouse_combobox"):
            self._apply_cached_inventory_warehouses_for_current_shop()


def main() -> None:
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
